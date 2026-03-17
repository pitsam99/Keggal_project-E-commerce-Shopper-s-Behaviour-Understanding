"""
╔══════════════════════════════════════════════════════════════════════╗
║  LineageIQ — WORKFLOW PARSER  v4                                     ║
║  Script : lineageiq_workflow.py                                      ║
║                                                                      ║
║  Maplet architecture correctly handled (same as mapping parser):     ║
║   • Global MAPPLET definitions at FOLDER level → registry           ║
║   • Inline TRANSFORMATION TYPE="Maplet" inside MAPPING              ║
║   • INSTANCE TRANSFORMATION_TYPE="Maplet" → registry lookup         ║
║   • CONNECTOR traces through maplet input/output ports               ║
║   • Nested maplets → recursive resolution with cycle detection       ║
║   • Session PIPELINE/STAGE used for exec order (PC 10.5.x)          ║
║                                                                      ║
║  USAGE                                                               ║
║    python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║    python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║                                  --par params_prod.par               ║
║    python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║                                  --par params_prod.par               ║
║                                  --out MyOutput.xlsx                 ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict
from typing import Optional

import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Re-use all data models and modules from the mapping parser ────────
# Rather than duplicate code, we import from lineageiq_mapping.
# If running standalone, the mapping parser must be in the same folder.
import importlib.util, sys, os

_mp_path = Path(__file__).parent / "lineageiq_mapping.py"
_spec    = importlib.util.spec_from_file_location("lineageiq_mapping",
                                                    str(_mp_path))
_mod     = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

# Pull everything we need from the mapping module
ParamResolver    = _mod.ParamResolver
MappingParser    = _mod.MappingParser
LineageGraph     = _mod.LineageGraph
MappingExcelWriter = _mod.MappingExcelWriter

# Re-use dataclasses from mapping module
ColumnDef        = _mod.ColumnDef
TransformPort    = _mod.TransformPort
MapletPort       = _mod.MapletPort
Transformation   = _mod.Transformation
Connector        = _mod.Connector
Mapping          = _mod.Mapping


# ══════════════════════════════════════════════════════════════════════
# WORKFLOW-SPECIFIC DATA MODELS
# ══════════════════════════════════════════════════════════════════════

@dataclass
class SessionTransformInst:
    instance_name:       str
    transformation_name: str
    transformation_type: str
    pipeline:            int
    stage:               int
    is_repartition:      bool
    partition_type:      str


@dataclass
class Session:
    name:             str
    mapping_name:     str
    description:      str  = ""
    src_connection:   str  = ""
    tgt_connection:   str  = ""
    is_reusable:      bool = False
    sess_transforms:  list = field(default_factory=list)


@dataclass
class WorkflowTask:
    name:      str
    task_type: str
    task_ref:  str
    condition: str = ""


@dataclass
class Workflow:
    name:            str
    description:     str  = ""
    server:          str  = ""
    tasks:           list = field(default_factory=list)
    execution_order: list = field(default_factory=list)
    worklets:        dict = field(default_factory=dict)
    sessions:        list = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════
# WORKFLOW XML PARSER
# ══════════════════════════════════════════════════════════════════════

class WorkflowParser:
    """
    Parses a complete Informatica workflow XML.
    Delegates all mapping/maplet parsing to MappingParser
    (which handles the full maplet architecture).
    Adds workflow/worklet/session layer on top.
    """

    def __init__(self, resolver: ParamResolver):
        self.resolver            = resolver
        self.mapping_parser      = MappingParser(resolver)
        self.workflows:  dict    = {}
        self.sessions:   dict    = {}

    # Expose mapping_parser's data for convenience
    @property
    def mappings(self):
        return self.mapping_parser.mappings

    @property
    def maplets(self):
        return self.mapping_parser.maplets

    def parse(self, xml_path: str) -> None:
        print(f"\n[PARSE] {xml_path}")
        xp = ET.XMLParser(load_dtd=False, no_network=True,
                           resolve_entities=False, recover=True)
        tree  = ET.parse(xml_path, xp)
        root  = tree.getroot()
        folder = (root.find(".//FOLDER")
                  if root.find(".//FOLDER") is not None
                  else root)

        # Let mapping parser handle sources, targets, maplets, mappings
        self.mapping_parser.parse(xml_path)

        # Now handle workflow-specific elements
        self._parse_sessions(folder)
        self._parse_workflows(folder)

        print(f"\n  [SUMMARY]")
        print(f"    Workflows : {len(self.workflows)}")
        print(f"    Sessions  : {len(self.sessions)}")
        print(f"    Mappings  : {len(self.mappings)}")
        print(f"    Maplets   : {len(self.maplets)}")
        maplet_users = sum(
            1 for m in self.mappings.values()
            if m.maplet_instances)
        print(f"    Mappings using maplets: {maplet_users}")

    # ── Sessions ──────────────────────────────────────────────────

    def _parse_sessions(self, folder) -> None:
        for sess in folder.findall(".//SESSION"):
            # PC 10.x: MAPPINGNAME as direct attribute
            mapping_name = sess.get("MAPPINGNAME","")
            # PC 9.x fallback
            if not mapping_name:
                for a in sess.findall("ATTRIBUTE"):
                    if "Mapping name" in a.get("NAME",""):
                        mapping_name = self.resolver.resolve(
                            a.get("VALUE",""))
                        break

            src_conn = tgt_conn = ""
            sti_list = []

            for sti in sess.findall("SESSTRANSFORMATIONINST"):
                ttype    = sti.get("TRANSFORMATIONTYPE","")
                sname    = sti.get("SINSTANCENAME","")
                pipeline = int(sti.get("PIPELINE","0"))
                stage    = int(sti.get("STAGE","0"))

                for attr in sti.findall("ATTRIBUTE"):
                    aname = attr.get("NAME","")
                    aval  = self.resolver.resolve(
                        attr.get("VALUE",""))
                    if (aname == "Table Name Prefix"
                            and ttype == "Target Definition"):
                        tgt_conn = aval
                    if aname == "$Source connection value":
                        src_conn = aval
                    if (aname == "$Target connection value"
                            and not tgt_conn):
                        tgt_conn = aval

                if ttype == "Source Definition" and not src_conn:
                    src_conn = sname

                sti_list.append(SessionTransformInst(
                    sname,
                    sti.get("TRANSFORMATIONNAME", sname),
                    ttype, pipeline, stage,
                    sti.get("ISREPARTITIONPOINT","NO") == "YES",
                    sti.get("PARTITIONTYPE","")))

            s = Session(
                name             = sess.get("NAME",""),
                mapping_name     = mapping_name,
                description      = sess.get("DESCRIPTION",""),
                src_connection   = src_conn,
                tgt_connection   = tgt_conn,
                is_reusable      = sess.get("REUSABLE","NO")=="YES",
                sess_transforms  = sti_list,
            )
            self.sessions[s.name] = s
            print(f"  [SESS] {s.name}  →  mapping:{s.mapping_name}  "
                  f"pipelines:{sorted(set(st.pipeline for st in sti_list))}")

    # ── Workflows + Worklets ──────────────────────────────────────

    def _parse_workflows(self, folder) -> None:
        worklet_defs: dict = {}
        for wkl in folder.findall(".//WORKLET"):
            wname = wkl.get("NAME","")
            tasks = [(ti.get("NAME",""),
                      ti.get("TASKTYPE",""),
                      ti.get("TASKNAME",""))
                     for ti in wkl.findall("TASKINSTANCE")]
            worklet_defs[wname] = tasks

        for wf in folder.findall(".//WORKFLOW"):
            w = Workflow(
                name        = wf.get("NAME",""),
                description = wf.get("DESCRIPTION",""),
                server      = wf.get("SERVERNAME",""),
            )
            for ti in wf.findall("TASKINSTANCE"):
                w.tasks.append(WorkflowTask(
                    name      = ti.get("NAME",""),
                    task_type = ti.get("TASKTYPE",""),
                    task_ref  = ti.get("TASKNAME",""),
                ))

            # Build execution order from LINK chain
            links: dict = {}
            for lnk in wf.findall("LINK"):
                links[lnk.get("FROMTASK","")] = (
                    lnk.get("TOTASK",""),
                    lnk.get("CONDITION",""))
            cur = "start"; visited = set()
            while cur in links and cur not in visited:
                visited.add(cur)
                nxt, cond = links[cur]
                if nxt and nxt != "start":
                    w.execution_order.append((nxt, cond))
                cur = nxt

            # Flatten sessions in order
            for task in w.tasks:
                if (task.task_type == "Worklet"
                        and task.task_ref in worklet_defs):
                    for (tname, ttype, tref) in \
                            worklet_defs[task.task_ref]:
                        if ttype == "Session":
                            w.sessions.append(tname)
                elif task.task_type == "Session":
                    w.sessions.append(task.name)

            w.worklets = {k: [t[0] for t in v]
                          for k, v in worklet_defs.items()}
            self.workflows[w.name] = w
            print(f"  [WF  ] {w.name}  sessions={w.sessions}")


# ══════════════════════════════════════════════════════════════════════
# WORKFLOW EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════

class WorkflowExcelWriter:
    """
    Generates the workflow-level STTM Excel.
    Sheets:
      1_WorkflowSummary    — WF → Worklet → Session → Mapping → Maplets
      2_SessionDetail      — PIPELINE / STAGE for each session
      3_ExecutionSequence  — Ordered list of sessions in WF
      Per mapping (one set per mapping):
        <M>_A_MappingParse — Full 5-section parse (delegates to mapping writer)
        <M>_B_Sources
        <M>_C_Targets
        <M>_D_Transforms   — With maplet expansion
        <M>_E_ColFlow      — With maplet traces
        <M>_F_Lookups
        <M>_G_MapletDetail — Full maplet breakdown
        <M>_H_ExecOrder
    """

    C = {
        "navy":   "0D1B2A","teal":  "065A82","purple":"4A235A",
        "dark":   "2C3E50","brown": "7B341E","maplet":"7B2D8B",
        "sub":    "1A3A5C","gold":  "F4A621","green": "22C55E",
        "cyan":   "00B4D8","coral": "E8593C","white": "FFFFFF",
        "alt_a":  "E8F4FD","alt_b": "E8FDF4","alt_d": "FEF9E7",
        "alt_maplet": "F9EAF9",
    }
    BORD  = Border(**{s: Side(style="thin", color="CCCCCC")
                      for s in ("left","right","top","bottom")})
    HDR_F = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    BOD_F = Font(name="Calibri", bold=False, color="000000", size=10)
    BLD_F = Font(name="Calibri", bold=True,  color="000000", size=10)

    def __init__(self, output_path: str):
        self.out = output_path
        self.wb  = Workbook()
        self.wb.remove(self.wb.active)
        # Reuse mapping writer for per-mapping sheets
        self._mw = MappingExcelWriter.__new__(MappingExcelWriter)
        self._mw.out  = output_path
        self._mw.wb   = self.wb
        self._mw.BORD = self.BORD
        self._mw.HDR_F= self.HDR_F
        self._mw.BOD_F= self.BOD_F
        self._mw.BLD_F= self.BLD_F
        self._mw.C    = MappingExcelWriter.C

    def _fill(self, c: str) -> PatternFill:
        hex_c = self.C.get(c, c)
        hex_c = hex_c.lstrip("#")
        if len(hex_c) not in (6,8):
            hex_c = "FFFFFF"
        return PatternFill("solid", fgColor=hex_c)

    def _ws(self, t: str):
        ws = self.wb.create_sheet(title=t[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _hc(self, ws, r, c, v, w=None, clr="navy"):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.HDR_F
        cell.fill      = self._fill(clr)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center", wrap_text=True)
        cell.border    = self.BORD
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[r].height = 30
        return cell

    def _dc(self, ws, r, c, v, shade=False, wrap=False,
            align="left", clr=None, bold=False):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.BLD_F if bold else self.BOD_F
        cell.fill      = (self._fill(clr) if clr
                          else self._fill("alt_a") if shade
                          else self._fill("white"))
        cell.alignment = Alignment(horizontal=align,
                                   vertical="top" if wrap else "center",
                                   wrap_text=wrap)
        cell.border    = self.BORD
        ws.row_dimensions[r].height = 38 if wrap else 22
        return cell

    def _ch(self, ws, row, hdrs, wids, clr="sub"):
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._hc(ws, row, c, h, w, clr)
        return row + 1

    def _banner(self, ws, row, text, clr="navy", cols=14):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Calibri", bold=True,
                           color="FFFFFF", size=11)
        c.fill      = self._fill(clr)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = self.BORD
        ws.row_dimensions[row].height = 26
        return row + 1

    # ─────────────────────────────────────────────────────────────
    # SHEET 1: Workflow Summary
    # ─────────────────────────────────────────────────────────────

    def sheet_wf_summary(self, parser: WorkflowParser):
        ws   = self._ws("1_WorkflowSummary")
        hdrs = ["Workflow","Server","Step #","Task Name","Task Type",
                "Link Condition","Session","Mapping",
                "Source Tables","Target Tables",
                "Transforms","Maplets Used","Pipelines"]
        wids = [28,20,8,30,16,30,30,35,28,28,12,35,14]
        self._ch(ws, 1, hdrs, wids, "navy")
        row = 2

        for wf_name, wf in parser.workflows.items():
            eo_map = {task:(i+1,cond)
                      for i,(task,cond) in
                      enumerate(wf.execution_order)}
            for task in wf.tasks:
                if task.task_type in ("Start","start","End"):
                    continue
                sess  = parser.sessions.get(task.name)
                m     = parser.mappings.get(
                    sess.mapping_name if sess else "", None)
                eo_num, cond = eo_map.get(task.name,("?",""))
                pipe_ids = sorted(set(
                    st.pipeline for st in
                    (sess.sess_transforms if sess else [])))
                maplets_used = (", ".join(m.maplet_instances)
                                if m and m.maplet_instances else "None")
                shade = (row % 2 == 0)
                for c, v in enumerate([
                    wf_name, wf.server, str(eo_num),
                    task.name, task.task_type, cond,
                    task.name if task.task_type=="Session" else "",
                    sess.mapping_name if sess else "",
                    ", ".join(m.sources.keys()) if m else "",
                    ", ".join(m.targets.keys()) if m else "",
                    str(len(m.transformations)) if m else "",
                    maplets_used,
                    str(pipe_ids) if pipe_ids else "[0]",
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             wrap=(c in (6,12)),
                             clr=("alt_maplet"
                                  if m and m.maplet_instances
                                  and shade else None))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET 2: Session Detail  (Pipeline / Stage / Maplet flags)
    # ─────────────────────────────────────────────────────────────

    def sheet_session_detail(self, parser: WorkflowParser):
        ws   = self._ws("2_SessionDetail")
        hdrs = ["Session","Mapping","Pipeline","Stage",
                "Instance Name","Transform Type",
                "Is Maplet","Is Repartition","Partition Type",
                "Exec Mode (derived)","Src Connection","Tgt Connection"]
        wids = [30,35,10,8,32,22,10,14,18,35,22,22]
        self._ch(ws, 1, hdrs, wids, "teal")
        row = 2

        for s_name, sess in parser.sessions.items():
            stis = sorted(sess.sess_transforms,
                          key=lambda x:(x.pipeline, x.stage))
            if not stis:
                shade = (row % 2 == 0)
                for c, v in enumerate([
                    s_name, sess.mapping_name,
                    "-","-","-","-","-","-","-",
                    "No SESSTRANSFORMATIONINST data",
                    sess.src_connection, sess.tgt_connection
                ], 1):
                    self._dc(ws, row, c, v, shade=shade)
                row += 1
                continue

            pipe_ids    = sorted(set(st.pipeline for st in stis))
            is_parallel = len(pipe_ids) > 1
            # Determine which instances are maplets
            m = parser.mappings.get(sess.mapping_name)
            maplet_names = set(m.maplet_instances) if m else set()

            for sti in stis:
                shade    = (row % 2 == 0)
                is_maplet = sti.instance_name in maplet_names
                if is_parallel:
                    letter    = chr(97 + pipe_ids.index(sti.pipeline))
                    exec_mode = (f"Parallel — Pipeline {sti.pipeline},"
                                 f" Stage {sti.stage}")
                    pc = "gold"
                else:
                    exec_mode = f"Sequential — Stage {sti.stage}"
                    pc = None
                if sti.is_repartition:
                    exec_mode += f" [REPARTITION:{sti.partition_type}]"
                    pc = "cyan"
                if is_maplet:
                    exec_mode += " [MAPLET]"

                for c, v in enumerate([
                    s_name, sess.mapping_name,
                    str(sti.pipeline), str(sti.stage),
                    sti.instance_name, sti.transformation_type,
                    "YES" if is_maplet else "NO",
                    "YES" if sti.is_repartition else "NO",
                    sti.partition_type, exec_mode,
                    sess.src_connection, sess.tgt_connection
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             clr=("alt_maplet"
                                  if is_maplet and shade else
                                  "F9F0FA" if is_maplet else
                                  pc if c in (3,10) else None))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET 3: Execution Sequence
    # ─────────────────────────────────────────────────────────────

    def sheet_exec_sequence(self, parser: WorkflowParser):
        ws   = self._ws("3_ExecutionSequence")
        hdrs = ["Workflow","Step","Task Name","Task Type",
                "Link Condition","Mapping",
                "Source Tables","Target Tables",
                "Maplets Used","Pipelines","Notes"]
        wids = [28,8,30,16,35,35,28,28,35,16,35]
        self._ch(ws, 1, hdrs, wids, "dark")
        row = 2

        for wf_name, wf in parser.workflows.items():
            for step, (task_name, cond) in \
                    enumerate(wf.execution_order, 1):
                sess = parser.sessions.get(task_name)
                m    = parser.mappings.get(
                    sess.mapping_name if sess else "", None)
                pipe_ids = sorted(set(
                    st.pipeline for st in
                    (sess.sess_transforms if sess else [])))
                note = ""
                if len(pipe_ids) > 1:
                    note = f"⚡ {len(pipe_ids)} parallel pipelines"
                if m and m.maplet_instances:
                    note += (f"  ◈ maplets: "
                             f"{', '.join(m.maplet_instances)}")
                shade = (row % 2 == 0)
                for c, v in enumerate([
                    wf_name, str(step), task_name,
                    "Session" if sess else "Worklet/Other",
                    cond,
                    sess.mapping_name if sess else "",
                    ", ".join(m.sources.keys()) if m else "",
                    ", ".join(m.targets.keys()) if m else "",
                    ", ".join(m.maplet_instances)
                    if m and m.maplet_instances else "None",
                    str(pipe_ids) if pipe_ids else "[0]",
                    note
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             wrap=(c in (5,9,11)))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # GENERATE
    # ─────────────────────────────────────────────────────────────

    def generate(self, parser: WorkflowParser) -> None:
        print(f"\n[WRITE] {self.out}")

        # Workflow-level sheets
        self.sheet_wf_summary(parser)
        self.sheet_session_detail(parser)
        self.sheet_exec_sequence(parser)

        # Per-mapping sheets — uses mapping writer
        # Annotate exec order from session PIPELINE/STAGE data first
        for m_name, mapping in parser.mappings.items():
            graph = LineageGraph(mapping)

            sess = next(
                (s for s in parser.sessions.values()
                 if s.mapping_name == m_name), None)

            if sess and sess.sess_transforms:
                # Use PIPELINE/STAGE for accurate parallel detection
                ts          = mapping.transformations
                pipe_groups = defaultdict(list)
                for sti in sess.sess_transforms:
                    pipe_groups[sti.pipeline].append(sti)
                pipe_ids    = sorted(pipe_groups.keys())
                is_parallel = len(pipe_ids) > 1
                maplet_names = set(mapping.maplet_instances)

                for pid in pipe_ids:
                    stis = sorted(pipe_groups[pid],
                                  key=lambda x: x.stage)
                    for sti in stis:
                        t = ts.get(sti.instance_name)
                        if not t:
                            continue
                        if is_parallel:
                            letter       = chr(97 + pipe_ids.index(pid))
                            t.exec_order = f"{letter}{sti.stage}"
                            t.exec_mode  = (
                                f"Parallel — "
                                f"Pipeline {pid}, Stage {sti.stage}")
                        else:
                            t.exec_order = str(sti.stage)
                            t.exec_mode  = (
                                f"Sequential — Stage {sti.stage}")
                        if t.trans_type in ("Aggregator","Sorter"):
                            t.exec_mode += " [BARRIER]"
                        if sti.is_repartition:
                            t.exec_mode += (
                                f" [REPARTITION:{sti.partition_type}]")
                        if sti.instance_name in maplet_names:
                            t.exec_mode += " [MAPLET]"
            else:
                graph.annotate_exec_order()

            safe = re.sub(r'[^A-Za-z0-9_]', '_', m_name)[:14]

            # Delegate to mapping writer's sheet methods
            mw = self._mw
            mw.sheet_mapping_parse(mapping, graph,
                                    f"{safe}_A_MappingParse")
            mw.sheet_source_detail(mapping,  f"{safe}_B_Sources")
            mw.sheet_target_detail(mapping,  f"{safe}_C_Targets")
            mw.sheet_transforms   (mapping,  f"{safe}_D_Transforms")
            mw.sheet_column_flow  (mapping, graph,
                                   f"{safe}_E_ColFlow")
            mw.sheet_lookups      (mapping,  f"{safe}_F_Lookups")
            mw.sheet_maplet_detail(mapping, parser.maplets,
                                   f"{safe}_G_MapletDetail")
            mw.sheet_exec_order   (mapping, graph,
                                   f"{safe}_H_ExecOrder")

            print(f"  [OK  ] {m_name}  "
                  f"maplets={mapping.maplet_instances}")

        self.wb.save(self.out)
        print(f"\n  ✅  Saved → {Path(self.out).resolve()}")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="LineageIQ Workflow Parser v4 — "
                    "full maplet architecture support",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
--------
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml --par params_prod.par
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml --par params_prod.par --out MyOutput.xlsx
        """
    )
    ap.add_argument("--xml", required=True)
    ap.add_argument("--par", nargs="*", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    xml_path  = args.xml
    par_files = args.par or []
    out_path  = args.out or \
        f"LineageIQ_Workflow_{Path(xml_path).stem}.xlsx"

    print()
    print("═" * 65)
    print("  LineageIQ — WORKFLOW PARSER  v4  (full maplet support)")
    print("═" * 65)
    print(f"  XML : {xml_path}")
    print(f"  PAR : {par_files or 'None ($$PARAMS kept as-is)'}")
    print(f"  OUT : {out_path}")
    print("═" * 65)

    resolver = ParamResolver(par_files)
    resolver.report()

    parser = WorkflowParser(resolver)
    parser.parse(xml_path)

    if not parser.workflows and not parser.mappings:
        print("\n  [ERROR] No workflow or mapping tags found.")
        return

    writer = WorkflowExcelWriter(out_path)
    writer.generate(parser)

    print()
    print("═" * 65)
    print(f"  ✅  Done!  → {Path(out_path).resolve()}")
    print("═" * 65)


if __name__ == "__main__":
    main()
