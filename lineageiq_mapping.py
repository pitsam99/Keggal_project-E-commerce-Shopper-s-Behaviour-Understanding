"""
╔══════════════════════════════════════════════════════════════════════╗
║   LineageIQ — MAPPING PARSER                                         ║
║   Script : lineageiq_mapping.py                                      ║
║   Purpose: Parse a single mapping XML and generate a full STTM Excel ║
║                                                                       ║
║   Sheets produced                                                     ║
║     A_MappingParse    — Full mapping-level parse (5 sections)        ║
║     B_SourceDetail    — Source schema / table / columns / SQ SQL     ║
║     C_TargetDetail    — Target schema / table / columns              ║
║     D_Transforms      — All transforms in exec order + logic         ║
║     E_ColumnFlowMap   — SQ col → T1 → T2 → … → Target col           ║
║     F_Lookups         — Lookup table / condition / return ports       ║
║     G_Maplets         — Maplet ports + CRC/hash expressions          ║
║     H_ExecOrder       — Pipeline / stage / parallel detection         ║
║                                                                       ║
║   Works WITH or WITHOUT a parameter file                              ║
║                                                                       ║
║   USAGE                                                               ║
║     # With parameter file                                             ║
║     python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║                                  --par params_prod.par                ║
║                                                                       ║
║     # Without parameter file ($$PARAMS kept as-is)                   ║
║     python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║                                                                       ║
║     # Custom output name                                              ║
║     python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║                                  --par params_prod.par                ║
║                                  --out MyMapping_STTM.xlsx            ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict

import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════

@dataclass
class ColumnDef:
    name:      str
    datatype:  str
    precision: str
    scale:     str
    nullable:  str
    key_type:  str = "NOT A KEY"


@dataclass
class TransformPort:
    name:       str
    port_type:  str
    datatype:   str
    precision:  str
    scale:      str
    expression: str = ""


@dataclass
class Transformation:
    name:       str
    trans_type: str
    ports:      list = field(default_factory=list)
    attributes: dict = field(default_factory=dict)
    exec_order: str  = ""
    exec_mode:  str  = ""


@dataclass
class Connector:
    from_instance: str
    from_field:    str
    to_instance:   str
    to_field:      str


@dataclass
class Mapping:
    name:            str
    description:     str
    folder_name:     str = ""
    sources:         dict = field(default_factory=dict)
    targets:         dict = field(default_factory=dict)
    source_schemas:  dict = field(default_factory=dict)   # table → schema/owner
    target_schemas:  dict = field(default_factory=dict)
    transformations: dict = field(default_factory=dict)
    connectors:      list = field(default_factory=list)
    maplets_used:    list = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════
# MODULE 1 — PARAMETER RESOLVER
# ══════════════════════════════════════════════════════════════════════

class ParamResolver:
    """
    Reads Informatica .par files and substitutes $$PARAM references.
    If no par file is provided, $$PARAMS are returned unchanged.
    """

    def __init__(self, par_files: list = None):
        self.params:     dict = {}
        self.unresolved: list = []
        self.has_params: bool = False

        if par_files:
            for pf in par_files:
                self._load(pf)

    def _load(self, par_path: str) -> None:
        p = Path(par_path)
        if not p.exists():
            print(f"  [WARN] Par file not found: {par_path} — continuing without it")
            return
        print(f"  [PAR ] Loading: {par_path}")
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip()
                    if k not in self.params:
                        self.params[k] = v
                        print(f"         {k} = {v}")
        self.has_params = True

    def resolve(self, text: str) -> str:
        """Replace $$PARAMS. If no par loaded, return text unchanged."""
        if not text:
            return text
        if not self.has_params:
            return text   # keep $$PARAMS as-is when no par file

        def sub(m):
            pname = m.group(0)
            if pname in self.params:
                return self.params[pname]
            self.unresolved.append(pname)
            return f"[UNRESOLVED:{pname}]"

        return re.sub(r'\$\$[A-Z0-9_]+', sub, text)

    def report(self) -> None:
        if not self.has_params:
            print("  [INFO] No parameter file supplied — $$PARAMS kept as-is")
            return
        seen   = set()
        unique = [x for x in self.unresolved
                  if not (x in seen or seen.add(x))]
        if unique:
            print(f"  [WARN] Unresolved params: {unique}")
        else:
            print("  [OK  ] All parameters resolved")


# ══════════════════════════════════════════════════════════════════════
# MODULE 2 — MAPPING XML PARSER
# ══════════════════════════════════════════════════════════════════════

class MappingParser:
    """
    Parses Informatica PowerCenter mapping XML.
    Handles PC 10.5.x (v189.x) and PC 9.x formats.
    Supports both:
      - Single mapping export  (just <MAPPING> tags)
      - Workflow export        (extracts MAPPINGs from inside workflow)
    """

    def __init__(self, resolver: ParamResolver):
        self.resolver        = resolver
        self.mappings:  dict = {}
        self.maplets:   dict = {}
        self._global_sources: dict = {}
        self._global_targets: dict = {}

    def parse(self, xml_path: str) -> None:
        print(f"\n[PARSE] {xml_path}")

        # Handle DTD declaration and encoding (PC 10.x)
        xml_parser = ET.XMLParser(
            load_dtd        = False,
            no_network      = True,
            resolve_entities= False,
            recover         = True,
        )
        tree = ET.parse(xml_path, xml_parser)
        root = tree.getroot()

        # Find folder (may be absent in minimal mapping exports)
        folder = root.find(".//FOLDER") or root

        self._parse_sources(folder)
        self._parse_targets(folder)
        self._parse_maplets(folder)
        self._parse_mappings(folder)

        print(f"  [OK  ] Found {len(self.mappings)} mapping(s), "
              f"{len(self.maplets)} maplet(s)")

    # ── Sources ───────────────────────────────────────────────────

    def _parse_sources(self, parent) -> None:
        for src in parent.findall(".//SOURCE"):
            name   = src.get("NAME", "")
            owner  = self.resolver.resolve(
                src.get("OWNERNAME", src.get("DBDNAME", "")))
            cols   = [
                ColumnDef(
                    name      = sf.get("NAME", ""),
                    datatype  = sf.get("DATATYPE", ""),
                    precision = sf.get("PRECISION", ""),
                    scale     = sf.get("SCALE", "0"),
                    nullable  = sf.get("NULLABLE", ""),
                )
                for sf in src.findall("SOURCEFIELD")
            ]
            self._global_sources[name]       = cols
            self._src_owner = getattr(self, "_src_owner", {})
            self._src_owner[name] = owner
            print(f"  [SRC ] {name}  schema={owner}  ({len(cols)} cols)")

    # ── Targets ───────────────────────────────────────────────────

    def _parse_targets(self, parent) -> None:
        for tgt in parent.findall(".//TARGET"):
            name  = tgt.get("NAME", "")
            owner = self.resolver.resolve(
                tgt.get("OWNERNAME", tgt.get("DBDNAME", "")))
            cols  = [
                ColumnDef(
                    name      = tf.get("NAME", ""),
                    datatype  = tf.get("DATATYPE", ""),
                    precision = tf.get("PRECISION", ""),
                    scale     = tf.get("SCALE", "0"),
                    nullable  = tf.get("NULLABLE", ""),
                    key_type  = tf.get("KEYTYPE", "NOT A KEY"),
                )
                for tf in tgt.findall("TARGETFIELD")
            ]
            self._global_targets[name]       = cols
            self._tgt_owner = getattr(self, "_tgt_owner", {})
            self._tgt_owner[name] = owner
            print(f"  [TGT ] {name}  schema={owner}  ({len(cols)} cols)")

    # ── Maplets ───────────────────────────────────────────────────

    def _parse_maplets(self, parent) -> None:
        for ml in parent.findall(".//MAPPLET"):
            name = ml.get("NAME", "")
            t    = Transformation(
                name       = name,
                trans_type = "Maplet",
                attributes = {"description": ml.get("DESCRIPTION", "")}
            )
            for inner in ml.findall(".//TRANSFORMATION"):
                for tf in inner.findall("TRANSFORMFIELD"):
                    t.ports.append(TransformPort(
                        name       = tf.get("NAME", ""),
                        port_type  = tf.get("PORTTYPE", ""),
                        datatype   = tf.get("DATATYPE", ""),
                        precision  = tf.get("PRECISION", ""),
                        scale      = tf.get("SCALE", "0"),
                        expression = self.resolver.resolve(
                            tf.get("EXPRESSION", "")),
                    ))
            self.maplets[name] = t
            print(f"  [MLET] {name}")

    # ── Mappings ──────────────────────────────────────────────────

    def _parse_mappings(self, parent) -> None:
        for mp in parent.findall(".//MAPPING"):
            m = Mapping(
                name        = mp.get("NAME", "UNKNOWN"),
                description = mp.get("DESCRIPTION", ""),
            )

            # Source / target instances
            for inst in mp.findall("INSTANCE"):
                itype = inst.get("TYPE", "")
                iname = inst.get("TRANSFORMATION_NAME",
                                 inst.get("NAME", ""))
                if itype == "SOURCE":
                    if iname in self._global_sources:
                        m.sources[iname] = self._global_sources[iname]
                        m.source_schemas[iname] = getattr(
                            self, "_src_owner", {}).get(iname, "")
                elif itype == "TARGET":
                    if iname in self._global_targets:
                        m.targets[iname] = self._global_targets[iname]
                        m.target_schemas[iname] = getattr(
                            self, "_tgt_owner", {}).get(iname, "")

            # Transformations
            for trans in mp.findall("TRANSFORMATION"):
                t = self._parse_transform(trans)
                m.transformations[t.name] = t
                if t.trans_type == "Maplet":
                    m.maplets_used.append(t.name)

            # Reusable maplet instances
            for inst in mp.findall("INSTANCE"):
                if inst.get("TRANSFORMATION_TYPE") == "Maplet":
                    mname = inst.get("TRANSFORMATION_NAME", "")
                    if mname in self.maplets \
                            and mname not in m.transformations:
                        m.transformations[mname] = self.maplets[mname]
                        if mname not in m.maplets_used:
                            m.maplets_used.append(mname)

            # Connectors
            for conn in mp.findall("CONNECTOR"):
                m.connectors.append(Connector(
                    from_instance = conn.get("FROMINSTANCE", ""),
                    from_field    = conn.get("FROMFIELD", ""),
                    to_instance   = conn.get("TOINSTANCE", ""),
                    to_field      = conn.get("TOFIELD", ""),
                ))

            self.mappings[m.name] = m
            print(f"  [MAP ] {m.name}:  "
                  f"src={list(m.sources.keys())}  "
                  f"tgt={list(m.targets.keys())}  "
                  f"transforms={len(m.transformations)}  "
                  f"connectors={len(m.connectors)}")

    def _parse_transform(self, el) -> Transformation:
        t = Transformation(
            name       = el.get("NAME", ""),
            trans_type = el.get("TYPE", ""),
        )
        for tf in el.findall("TRANSFORMFIELD"):
            t.ports.append(TransformPort(
                name       = tf.get("NAME", ""),
                port_type  = tf.get("PORTTYPE", ""),
                datatype   = tf.get("DATATYPE", ""),
                precision  = tf.get("PRECISION", ""),
                scale      = tf.get("SCALE", "0"),
                expression = self.resolver.resolve(
                    tf.get("EXPRESSION", "")),
            ))
        for ta in el.findall("TABLEATTRIBUTE"):
            k = ta.get("NAME", "")
            v = self.resolver.resolve(ta.get("VALUE", ""))
            t.attributes[k] = v
        return t


# ══════════════════════════════════════════════════════════════════════
# MODULE 3 — LINEAGE GRAPH
# ══════════════════════════════════════════════════════════════════════

class LineageGraph:
    """
    Builds directed graph from CONNECTOR tags.
    Detects parallel branches and merge points via topology.
    """

    def __init__(self, mapping: Mapping):
        self.mapping         = mapping
        self.G               = nx.DiGraph()   # port-level
        self.IG              = nx.DiGraph()   # instance-level
        self._build()

    def _build(self) -> None:
        for conn in self.mapping.connectors:
            src = f"{conn.from_instance}.{conn.from_field}"
            tgt = f"{conn.to_instance}.{conn.to_field}"
            self.G.add_edge(src, tgt,
                from_instance=conn.from_instance,
                from_field=conn.from_field,
                to_instance=conn.to_instance,
                to_field=conn.to_field)
            if not self.IG.has_edge(conn.from_instance, conn.to_instance):
                self.IG.add_edge(conn.from_instance, conn.to_instance)

    def annotate_exec_order(self) -> None:
        """
        Topological sort on instance graph.
        Detects split points (out_degree > 1) → parallel branches.
        Detects merge points (in_degree > 1)  → convergence.
        """
        try:
            topo = list(nx.topological_sort(self.IG))
        except nx.NetworkXUnfeasible:
            topo = list(self.IG.nodes)

        ts          = self.mapping.transformations
        split_nodes = {n for n in self.IG.nodes
                       if self.IG.out_degree(n) > 1}
        merge_nodes = {n for n in self.IG.nodes
                       if self.IG.in_degree(n) > 1}

        level          = 0
        parallel_level = None
        branch_count:  dict = {}
        in_parallel:   set  = set()

        for inst in topo:
            if inst in split_nodes:
                level += 1
                parallel_level = level + 1
                branch_count[parallel_level] = 0
                _ord, _mode = str(level), "Sequential — Split Point"
            elif inst in merge_nodes:
                in_parallel.discard(inst)
                parallel_level = None
                level += 1
                _ord, _mode = str(level), "Sequential — Merge Point"
            elif parallel_level is not None:
                in_parallel.add(inst)
                branch_count[parallel_level] = \
                    branch_count.get(parallel_level, 0) + 1
                letter = chr(96 + branch_count[parallel_level])
                _ord   = f"{parallel_level}{letter}"
                _mode  = f"Parallel Branch {letter.upper()}"
            else:
                level += 1
                _ord, _mode = str(level), "Sequential"

            if inst in ts:
                t = ts[inst]
                t.exec_order = _ord
                t.exec_mode  = _mode
                if t.trans_type in ("Aggregator", "Sorter"):
                    t.exec_mode += " [BARRIER]"

    def topo_order(self) -> list:
        try:
            return list(nx.topological_sort(self.IG))
        except Exception:
            return list(self.IG.nodes)

    def trace_back(self, tgt_inst: str, tgt_col: str) -> list:
        """BFS backward from target port → list of edge dicts."""
        node = f"{tgt_inst}.{tgt_col}"
        if node not in self.G:
            return []
        edges, queue, visited = [], [node], set()
        while queue:
            n = queue.pop()
            if n in visited:
                continue
            visited.add(n)
            for pred in self.G.predecessors(n):
                e = self.G.edges[pred, n]
                edges.append(dict(e))
                queue.append(pred)
        return edges

    def build_logic(self, t: Transformation, port: str) -> str:
        tt = t.trans_type
        if tt == "Expression":
            for p in t.ports:
                if p.name == port and p.expression:
                    return p.expression
            return "Pass-through"
        elif tt == "Lookup Procedure":
            table = t.attributes.get("Lookup table name", "")
            cond  = t.attributes.get("Lookup condition", "")
            cache = t.attributes.get("Lookup cache persistent", "NO")
            return f"LOOKUP: {table}  ON ({cond})  RETURN {port}  [Cache={cache}]"
        elif tt == "Filter":
            return "FILTER: " + t.attributes.get("Filter Condition", "")
        elif tt == "Update Strategy":
            return "UPD_STRATEGY: " + \
                   t.attributes.get("Update Strategy Expression", "")
        elif tt == "Maplet":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"MAPLET [{t.name}]: {p.expression}"
            return f"MAPLET [{t.name}]: Pass-through"
        elif tt == "Source Qualifier":
            sql = t.attributes.get("Sql Query", "")
            flt = t.attributes.get("Source Filter", "")
            return f"SQ_SQL: {sql}" if sql else \
                   (f"SQ_FILTER: {flt}" if flt else "Pass-through")
        elif tt == "Router":
            return "ROUTER: " + "; ".join(
                f"{k}={v}" for k, v in t.attributes.items()
                if "Group" in k)
        elif tt == "Aggregator":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"AGG: {p.expression}"
            return "Pass-through"
        elif tt == "Sequence":
            return (f"SEQUENCE  start={t.attributes.get('Start Value','1')} "
                    f"incr={t.attributes.get('Increment By','1')}")
        elif tt == "Joiner":
            return ("JOINER  condition=" +
                    t.attributes.get("Join Condition", "") +
                    "  type=" + t.attributes.get("Join Type", ""))
        elif tt == "Normalizer":
            return "NORMALIZER"
        elif tt == "Stored Procedure":
            return "SP: " + t.attributes.get("Stored Procedure Name", "")
        return "Pass-through"


# ══════════════════════════════════════════════════════════════════════
# MODULE 4 — EXCEL WRITER (Mapping-level)
# ══════════════════════════════════════════════════════════════════════

class MappingExcelWriter:

    # Colour palette
    C = {
        "navy":    "0D1B2A",
        "teal":    "065A82",
        "purple":  "4A235A",
        "dark":    "2C3E50",
        "brown":   "7B341E",
        "green_d": "1A6B5A",
        "sub":     "1A3A5C",
        "gold":    "F4A621",
        "green":   "22C55E",
        "cyan":    "00B4D8",
        "red":     "EF4444",
        "white":   "FFFFFF",
        "alt_a":   "E8F4FD",
        "alt_b":   "E8FDF4",
        "alt_c":   "F5EEF8",
        "alt_d":   "FEF9E7",
        "alt_e":   "FEF5E4",
        "pale":    "D6E4F0",
    }

    THIN  = Side(style="thin",   color="CCCCCC")
    BORD  = Border(**{s: Side(style="thin", color="CCCCCC")
                      for s in ("left","right","top","bottom")})
    HDR_F = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    BOD_F = Font(name="Calibri", size=10)
    BLD_F = Font(name="Calibri", bold=True, size=10)

    def __init__(self, output_path: str):
        self.out = output_path
        self.wb  = Workbook()
        self.wb.remove(self.wb.active)

    def _fill(self, color_key: str) -> PatternFill:
        hex_color = self.C.get(color_key, color_key)
        hex_color = hex_color.lstrip("#")
        if len(hex_color) not in (6, 8):
            hex_color = "FFFFFF"
        return PatternFill("solid", fgColor=hex_color)

    def _ws(self, title: str):
        ws = self.wb.create_sheet(title=title[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _hcell(self, ws, r, c, v, w=None, color="navy"):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.HDR_F
        cell.fill      = self._fill(self.C[color])
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center",
                                   wrap_text=True)
        cell.border    = self.BORD
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[r].height = 30
        return cell

    def _dcell(self, ws, r, c, v, shade=False, wrap=False,
               align="left", color=None, bold=False):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.BLD_F if bold else self.BOD_F
        cell.fill      = (self._fill(color) if color
                          else self._fill("alt_a") if shade
                          else self._fill("white"))
        cell.alignment = Alignment(horizontal=align,
                                   vertical="top" if wrap else "center",
                                   wrap_text=wrap)
        cell.border    = self.BORD
        ws.row_dimensions[r].height = 38 if wrap else 22
        return cell

    def _banner(self, ws, row, text, color="navy", cols=14):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Calibri", bold=True,
                           color="FFFFFF", size=11)
        c.fill      = self._fill(self.C[color])
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = self.BORD
        ws.row_dimensions[row].height = 26
        return row + 1

    def _col_hdrs(self, ws, row, hdrs, wids, color="sub"):
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._hcell(ws, row, c, h, w, color)
        return row + 1

    # ─────────────────────────────────────────────────────────────
    # SHEET A: Full Mapping Parse (5 sections in one sheet)
    # ─────────────────────────────────────────────────────────────

    def sheet_mapping_parse(self, mapping: Mapping,
                             graph: LineageGraph,
                             maplets: dict,
                             sheet_name: str):
        ws = self._ws(sheet_name)

        # ── A1: SOURCE DETAILS ────────────────────────────────────
        row = self._banner(ws, 1,
            f"  A — SOURCE DETAILS     Mapping: {mapping.name}",
            "navy", cols=14)
        hdrs = ["Source Schema", "Source Table",
                "Source Column", "Datatype", "Precision",
                "Nullable", "SQ Name", "SQ Column",
                "SQL Override / Filter"]
        wids = [22, 25, 25, 14, 10, 10, 25, 25, 65]
        row  = self._col_hdrs(ws, row, hdrs, wids, "sub")

        for src_name, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src_name, "")

            # Build SQ column map: src_col → sq_col
            sq_name = "-"
            sq_sql  = "-"
            sq_col_map: dict = {}
            for conn in mapping.connectors:
                if conn.from_instance == src_name:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_col_map[conn.from_field] = conn.to_field
                        sql = t.attributes.get("Sql Query", "")
                        flt = t.attributes.get("Source Filter", "")
                        sq_sql = sql or flt or "Default (no SQL override)"

            for i, col in enumerate(cols):
                shade  = (i % 2 == 0)
                sq_col = sq_col_map.get(col.name, "-")
                vals   = [schema, src_name, col.name,
                          col.datatype, col.precision, col.nullable,
                          sq_name, sq_col, sq_sql]
                for c, v in enumerate(vals, 1):
                    self._dcell(ws, row, c, v, shade=shade,
                                wrap=(c == 9))
                row += 1

        row += 1

        # ── A2: TARGET DETAILS ────────────────────────────────────
        row = self._banner(ws, row,
            "  B — TARGET DETAILS", "teal", cols=14)
        hdrs = ["Target Schema", "Target Table", "Target Column",
                "Datatype", "Precision", "Scale",
                "Nullable", "Key Type"]
        wids = [22, 25, 25, 14, 10, 8, 10, 16]
        row  = self._col_hdrs(ws, row, hdrs, wids, "teal")

        for tgt_name, cols in mapping.targets.items():
            schema = mapping.target_schemas.get(tgt_name, "")
            for i, col in enumerate(cols):
                shade  = (i % 2 == 0)
                pk     = col.key_type == "PRIMARY KEY"
                clr    = "gold" if pk else None
                for c, v in enumerate([
                    schema, tgt_name, col.name,
                    col.datatype, col.precision, col.scale,
                    col.nullable, col.key_type
                ], 1):
                    self._dcell(ws, row, c, v, shade=shade, color=clr)
                row += 1

        row += 1

        # ── A3: TRANSFORMATION INVENTORY ──────────────────────────
        row = self._banner(ws, row,
            "  C — TRANSFORMATION INVENTORY  (Sequence | Type | Logic)",
            "purple", cols=14)
        hdrs = ["Exec Order", "Exec Mode", "Transform Name",
                "Transform Type", "Port Name", "Port Type",
                "Expression / Logic", "Attributes"]
        wids = [12, 32, 28, 22, 25, 16, 58, 58]
        row  = self._col_hdrs(ws, row, hdrs, wids, "purple")

        def eo_key(t):
            eo = t.exec_order or "z99"
            l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
            n  = int(re.sub(r'[^0-9]', '', eo) or "0")
            return (l, n)

        for t in sorted(mapping.transformations.values(), key=eo_key):
            attr_str = " | ".join(
                f"{k}: {v}" for k, v in t.attributes.items()
                if v and k != "description")

            em_color = None
            if "Parallel" in (t.exec_mode or ""):
                em_color = "gold"
            elif "BARRIER" in (t.exec_mode or ""):
                em_color = "green"
            elif "REPARTITION" in (t.exec_mode or ""):
                em_color = "cyan"

            ports = t.ports or [TransformPort(
                "-", "-", "-", "-", "0", "-")]
            for pi, p in enumerate(ports):
                shade = (pi % 2 == 0)
                vals  = [
                    t.exec_order      if pi == 0 else "",
                    t.exec_mode       if pi == 0 else "",
                    t.name            if pi == 0 else "",
                    t.trans_type      if pi == 0 else "",
                    p.name, p.port_type,
                    p.expression or "Pass-through",
                    attr_str          if pi == 0 else "",
                ]
                for c, v in enumerate(vals, 1):
                    clr = em_color if (c == 2 and pi == 0) else None
                    self._dcell(ws, row, c, v, shade=shade,
                                wrap=(c in (7, 8)), color=clr)
                row += 1

        row += 1

        # ── A4: COLUMN FLOW MAP ───────────────────────────────────
        row = self._banner(ws, row,
            "  D — COLUMN FLOW MAP"
            "  [ Source Column  →  SQ  →  Transformations  →  Target Column ]",
            "dark", cols=14)
        hdrs = ["#", "Source Schema", "Source Table",
                "Source Column",
                "SQ Name", "SQ Column",
                "Transformation Chain\n(Name|Type|Order|Mode)",
                "Logic at Each Step",
                "Target Schema", "Target Table",
                "Target Column", "Target DT",
                "Key Type", "Remarks"]
        wids = [5, 18, 20, 20, 20, 20, 55, 62, 18, 20, 20, 13, 13, 14]
        row  = self._col_hdrs(ws, row, hdrs, wids, "dark")

        seq = 1
        for tgt_name, tgt_cols in mapping.targets.items():
            tgt_schema = mapping.target_schemas.get(tgt_name, "")
            for col in tgt_cols:
                edges = graph.trace_back(tgt_name, col.name)

                if not edges:
                    for c, v in enumerate([
                        str(seq), "-", "-", "-", "-", "-",
                        "UNCONNECTED", "-",
                        tgt_schema, tgt_name, col.name,
                        col.datatype, col.key_type,
                        "No upstream connection"
                    ], 1):
                        self._dcell(ws, row, c, v)
                    ws.row_dimensions[row].height = 22
                    row += 1
                    seq += 1
                    continue

                # Identify src, SQ, intermediate transforms
                src_table  = src_col  = "-"
                src_schema = "-"
                sq_name    = sq_col   = "-"
                chain_steps = []
                logic_steps = []

                # Sort edges by instance exec_order
                def inst_ord(inst):
                    t = mapping.transformations.get(inst)
                    if t and t.exec_order:
                        eo = t.exec_order
                        l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
                        n  = int(re.sub(r'[^0-9]', '', eo) or "0")
                        return (l, n)
                    return (0, 0) if inst in mapping.sources else (99, 99)

                seen_inst = set()
                ordered   = []
                for e in edges:
                    for inst in [e["from_instance"], e["to_instance"]]:
                        if inst not in seen_inst:
                            seen_inst.add(inst)
                            ordered.append(inst)
                ordered.sort(key=inst_ord)

                for inst in ordered:
                    t = mapping.transformations.get(inst)
                    # Find field at this instance
                    field = "-"
                    for e in edges:
                        if e["from_instance"] == inst:
                            field = e["from_field"]
                            break
                        if e["to_instance"] == inst:
                            field = e["to_field"]

                    if inst in mapping.sources:
                        src_table  = inst
                        src_col    = field
                        src_schema = mapping.source_schemas.get(inst, "")
                    elif t and t.trans_type == "Source Qualifier":
                        sq_name = inst
                        sq_col  = field
                    elif inst not in mapping.targets and t:
                        label = (f"{inst}  [{t.trans_type}]  "
                                 f"Order:{t.exec_order}  {t.exec_mode}")
                        chain_steps.append(label)
                        logic = graph.build_logic(t, field)
                        if logic and logic != "-":
                            logic_steps.append(f"▶ {inst}: {logic}")

                trans_chain = "\n→ ".join(chain_steps) \
                              if chain_steps else "Direct / Pass-through"
                logic_chain = "\n".join(logic_steps) \
                              if logic_steps else "Pass-through"

                shade = (seq % 2 == 0)
                vals  = [
                    str(seq),
                    src_schema, src_table, src_col,
                    sq_name, sq_col,
                    trans_chain, logic_chain,
                    tgt_schema, tgt_name,
                    col.name, col.datatype, col.key_type,
                    "PK" if col.key_type == "PRIMARY KEY" else ""
                ]

                lines = max(len(chain_steps), len(logic_steps), 1)
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c,
                                   value=str(v) if v else "")
                    cell.font   = self.BOD_F
                    cell.fill   = self._fill("alt_d") if shade \
                                  else self._fill("white")
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top",
                        wrap_text=True)
                    cell.border = self.BORD
                ws.row_dimensions[row].height = max(38, lines * 18)
                row += 1
                seq += 1

        row += 1

        # ── A5: FLOW DIAGRAM ROADMAP ──────────────────────────────
        row = self._banner(ws, row,
            "  E — FLOW DIAGRAM ROADMAP  "
            "(Steps to build a visual lineage diagram)",
            "brown", cols=14)
        roadmap = self._build_roadmap(mapping, graph)
        for line in roadmap:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=14)
            cell  = ws.cell(row=row, column=1, value=line)
            cell.font = Font(name="Consolas", size=9,
                             bold=("STEP" in line or
                                   "════" in line),
                             color=("FFFFFF"
                                    if line.startswith("  STEP")
                                    else "1A1A1A"))
            cell.fill = (self._fill("sub")
                         if line.startswith("  STEP")
                         else self._fill("alt_a")
                         if "→" in line
                         else self._fill("white"))
            cell.alignment = Alignment(horizontal="left",
                                        vertical="center")
            cell.border = self.BORD
            ws.row_dimensions[row].height = 15
            row += 1

        ws.freeze_panes = "A4"

    def _build_roadmap(self, mapping: Mapping,
                        graph: LineageGraph) -> list:
        lines = []
        W     = 110

        lines.append("═" * W)
        lines.append(f"  FLOW DIAGRAM ROADMAP   Mapping: {mapping.name}")
        lines.append("═" * W)
        lines.append("")

        # STEP 1: Sources
        lines.append(f"  STEP 1 — SOURCE LAYER")
        lines.append("  " + "─" * 60)
        for src, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src, "?")
            sq_name, sq_sql = "-", "-"
            for conn in mapping.connectors:
                if conn.from_instance == src:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_sql  = t.attributes.get(
                            "Sql Query", "")[:100]
                        break
            lines.append(f"   SCHEMA: {schema}  |  TABLE: {src}  "
                         f"({len(cols)} cols)")
            lines.append(f"   SQ    : {sq_name}")
            if sq_sql:
                lines.append(f"   SQL   : {sq_sql}"
                             f"{'...' if len(sq_sql)==100 else ''}")
        lines.append("")

        # STEP 2: Transformation chain
        lines.append("  STEP 2 — TRANSFORMATION CHAIN  (Execution Order)")
        lines.append("  " + "─" * 60)

        def eo_key(t):
            eo = t.exec_order or "z99"
            l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
            n  = int(re.sub(r'[^0-9]', '', eo) or "0")
            return (l, n)

        split_nodes = {n for n in graph.IG.nodes
                       if graph.IG.out_degree(n) > 1}
        merge_nodes = {n for n in graph.IG.nodes
                       if graph.IG.in_degree(n) > 1}

        for t in sorted(mapping.transformations.values(), key=eo_key):
            flag = ""
            if t.name in split_nodes:
                flag = "  ⚡ SPLIT → parallel branches start here"
            elif t.name in merge_nodes:
                flag = "  🔀 MERGE ← branches converge here"
            if t.trans_type in ("Aggregator", "Sorter"):
                flag += "  🛑 BARRIER"

            lines.append(
                f"   [{t.exec_order:>5}]  {t.name:<35}  "
                f"({t.trans_type})  {t.exec_mode}{flag}")

            # Show key logic
            attrs = t.attributes
            if t.trans_type == "Lookup Procedure":
                lines.append(
                    f"            ▶ LOOKUP: "
                    f"{attrs.get('Lookup table name', '')}  "
                    f"ON {attrs.get('Lookup condition', '')}")
            elif t.trans_type == "Filter":
                lines.append(
                    f"            ▶ FILTER: "
                    f"{attrs.get('Filter Condition', '')}")
            elif t.trans_type == "Update Strategy":
                lines.append(
                    f"            ▶ UPD: "
                    f"{attrs.get('Update Strategy Expression', '')}")
            elif t.trans_type == "Expression":
                expr_ports = [p for p in t.ports
                              if p.expression
                              and "OUTPUT" in p.port_type][:3]
                for p in expr_ports:
                    lines.append(
                        f"            ▶ {p.name} = "
                        f"{p.expression[:80]}")
        lines.append("")

        # STEP 3: Targets
        lines.append("  STEP 3 — TARGET LAYER")
        lines.append("  " + "─" * 60)
        for tgt, cols in mapping.targets.items():
            schema  = mapping.target_schemas.get(tgt, "?")
            pk_cols = [c.name for c in cols
                       if c.key_type == "PRIMARY KEY"]
            lines.append(
                f"   SCHEMA: {schema}  |  TABLE: {tgt}  "
                f"({len(cols)} cols)")
            if pk_cols:
                lines.append(
                    f"   PK    : {', '.join(pk_cols)}")
        lines.append("")

        # Additional steps for visual diagram
        lines.append("  ADDITIONAL STEPS — BUILDING A VISUAL FLOW DIAGRAM")
        lines.append("  " + "─" * 60)
        steps = [
            ("STEP A", "Use Section D rows as directed graph edges",
             "Each row = SQ_Column → Transform1 → Transform2 → Target_Column"),
            ("STEP B", "Assign node types and colours",
             "SOURCE=blue  SQ=teal  EXPR=orange  LKP=purple  FILTER=red  "
             "MAPLET=pink  TARGET=green"),
            ("STEP C", "Assign edge labels",
             "Edge label = expression / filter condition / strategy"),
            ("STEP D", "Group by parallel branch",
             "Parallel branches → swim-lanes  |  "
             "Merge points → converge lanes"),
            ("STEP E", "Recommended tools",
             "Python: graphviz / networkx+matplotlib  |  "
             "Web: D3.js / vis.js  |  Manual: draw.io / Lucidchart"),
            ("STEP F", "Cross-mapping lineage",
             "Match target cols of this mapping to source cols of next mapping "
             "to build TPR→TT→DDM end-to-end lineage"),
        ]
        for sid, title, detail in steps:
            lines.append(f"   {sid}  {title}")
            lines.append(f"         • {detail}")
        lines.append("")
        lines.append("═" * W)
        return lines

    # ─────────────────────────────────────────────────────────────
    # SHEET B: Source Detail (dedicated)
    # ─────────────────────────────────────────────────────────────

    def sheet_source_detail(self, mapping: Mapping, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Source Schema", "Source Table",
                "Source Column", "Datatype",
                "Precision", "Scale", "Nullable",
                "SQ Name", "SQ Column",
                "SQL Override / Filter"]
        wids = [22, 25, 25, 14, 10, 8, 10, 25, 25, 70]
        self._col_hdrs(ws, 1, hdrs, wids, "navy")
        row = 2
        for src_name, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src_name, "")
            sq_name, sq_sql, sq_col_map = "-", "-", {}
            for conn in mapping.connectors:
                if conn.from_instance == src_name:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_col_map[conn.from_field] = conn.to_field
                        sql = t.attributes.get("Sql Query", "")
                        flt = t.attributes.get("Source Filter", "")
                        sq_sql = sql or flt or "Default"
            for i, col in enumerate(cols):
                shade = (i % 2 == 0)
                sq_col = sq_col_map.get(col.name, "-")
                for c, v in enumerate([
                    schema, src_name, col.name, col.datatype,
                    col.precision, col.scale, col.nullable,
                    sq_name, sq_col, sq_sql
                ], 1):
                    self._dcell(ws, row, c, v, shade=shade,
                                wrap=(c == 10))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET C: Target Detail (dedicated)
    # ─────────────────────────────────────────────────────────────

    def sheet_target_detail(self, mapping: Mapping, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Target Schema", "Target Table",
                "Target Column", "Datatype",
                "Precision", "Scale", "Nullable", "Key Type"]
        wids = [22, 25, 25, 14, 10, 8, 10, 16]
        self._col_hdrs(ws, 1, hdrs, wids, "teal")
        row = 2
        for tgt_name, cols in mapping.targets.items():
            schema = mapping.target_schemas.get(tgt_name, "")
            for i, col in enumerate(cols):
                shade = (i % 2 == 0)
                pk    = col.key_type == "PRIMARY KEY"
                for c, v in enumerate([
                    schema, tgt_name, col.name, col.datatype,
                    col.precision, col.scale,
                    col.nullable, col.key_type
                ], 1):
                    self._dcell(ws, row, c, v, shade=shade,
                                color="gold" if pk else None)
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET D: Transformation Detail (dedicated)
    # ─────────────────────────────────────────────────────────────

    def sheet_transforms(self, mapping: Mapping, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Exec Order", "Exec Mode", "Transform Name",
                "Transform Type", "Port Name", "Port Type",
                "Datatype", "Expression / Logic", "Attributes"]
        wids = [12, 32, 28, 22, 25, 16, 14, 58, 58]
        self._col_hdrs(ws, 1, hdrs, wids, "purple")
        row = 2

        def eo_key(t):
            eo = t.exec_order or "z99"
            l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
            n  = int(re.sub(r'[^0-9]', '', eo) or "0")
            return (l, n)

        for t in sorted(mapping.transformations.values(), key=eo_key):
            attr_str = " | ".join(
                f"{k}: {v}" for k, v in t.attributes.items()
                if v and k != "description")
            ports = t.ports or [TransformPort("-", "-", "-", "-", "0")]
            for pi, p in enumerate(ports):
                shade = (pi % 2 == 0)
                em_c  = ("gold"  if "Parallel" in (t.exec_mode or "")
                         else "green" if "BARRIER" in (t.exec_mode or "")
                         else "cyan" if "REPARTITION" in (t.exec_mode or "")
                         else None)
                for c, v in enumerate([
                    t.exec_order if pi == 0 else "",
                    t.exec_mode  if pi == 0 else "",
                    t.name       if pi == 0 else "",
                    t.trans_type if pi == 0 else "",
                    p.name, p.port_type, p.datatype,
                    p.expression or "Pass-through",
                    attr_str     if pi == 0 else "",
                ], 1):
                    self._dcell(ws, row, c, v, shade=shade,
                                wrap=(c in (8, 9)),
                                color=em_c if c == 2 and pi == 0
                                      else None)
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET E: Column Flow Map (dedicated)
    # ─────────────────────────────────────────────────────────────

    def sheet_column_flow(self, mapping: Mapping,
                           graph: LineageGraph,
                           sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["#", "Source Schema", "Source Table",
                "Source Column",
                "SQ Name", "SQ Column",
                "Transformation Chain\n(Name|Type|Order|Mode)",
                "Logic at Each Step",
                "Target Schema", "Target Table",
                "Target Column", "Target DT",
                "Key Type", "Remarks"]
        wids = [5, 18, 22, 22, 22, 22, 55, 62, 18, 22, 22, 14, 13, 14]
        self._col_hdrs(ws, 1, hdrs, wids, "dark")
        row = 2
        seq = 1

        for tgt_name, tgt_cols in mapping.targets.items():
            tgt_schema = mapping.target_schemas.get(tgt_name, "")
            for col in tgt_cols:
                edges = graph.trace_back(tgt_name, col.name)
                if not edges:
                    for c, v in enumerate([
                        str(seq), "-", "-", "-", "-", "-",
                        "UNCONNECTED", "-",
                        tgt_schema, tgt_name, col.name,
                        col.datatype, col.key_type,
                        "No upstream connection"
                    ], 1):
                        self._dcell(ws, row, c, v)
                    row += 1
                    seq += 1
                    continue

                src_table = src_col = src_schema = "-"
                sq_name   = sq_col  = "-"
                chain_steps, logic_steps = [], []

                def inst_ord(inst):
                    t = mapping.transformations.get(inst)
                    if t and t.exec_order:
                        eo = t.exec_order
                        l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
                        n  = int(re.sub(r'[^0-9]', '', eo) or "0")
                        return (l, n)
                    return (0, 0) if inst in mapping.sources else (99, 99)

                seen, ordered = set(), []
                for e in edges:
                    for inst in [e["from_instance"], e["to_instance"]]:
                        if inst not in seen:
                            seen.add(inst)
                            ordered.append(inst)
                ordered.sort(key=inst_ord)

                for inst in ordered:
                    t = mapping.transformations.get(inst)
                    field = "-"
                    for e in edges:
                        if e["from_instance"] == inst:
                            field = e["from_field"]
                            break
                        if e["to_instance"] == inst:
                            field = e["to_field"]

                    if inst in mapping.sources:
                        src_table  = inst
                        src_col    = field
                        src_schema = mapping.source_schemas.get(inst, "")
                    elif t and t.trans_type == "Source Qualifier":
                        sq_name = inst
                        sq_col  = field
                    elif inst not in mapping.targets and t:
                        label = (f"{inst}  [{t.trans_type}]  "
                                 f"Ord:{t.exec_order}  {t.exec_mode}")
                        chain_steps.append(label)
                        lg = graph.build_logic(t, field)
                        if lg and lg != "-":
                            logic_steps.append(f"▶ {inst}: {lg}")

                tc = "\n→ ".join(chain_steps) if chain_steps \
                     else "Direct / Pass-through"
                lc = "\n".join(logic_steps)  if logic_steps \
                     else "Pass-through"

                shade = (seq % 2 == 0)
                lines = max(len(chain_steps), len(logic_steps), 1)
                for c, v in enumerate([
                    str(seq),
                    src_schema, src_table, src_col,
                    sq_name, sq_col, tc, lc,
                    tgt_schema, tgt_name,
                    col.name, col.datatype, col.key_type,
                    "PK" if col.key_type == "PRIMARY KEY" else ""
                ], 1):
                    cell = ws.cell(row=row, column=c,
                                   value=str(v) if v else "")
                    cell.font = self.BOD_F
                    cell.fill = (self._fill("alt_d") if shade
                                 else self._fill("white"))
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top",
                        wrap_text=True)
                    cell.border = self.BORD
                ws.row_dimensions[row].height = max(38, lines * 18)
                row += 1
                seq += 1

        ws.freeze_panes = "B2"

    # ─────────────────────────────────────────────────────────────
    # SHEET F: Lookups
    # ─────────────────────────────────────────────────────────────

    def sheet_lookups(self, mapping: Mapping, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Lookup Name", "Lookup Table / View",
                "Lookup Condition", "Return Columns",
                "Cache Type", "Exec Order", "Exec Mode"]
        wids = [28, 32, 45, 28, 14, 12, 30]
        self._col_hdrs(ws, 1, hdrs, wids, "teal")
        row = 2
        for t in mapping.transformations.values():
            if t.trans_type == "Lookup Procedure":
                s    = (row % 2 == 0)
                rets = [p.name for p in t.ports
                        if "OUTPUT" in p.port_type]
                for c, v in enumerate([
                    t.name,
                    t.attributes.get("Lookup table name", ""),
                    t.attributes.get("Lookup condition", ""),
                    ", ".join(rets),
                    t.attributes.get("Lookup cache persistent", "NO"),
                    t.exec_order, t.exec_mode
                ], 1):
                    self._dcell(ws, row, c, v, shade=s,
                                wrap=(c == 3))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET G: Maplets
    # ─────────────────────────────────────────────────────────────

    def sheet_maplets(self, mapping: Mapping,
                       maplets: dict, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Maplet Name", "Port Name", "Port Type",
                "Datatype", "Expression / Formula", "Purpose"]
        wids = [28, 22, 14, 14, 60, 30]
        self._col_hdrs(ws, 1, hdrs, wids, "purple")
        row = 2
        for mname in mapping.maplets_used:
            t = maplets.get(mname) or mapping.transformations.get(mname)
            if not t:
                continue
            purpose = ("CRC Checksum for CDC"
                       if "CRC"  in mname.upper() else
                       "Hash Key Generation"
                       if "HASH" in mname.upper() else
                       "Sequence / Surrogate Key"
                       if "SEQ"  in mname.upper() else "")
            for p in t.ports:
                s = (row % 2 == 0)
                for c, v in enumerate([
                    t.name, p.name, p.port_type, p.datatype,
                    p.expression or "Pass-through", purpose
                ], 1):
                    self._dcell(ws, row, c, v, shade=s,
                                wrap=(c == 5))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET H: Execution Order
    # ─────────────────────────────────────────────────────────────

    def sheet_exec_order(self, mapping: Mapping,
                          graph: LineageGraph, sheet_name: str):
        ws   = self._ws(sheet_name)
        hdrs = ["Exec Order", "Instance Name", "Transform Type",
                "Exec Mode", "In-Degree", "Out-Degree", "Notes"]
        wids = [12, 35, 22, 38, 10, 10, 50]
        self._col_hdrs(ws, 1, hdrs, wids, "navy")
        topo = graph.topo_order()
        ts   = mapping.transformations
        row  = 2
        for inst in topo:
            t    = ts.get(inst)
            in_d = graph.IG.in_degree(inst)
            out_d= graph.IG.out_degree(inst)
            s    = (row % 2 == 0)
            note = ""
            if out_d > 1:
                note = f"⚡ SPLIT → {out_d} parallel branches"
            if in_d > 1:
                note = f"🔀 MERGE ← {in_d} branches converge"
            if t and t.trans_type in ("Aggregator", "Sorter"):
                note += " 🛑 BARRIER"
            em_c = (
                "gold"  if t and "Parallel" in (t.exec_mode or "")
                else "green" if t and "Merge"  in (t.exec_mode or "")
                else "cyan"  if t and "Split"  in (t.exec_mode or "")
                else None
            )
            for c, v in enumerate([
                t.exec_order if t else "?",
                inst,
                t.trans_type if t else "Source/Target",
                t.exec_mode  if t else "Sequential",
                str(in_d), str(out_d), note
            ], 1):
                self._dcell(ws, row, c, v, shade=s,
                            align="center" if c in (1,5,6) else "left",
                            color=em_c if c == 4 else None,
                            wrap=(c == 7))
            row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # GENERATE (all sheets)
    # ─────────────────────────────────────────────────────────────

    def generate(self, parser: MappingParser) -> None:
        print(f"\n[WRITE] {self.out}")
        for m_name, mapping in parser.mappings.items():
            graph = LineageGraph(mapping)
            graph.annotate_exec_order()
            safe  = re.sub(r'[^A-Za-z0-9_]', '_', m_name)[:16]

            self.sheet_mapping_parse(
                mapping, graph, parser.maplets,
                f"{safe}_MappingParse")
            self.sheet_source_detail(mapping, f"{safe}_B_Sources")
            self.sheet_target_detail(mapping, f"{safe}_C_Targets")
            self.sheet_transforms   (mapping, f"{safe}_D_Transforms")
            self.sheet_column_flow  (mapping, graph, f"{safe}_E_ColFlow")
            self.sheet_lookups      (mapping, f"{safe}_F_Lookups")
            self.sheet_maplets      (mapping, parser.maplets,
                                     f"{safe}_G_Maplets")
            self.sheet_exec_order   (mapping, graph, f"{safe}_H_ExecOrder")

            print(f"  [OK  ] Sheets written for: {m_name}")

        self.wb.save(self.out)
        print(f"\n  ✅  Saved → {Path(self.out).resolve()}")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="LineageIQ Mapping Parser — "
                    "generates STTM Excel from an Informatica mapping XML",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
--------
  # With parameter file
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml --par params_prod.par

  # Without parameter file ($$PARAMS kept as-is in output)
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml

  # Custom output name
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml --par params_prod.par
                               --out MyMapping_STTM.xlsx

  # Workflow XML (extracts all mappings inside)
  python lineageiq_mapping.py --xml wf_TCOM_RR.xml --par params_prod.par
        """
    )
    ap.add_argument("--xml", required=True,
                    help="Path to Informatica mapping or workflow XML file")
    ap.add_argument("--par", nargs="*", default=None,
                    help="Path(s) to .par parameter file(s) — optional")
    ap.add_argument("--out", default=None,
                    help="Output Excel path — auto-named if not given")
    args = ap.parse_args()

    xml_path  = args.xml
    par_files = args.par or []
    xml_stem  = Path(xml_path).stem
    out_path  = args.out or f"LineageIQ_Mapping_{xml_stem}.xlsx"

    print()
    print("═" * 65)
    print("  LineageIQ — MAPPING PARSER")
    print("═" * 65)
    print(f"  XML  : {xml_path}")
    print(f"  PAR  : {par_files or 'None ($$PARAMS kept as-is)'}")
    print(f"  OUT  : {out_path}")
    print("═" * 65)

    resolver = ParamResolver(par_files)
    resolver.report()

    parser = MappingParser(resolver)
    parser.parse(xml_path)

    if not parser.mappings:
        print("\n  [ERROR] No <MAPPING> tags found in the XML file.")
        return

    writer = MappingExcelWriter(out_path)
    writer.generate(parser)

    print()
    print("═" * 65)
    print(f"  ✅  Done!  Output → {Path(out_path).resolve()}")
    print("═" * 65)


if __name__ == "__main__":
    main()
