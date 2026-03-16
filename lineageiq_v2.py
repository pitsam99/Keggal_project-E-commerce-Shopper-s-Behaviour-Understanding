"""
╔══════════════════════════════════════════════════════════════════════╗
║            LineageIQ — Informatica Data Lineage Framework            ║
║            Revised for PowerCenter 10.5.8 (v189.x)                  ║
║                                                                      ║
║  Fix 1 — DTD parser    : handles <!DOCTYPE> & ISO-8859-1 encoding   ║
║  Fix 2 — SESSION tag   : reads MAPPINGNAME as direct attribute       ║
║  Fix 3 — PIPELINE/STAGE: uses SESSTRANSFORMATIONINST for exec order ║
║  Fix 4 — $$PARAM targets: resolves $$SCHEMA / $$TABLE_NAME attrs    ║
║                                                                      ║
║  Input  : <workflow_name>.xml   (Informatica XML export)             ║
║           <env>_params.par      (Unix parameter file)                ║
║  Output : LineageIQ_STTM_<mapping_name>.xlsx                         ║
╚══════════════════════════════════════════════════════════════════════╝

Input / Output Examples
-----------------------
  Input XML  : wf_TCOM_RR.xml
  Input PAR  : params_prod.par
  Output XLSX: LineageIQ_STTM_wf_TCOM_RR.xlsx

Usage
-----
  python lineageiq.py --xml wf_TCOM_RR.xml --par params_prod.par
  python lineageiq.py --xml wf_TCOM_RR.xml --par params_prod.par --out MyOutput.xlsx
  python lineageiq.py   (uses defaults at bottom of file)
"""

import os
import re
import sys
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict

import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════

@dataclass
class ColumnDef:
    name:      str
    datatype:  str
    precision: str
    scale:     str
    nullable:  str
    key_type:  str = "NOT A KEY"


@dataclass
class TransformPort:
    name:       str
    port_type:  str       # INPUT / OUTPUT / INPUT/OUTPUT / VARIABLE
    datatype:   str
    precision:  str
    scale:      str
    expression: str = ""


@dataclass
class Transformation:
    name:       str
    trans_type: str
    ports:      list = field(default_factory=list)
    attributes: dict = field(default_factory=dict)
    exec_order: str  = ""
    exec_mode:  str  = ""


@dataclass
class Connector:
    from_instance: str
    from_field:    str
    to_instance:   str
    to_field:      str


@dataclass
class SessionTransformInst:
    """One SESSTRANSFORMATIONINST entry — carries PIPELINE and STAGE."""
    instance_name:       str
    transformation_name: str
    transformation_type: str
    pipeline:            int    # parallel pipeline index (0,1,2,3…)
    stage:               int    # sequential stage within pipeline
    is_repartition:      bool
    partition_type:      str


@dataclass
class Mapping:
    name:         str
    description:  str
    sources:      dict = field(default_factory=dict)   # name → [ColumnDef]
    targets:      dict = field(default_factory=dict)
    transformations: dict = field(default_factory=dict)
    connectors:   list = field(default_factory=list)
    maplets_used: list = field(default_factory=list)


@dataclass
class Session:
    name:           str
    mapping_name:   str
    src_connection: str
    tgt_connection: str
    sess_transforms: list = field(default_factory=list)  # [SessionTransformInst]


@dataclass
class Workflow:
    name:            str
    sessions:        list = field(default_factory=list)
    execution_order: list = field(default_factory=list)


@dataclass
class LineageRow:
    seq:                  str
    source_table:         str
    source_column:        str
    source_datatype:      str
    sq_column:            str
    transformation_name:  str
    transformation_type:  str
    transformation_logic: str
    exec_order:           str
    exec_mode:            str
    target_column:        str
    target_datatype:      str
    target_table:         str
    remarks:              str = ""


# ══════════════════════════════════════════════════════════════════════
# MODULE 1 — PARAMETER FILE RESOLVER
# ══════════════════════════════════════════════════════════════════════

class ParamResolver:
    """
    Reads Informatica .par files and resolves $$PARAM references.
    Supports multi-environment priority: PROD > UAT > DEV.

    Par file format (Unix):
        # comment
        $$SRC_SCHEMA=TPR_PROD
        $$FILTER_DATE=20240101
    """

    def __init__(self):
        self.params:     dict = {}
        self.unresolved: list = []

    def load_file(self, par_path: str) -> None:
        path = Path(par_path)
        if not path.exists():
            print(f"  [WARN] Parameter file not found: {par_path}")
            return
        print(f"  [PAR ] Loading: {par_path}")
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, _, val = line.partition("=")
                    key = key.strip()
                    val = val.strip()
                    if key not in self.params:     # PROD first wins
                        self.params[key] = val
                        print(f"         {key} = {val}")

    def load_directory(self, dir_path: str,
                       priority: list = None) -> None:
        priority = priority or ["prod", "uat", "dev"]
        par_files = sorted(
            Path(dir_path).glob("*.par"),
            key=lambda p: next(
                (i for i, env in enumerate(priority)
                 if env in p.name.lower()), 99
            )
        )
        for pf in par_files:
            self.load_file(str(pf))

    def resolve(self, text: str) -> str:
        if not text:
            return text

        def replacer(match):
            pname = match.group(0)
            if pname in self.params:
                return self.params[pname]
            self.unresolved.append(pname)
            return f"[UNRESOLVED:{pname}]"

        return re.sub(r'\$\$[A-Z0-9_]+', replacer, text)

    def report(self) -> None:
        seen   = set()
        unique = [x for x in self.unresolved
                  if not (x in seen or seen.add(x))]
        if unique:
            print(f"\n  [WARN] Unresolved parameters ({len(unique)}): "
                  f"{unique}")
        else:
            print("  [OK  ] All parameters resolved.")


# ══════════════════════════════════════════════════════════════════════
# MODULE 2 — INFORMATICA XML PARSER  (PC 10.5.8 compatible)
# ══════════════════════════════════════════════════════════════════════

class InformaticaParser:
    """
    Parses Informatica PowerCenter 10.5.8 XML exports.

    Key fixes vs PC 9.x:
      - DTD / ISO-8859-1 handling via ET.XMLParser
      - SESSION.MAPPINGNAME read as direct XML attribute
      - SESSTRANSFORMATIONINST PIPELINE + STAGE captured
      - $$PARAM in target ATTRIBUTE tags resolved
    """

    def __init__(self, resolver: ParamResolver):
        self.resolver       = resolver
        self.mappings:  dict = {}
        self.sessions:  dict = {}
        self.workflows: dict = {}
        self.maplets:   dict = {}
        self.global_sources: dict = {}
        self.global_targets: dict = {}

    # ── Entry point ───────────────────────────────────────────────

    def parse(self, xml_path: str) -> None:
        print(f"\n[PARSE] {xml_path}")

        # ── FIX 1: Handle DTD declaration + encoding ──────────────
        # PowerCenter exports include <!DOCTYPE POWERMART SYSTEM "powrmart.dtd">
        # lxml must NOT try to fetch this file. recover=True handles minor XML errors.
        xml_parser = ET.XMLParser(
            load_dtd        = False,
            no_network      = True,
            resolve_entities= False,
            recover         = True,
        )
        tree = ET.parse(xml_path, xml_parser)
        root = tree.getroot()
        # ── end fix 1 ─────────────────────────────────────────────

        folder = root.find(".//FOLDER")
        if folder is None:
            raise ValueError("No <FOLDER> tag found in XML.")

        self._parse_sources(folder)
        self._parse_targets(folder)
        self._parse_maplets(folder)
        self._parse_mappings(folder)
        self._parse_sessions(folder)
        self._parse_workflows(folder)

        print(f"  [OK  ] Parsed: "
              f"{len(self.mappings)} mappings, "
              f"{len(self.sessions)} sessions, "
              f"{len(self.workflows)} workflows, "
              f"{len(self.maplets)} maplets")

    # ── Sources ───────────────────────────────────────────────────

    def _parse_sources(self, folder) -> None:
        for src in folder.findall("SOURCE"):
            name = src.get("NAME")
            cols = [
                ColumnDef(
                    name      = sf.get("NAME", ""),
                    datatype  = sf.get("DATATYPE", ""),
                    precision = sf.get("PRECISION", ""),
                    scale     = sf.get("SCALE", "0"),
                    nullable  = sf.get("NULLABLE", ""),
                )
                for sf in src.findall("SOURCEFIELD")
            ]
            self.global_sources[name] = cols
            print(f"  [SRC ] {name} ({len(cols)} cols)")

    # ── Targets ───────────────────────────────────────────────────

    def _parse_targets(self, folder) -> None:
        for tgt in folder.findall("TARGET"):
            name = tgt.get("NAME")
            cols = [
                ColumnDef(
                    name      = tf.get("NAME", ""),
                    datatype  = tf.get("DATATYPE", ""),
                    precision = tf.get("PRECISION", ""),
                    scale     = tf.get("SCALE", "0"),
                    nullable  = tf.get("NULLABLE", ""),
                    key_type  = tf.get("KEYTYPE", "NOT A KEY"),
                )
                for tf in tgt.findall("TARGETFIELD")
            ]
            self.global_targets[name] = cols
            print(f"  [TGT ] {name} ({len(cols)} cols)")

    # ── Maplets ───────────────────────────────────────────────────

    def _parse_maplets(self, folder) -> None:
        for ml in folder.findall("MAPPLET"):
            name = ml.get("NAME")
            t    = Transformation(
                name       = name,
                trans_type = "Maplet",
                attributes = {"description": ml.get("DESCRIPTION", "")}
            )
            for inner in ml.findall("TRANSFORMATION"):
                for tf in inner.findall("TRANSFORMFIELD"):
                    expr = self.resolver.resolve(tf.get("EXPRESSION", ""))
                    t.ports.append(TransformPort(
                        name      = tf.get("NAME", ""),
                        port_type = tf.get("PORTTYPE", ""),
                        datatype  = tf.get("DATATYPE", ""),
                        precision = tf.get("PRECISION", ""),
                        scale     = tf.get("SCALE", "0"),
                        expression= expr,
                    ))
            self.maplets[name] = t
            print(f"  [MLET] {name}")

    # ── Mappings ──────────────────────────────────────────────────

    def _parse_mappings(self, folder) -> None:
        for mp in folder.findall("MAPPING"):
            m = Mapping(
                name        = mp.get("NAME"),
                description = mp.get("DESCRIPTION", ""),
            )
            # Source / target instances
            for inst in mp.findall("INSTANCE"):
                itype = inst.get("TYPE")
                iname = inst.get("TRANSFORMATION_NAME",
                                 inst.get("NAME"))
                if itype == "SOURCE" and iname in self.global_sources:
                    m.sources[iname] = self.global_sources[iname]
                elif itype == "TARGET" and iname in self.global_targets:
                    m.targets[iname] = self.global_targets[iname]

            # Transformations
            for trans in mp.findall("TRANSFORMATION"):
                t = self._parse_transformation(trans)
                m.transformations[t.name] = t
                if t.trans_type == "Maplet":
                    m.maplets_used.append(t.name)

            # Maplet instances (REUSABLE="YES")
            for inst in mp.findall("INSTANCE"):
                if inst.get("TRANSFORMATION_TYPE") == "Maplet":
                    mname = inst.get("TRANSFORMATION_NAME")
                    if (mname in self.maplets
                            and mname not in m.transformations):
                        m.transformations[mname] = self.maplets[mname]
                        if mname not in m.maplets_used:
                            m.maplets_used.append(mname)

            # Connectors
            for conn in mp.findall("CONNECTOR"):
                m.connectors.append(Connector(
                    from_instance = conn.get("FROMINSTANCE"),
                    from_field    = conn.get("FROMFIELD"),
                    to_instance   = conn.get("TOINSTANCE"),
                    to_field      = conn.get("TOFIELD"),
                ))

            self.mappings[m.name] = m
            print(f"  [MAP ] {m.name}: "
                  f"{len(m.sources)} src, {len(m.targets)} tgt, "
                  f"{len(m.transformations)} transforms, "
                  f"{len(m.connectors)} connectors")

    def _parse_transformation(self, el) -> Transformation:
        t = Transformation(
            name      = el.get("NAME"),
            trans_type= el.get("TYPE"),
        )
        for tf in el.findall("TRANSFORMFIELD"):
            t.ports.append(TransformPort(
                name      = tf.get("NAME", ""),
                port_type = tf.get("PORTTYPE", ""),
                datatype  = tf.get("DATATYPE", ""),
                precision = tf.get("PRECISION", ""),
                scale     = tf.get("SCALE", "0"),
                expression= self.resolver.resolve(
                                tf.get("EXPRESSION", "")),
            ))
        for ta in el.findall("TABLEATTRIBUTE"):
            key = ta.get("NAME", "")
            val = self.resolver.resolve(ta.get("VALUE", ""))
            t.attributes[key] = val
        return t

    # ── Sessions — FIX 2 + FIX 3 + FIX 4 ────────────────────────

    def _parse_sessions(self, folder) -> None:
        for sess in folder.findall("SESSION"):

            # ── FIX 2: MAPPINGNAME is a direct attribute in PC 10.x ──
            mapping_name = sess.get("MAPPINGNAME", "")

            # Fallback: older PC 9.x style uses <ATTRIBUTE NAME="Mapping name">
            if not mapping_name:
                for attr in sess.findall("ATTRIBUTE"):
                    if "Mapping name" in attr.get("NAME", ""):
                        mapping_name = self.resolver.resolve(
                            attr.get("VALUE", ""))
                        break

            src_conn = ""
            tgt_conn = ""

            # ── FIX 3 + FIX 4: Parse SESSTRANSFORMATIONINST ──────────
            sess_transforms = []
            for sti in sess.findall("SESSTRANSFORMATIONINST"):
                ttype    = sti.get("TRANSFORMATIONTYPE", "")
                sname    = sti.get("SINSTANCENAME", "")
                pipeline = int(sti.get("PIPELINE", "0"))
                stage    = int(sti.get("STAGE",    "0"))

                # FIX 4: Resolve $$PARAM in target table attrs
                for attr in sti.findall("ATTRIBUTE"):
                    aname = attr.get("NAME", "")
                    aval  = self.resolver.resolve(attr.get("VALUE", ""))
                    if aname == "Table Name Prefix" and ttype == "Target Definition":
                        tgt_conn = aval
                    if aname == "$Source connection value":
                        src_conn = aval
                    if aname == "$Target connection value" and not tgt_conn:
                        tgt_conn = aval

                # Track first source connection name
                if ttype == "Source Definition" and not src_conn:
                    src_conn = sname

                sess_transforms.append(SessionTransformInst(
                    instance_name       = sname,
                    transformation_name = sti.get("TRANSFORMATIONNAME", sname),
                    transformation_type = ttype,
                    pipeline            = pipeline,
                    stage               = stage,
                    is_repartition      = sti.get(
                        "ISREPARTITIONPOINT", "NO") == "YES",
                    partition_type      = sti.get("PARTITIONTYPE", ""),
                ))

            s = Session(
                name             = sess.get("NAME"),
                mapping_name     = mapping_name,
                src_connection   = src_conn,
                tgt_connection   = tgt_conn,
                sess_transforms  = sess_transforms,
            )
            self.sessions[s.name] = s
            print(f"  [SESS] {s.name}")
            print(f"         mapping  : {s.mapping_name}")
            print(f"         src_conn : {s.src_connection}")
            print(f"         tgt_conn : {s.tgt_connection}")
            print(f"         instances: {len(sess_transforms)}")

    # ── Workflows + Worklets ──────────────────────────────────────

    def _parse_workflows(self, folder) -> None:
        worklet_tasks: dict = {}
        for wkl in folder.findall("WORKLET"):
            wname = wkl.get("NAME")
            tasks = [ti.get("NAME")
                     for ti in wkl.findall("TASKINSTANCE")]
            worklet_tasks[wname] = tasks

        for wf in folder.findall("WORKFLOW"):
            w = Workflow(name=wf.get("NAME"))
            for ti in wf.findall("TASKINSTANCE"):
                tname = ti.get("NAME")
                ttype = ti.get("TASKTYPE", "")
                if ttype == "Worklet" and tname in worklet_tasks:
                    w.sessions.extend(worklet_tasks[tname])
                elif ttype == "Session":
                    w.sessions.append(tname)

            # Execution order from LINK chain
            links: dict = {}
            for lnk in wf.findall("LINK"):
                links[lnk.get("FROMTASK")] = lnk.get("TOTASK")
            order   = []
            cur     = "start"
            visited = set()
            while cur in links and cur not in visited:
                visited.add(cur)
                nxt = links[cur]
                if nxt and nxt != "start":
                    order.append(nxt)
                cur = nxt
            w.execution_order = order
            self.workflows[w.name] = w
            print(f"  [WF  ] {w.name}: "
                  f"sessions={w.sessions}, order={w.execution_order}")


# ══════════════════════════════════════════════════════════════════════
# MODULE 3 — LINEAGE GRAPH  (NetworkX + PIPELINE/STAGE)
# ══════════════════════════════════════════════════════════════════════

class LineageGraph:
    """
    Builds directed graph from CONNECTOR tags.
    Uses PIPELINE + STAGE from SESSTRANSFORMATIONINST for execution
    order (more accurate than topology inference for PC 10.x).
    Falls back to topological sort when session data is unavailable.
    """

    def __init__(self, mapping: Mapping):
        self.mapping          = mapping
        self.G                = nx.DiGraph()   # port-level
        self._instance_graph  = nx.DiGraph()   # instance-level
        self._build()

    def _build(self) -> None:
        for conn in self.mapping.connectors:
            src = f"{conn.from_instance}.{conn.from_field}"
            tgt = f"{conn.to_instance}.{conn.to_field}"
            self.G.add_edge(src, tgt,
                            from_instance = conn.from_instance,
                            from_field    = conn.from_field,
                            to_instance   = conn.to_instance,
                            to_field      = conn.to_field)
            if not self._instance_graph.has_edge(
                    conn.from_instance, conn.to_instance):
                self._instance_graph.add_edge(
                    conn.from_instance, conn.to_instance)

    # ── FIX 3: Annotate from PIPELINE + STAGE ────────────────────

    def annotate_from_session(self,
                               sess_transforms: list) -> None:
        """
        Assign exec_order and exec_mode using PIPELINE and STAGE values
        from SESSTRANSFORMATIONINST.

        PIPELINE index — transformations with different PIPELINE numbers
                         run in truly parallel streams.
        STAGE    index — lower stage runs first within the same pipeline.

        Labelling convention:
          Single pipeline  : exec_order = stage number  (e.g. 0, 1, 2)
          Multiple pipelines: exec_order = pipe_letter + stage  (e.g. a0, b2)
        """
        transforms = self.mapping.transformations
        pipeline_groups = defaultdict(list)

        for sti in sess_transforms:
            pipeline_groups[sti.pipeline].append(sti)

        pipeline_ids = sorted(pipeline_groups.keys())
        is_parallel  = len(pipeline_ids) > 1

        print(f"  [EXEC] Pipelines detected: {pipeline_ids} "
              f"({'parallel' if is_parallel else 'single'})")

        for pid in pipeline_ids:
            stis = sorted(pipeline_groups[pid], key=lambda x: x.stage)
            for sti in stis:
                t = transforms.get(sti.instance_name)
                if not t:
                    continue

                if is_parallel:
                    pipe_letter  = chr(97 + pipeline_ids.index(pid))
                    t.exec_order = f"{pipe_letter}{sti.stage}"
                    t.exec_mode  = (f"Parallel — Pipeline {pid}, "
                                    f"Stage {sti.stage}")
                else:
                    t.exec_order = str(sti.stage)
                    t.exec_mode  = f"Sequential — Stage {sti.stage}"

                # Flag barriers and repartition points
                if t.trans_type in ("Aggregator", "Sorter"):
                    t.exec_mode += " [BARRIER]"
                if sti.is_repartition:
                    t.exec_mode += f" [REPARTITION:{sti.partition_type}]"

        # Print summary
        for inst in sorted(transforms,
                            key=lambda x: transforms[x].exec_order):
            t = transforms[inst]
            if t.exec_order:
                print(f"         [{t.exec_order:>5}] "
                      f"{inst:<40} ({t.trans_type})"
                      f"  →  {t.exec_mode}")

    # ── Fallback: topology-based annotation ──────────────────────

    def annotate_topology(self) -> None:
        """Fallback when no session data is available."""
        if not nx.is_directed_acyclic_graph(self._instance_graph):
            print("  [WARN] Cycle detected — skipping exec order annotation")
            return

        topo       = list(nx.topological_sort(self._instance_graph))
        transforms = self.mapping.transformations
        split_nodes = {n for n in self._instance_graph.nodes
                       if self._instance_graph.out_degree(n) > 1}
        merge_nodes = {n for n in self._instance_graph.nodes
                       if self._instance_graph.in_degree(n) > 1}

        level          = 0
        parallel_level = None
        branch_count   = {}
        in_parallel    = set()

        for inst in topo:
            if inst in split_nodes:
                level += 1
                parallel_level = level + 1
                branch_count[parallel_level] = 0
                _ord, _mode = str(level), "Sequential (Split Point)"
            elif inst in merge_nodes:
                in_parallel.discard(inst)
                parallel_level = None
                level += 1
                _ord, _mode = str(level), "Sequential (Merge Point)"
            elif parallel_level is not None:
                in_parallel.add(inst)
                branch_count[parallel_level] = \
                    branch_count.get(parallel_level, 0) + 1
                letter = chr(96 + branch_count[parallel_level])
                _ord   = f"{parallel_level}{letter}"
                _mode  = "Parallel Branch"
            else:
                level += 1
                _ord, _mode = str(level), "Sequential"

            if inst in transforms:
                t = transforms[inst]
                t.exec_order = _ord
                t.exec_mode  = _mode
                if t.trans_type in ("Aggregator", "Sorter"):
                    t.exec_mode += " [BARRIER]"

    # ── Column-level trace ────────────────────────────────────────

    def topological_order(self) -> list:
        try:
            return list(nx.topological_sort(self._instance_graph))
        except nx.NetworkXUnfeasible:
            return list(self._instance_graph.nodes)

    def trace_column(self, tgt_instance: str,
                     tgt_field: str) -> list:
        node = f"{tgt_instance}.{tgt_field}"
        if node not in self.G:
            return []
        edges   = []
        queue   = [node]
        visited = set()
        while queue:
            n = queue.pop()
            if n in visited:
                continue
            visited.add(n)
            for pred in self.G.predecessors(n):
                e = self.G.edges[pred, n]
                edges.append((e["from_instance"], e["from_field"],
                               e["to_instance"],  e["to_field"]))
                queue.append(pred)
        return edges

    def build_lineage_rows(self) -> list:
        rows = []
        seq  = 1
        for tgt_name, tgt_cols in self.mapping.targets.items():
            for col in tgt_cols:
                edges = self.trace_column(tgt_name, col.name)
                if not edges:
                    rows.append(LineageRow(
                        seq=str(seq),
                        source_table="UNCONNECTED",
                        source_column="-", source_datatype="-",
                        sq_column="-",
                        transformation_name="-",
                        transformation_type="-",
                        transformation_logic="-",
                        exec_order="-", exec_mode="-",
                        target_column=col.name,
                        target_datatype=col.datatype,
                        target_table=tgt_name,
                        remarks="Unconnected target port"
                    ))
                    seq += 1
                    continue

                src_table = src_col = sq_col = "-"
                trans_name = trans_type = "-"
                trans_logic = "Direct Mapping"
                exec_order = exec_mode = "-"

                for (fi, ff, ti, tf) in edges:
                    if fi in self.mapping.sources:
                        src_table = fi
                        src_col   = ff
                    t = self.mapping.transformations.get(ti)
                    if t and t.trans_type == "Source Qualifier":
                        sq_col = tf
                    if t and t.trans_type not in ("Source Qualifier",) \
                            and ti not in self.mapping.sources:
                        trans_name  = t.name
                        trans_type  = t.trans_type
                        exec_order  = t.exec_order
                        exec_mode   = t.exec_mode
                        trans_logic = self._build_logic(t, tf)

                rows.append(LineageRow(
                    seq               = str(seq),
                    source_table      = src_table,
                    source_column     = src_col,
                    source_datatype   = self._src_dt(src_table, src_col),
                    sq_column         = sq_col if sq_col != "-" else src_col,
                    transformation_name  = trans_name,
                    transformation_type  = trans_type,
                    transformation_logic = trans_logic,
                    exec_order        = exec_order,
                    exec_mode         = exec_mode,
                    target_column     = col.name,
                    target_datatype   = col.datatype,
                    target_table      = tgt_name,
                    remarks           = "PK" if col.key_type == "PRIMARY KEY"
                                        else ""
                ))
                seq += 1
        return rows

    def _build_logic(self, t: Transformation, port: str) -> str:
        tt = t.trans_type
        if tt == "Expression":
            for p in t.ports:
                if p.name == port and p.expression:
                    return p.expression
            return "Pass-through"
        elif tt == "Lookup Procedure":
            table = t.attributes.get("Lookup table name", "")
            cond  = t.attributes.get("Lookup condition", "")
            cache = t.attributes.get("Lookup cache persistent", "NO")
            return (f"LOOKUP: {table} "
                    f"ON ({cond}) "
                    f"RETURN {port} [Cache={cache}]")
        elif tt == "Filter":
            return "FILTER: " + t.attributes.get("Filter Condition", "")
        elif tt == "Update Strategy":
            return "UPD_STRATEGY: " + \
                   t.attributes.get("Update Strategy Expression", "")
        elif tt == "Maplet":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"MAPLET [{t.name}]: {p.expression}"
            return f"MAPLET [{t.name}]: Pass-through"
        elif tt == "Source Qualifier":
            sql = t.attributes.get("Sql Query", "")
            flt = t.attributes.get("Source Filter", "")
            return f"SQ_SQL: {sql}" if sql else \
                   f"SQ_FILTER: {flt}" if flt else "Pass-through"
        elif tt == "Router":
            return "ROUTER: " + "; ".join(
                f"{k}={v}" for k, v in t.attributes.items()
                if "Group" in k)
        elif tt == "Aggregator":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"AGG: {p.expression}"
            return "Pass-through"
        elif tt == "Sequence":
            return (f"SEQUENCE: start={t.attributes.get('Start Value','1')} "
                    f"incr={t.attributes.get('Increment By','1')}")
        return "Pass-through"

    def _src_dt(self, src_table: str, src_col: str) -> str:
        for c in self.mapping.sources.get(src_table, []):
            if c.name == src_col:
                return c.datatype
        return ""


# ══════════════════════════════════════════════════════════════════════
# MODULE 4 — STTM EXCEL GENERATOR
# ══════════════════════════════════════════════════════════════════════

class STTMGenerator:
    """
    Multi-sheet Excel STTM:
      Sheet 1  — Mapping Summary
      Sheet 2  — Column Lineage          (one per mapping)
      Sheet 3  — Transformation Detail   (one per mapping)
      Sheet 4  — Lookup Reference        (one per mapping)
      Sheet 5  — Source Qualifier SQL    (one per mapping)
      Sheet 6  — Maplet Detail           (one per mapping)
      Sheet 7  — Execution Order         (one per mapping)
      Sheet 8  — Session Pipeline Detail (one per session)
    """

    HDR_FILL  = PatternFill("solid", fgColor="0D1B2A")
    SUB_FILL  = PatternFill("solid", fgColor="1A3A5C")
    ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")
    GOLD_FILL = PatternFill("solid", fgColor="F4A621")
    GRN_FILL  = PatternFill("solid", fgColor="22C55E")
    TEAL_FILL = PatternFill("solid", fgColor="00B4D8")
    RED_FILL  = PatternFill("solid", fgColor="EF4444")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
    PAR_FILL  = PatternFill("solid", fgColor="D6E4F0")

    HDR_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    BODY_FONT = Font(name="Calibri", bold=False, color="000000", size=10)
    BOLD_FONT = Font(name="Calibri", bold=True,  color="000000", size=10)
    THIN  = Side(style="thin",   color="CCCCCC")
    BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def _ws(self, title: str):
        ws = self.wb.create_sheet(title=title[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _h(self, ws, r, c, v, w=None, fill=None):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.HDR_FONT
        cell.fill      = fill or self.HDR_FILL
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center", wrap_text=True)
        cell.border    = self.BORD
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[r].height = 32
        return cell

    def _c(self, ws, r, c, v, bold=False, shade=False,
           wrap=False, align="left", fill=None):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.BOLD_FONT if bold else self.BODY_FONT
        cell.fill      = fill or (self.ALT_FILL if shade else self.WHT_FILL)
        cell.alignment = Alignment(horizontal=align,
                                   vertical="center", wrap_text=wrap)
        cell.border    = self.BORD
        ws.row_dimensions[r].height = 40 if wrap else 22
        return cell

    # ── Sheet 1: Summary ─────────────────────────────────────────

    def _summary(self, parser: InformaticaParser, graph_map: dict):
        ws = self._ws("1_Mapping_Summary")
        hdrs = ["Workflow", "Execution Order", "Session",
                "Mapping", "Source Tables", "Target Tables",
                "Transforms", "Maplets", "Pipelines", "Description"]
        wids = [28, 18, 35, 35, 30, 30, 12, 25, 12, 45]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)

        row = 2
        for wf_name, wf in parser.workflows.items():
            for i, sess_name in enumerate(wf.sessions):
                sess = parser.sessions.get(sess_name)
                if not sess:
                    continue
                m = parser.mappings.get(sess.mapping_name)
                if not m:
                    continue
                g  = graph_map.get(m.name)
                pids = sorted(set(
                    st.pipeline for st in sess.sess_transforms))
                shade = (row % 2 == 0)
                vals = [
                    wf_name,
                    wf.execution_order[i]
                        if i < len(wf.execution_order) else str(i + 1),
                    sess_name,
                    m.name,
                    ", ".join(m.sources.keys()),
                    ", ".join(m.targets.keys()),
                    str(len(m.transformations)),
                    ", ".join(m.maplets_used) or "None",
                    f"{len(pids)} pipeline(s): {pids}",
                    m.description,
                ]
                for c, v in enumerate(vals, 1):
                    self._c(ws, row, c, v, shade=shade, wrap=True)
                row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 2: Column Lineage ───────────────────────────────────

    def _lineage(self, mapping: Mapping,
                 rows: list, name: str):
        ws   = self._ws(name)
        hdrs = ["#", "Source Table", "Source Column", "Src DT",
                "SQ Column", "Transform Name", "Transform Type",
                "Transformation Logic", "Exec Order", "Exec Mode",
                "Target Column", "Tgt DT", "Target Table", "Remarks"]
        wids = [5, 22, 22, 12, 20, 25, 20, 55, 12, 30, 22, 12, 22, 18]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)

        for i, r in enumerate(rows, 2):
            s = (i % 2 == 0)
            mf = None
            if "Parallel" in r.exec_mode:
                mf = self.GOLD_FILL
            elif "Merge" in r.exec_mode or "BARRIER" in r.exec_mode:
                mf = self.GRN_FILL
            elif "Split" in r.exec_mode or "REPARTITION" in r.exec_mode:
                mf = self.TEAL_FILL

            vals = [r.seq, r.source_table, r.source_column,
                    r.source_datatype, r.sq_column,
                    r.transformation_name, r.transformation_type,
                    r.transformation_logic, r.exec_order, r.exec_mode,
                    r.target_column, r.target_datatype,
                    r.target_table, r.remarks]
            for c, v in enumerate(vals, 1):
                self._c(ws, i, c, v, shade=s,
                        wrap=(c == 8),
                        fill=mf if c == 10 else None)
        ws.freeze_panes = "B2"

    # ── Sheet 3: Transformation Detail ───────────────────────────

    def _transforms(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Transform Name", "Type", "Port Name",
                "Port Type", "Datatype", "Expression / Logic",
                "Exec Order", "Exec Mode"]
        wids = [28, 22, 25, 16, 14, 55, 12, 30]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)

        row = 2
        for t in mapping.transformations.values():
            # Sub-header per transformation
            for c in range(1, 9):
                self._c(ws, row, c, "", fill=self.SUB_FILL)
            ws.cell(row=row, column=1).value = t.name
            ws.cell(row=row, column=1).font  = self.HDR_FONT
            ws.cell(row=row, column=2).value = t.trans_type
            ws.cell(row=row, column=2).font  = self.HDR_FONT
            ws.cell(row=row, column=7).value = t.exec_order
            ws.cell(row=row, column=7).font  = self.HDR_FONT
            ws.cell(row=row, column=8).value = t.exec_mode
            ws.cell(row=row, column=8).font  = self.HDR_FONT
            ws.row_dimensions[row].height = 22
            row += 1

            # Attributes
            for k, v in t.attributes.items():
                self._c(ws, row, 1, "")
                self._c(ws, row, 2, "")
                self._c(ws, row, 3, f"[{k}]", bold=True)
                self._c(ws, row, 4, "ATTRIBUTE")
                self._c(ws, row, 5, "")
                self._c(ws, row, 6, v, wrap=True, fill=self.PAR_FILL)
                self._c(ws, row, 7, "")
                self._c(ws, row, 8, "")
                row += 1

            # Ports
            for p in t.ports:
                s = (row % 2 == 0)
                self._c(ws, row, 1, t.name,       shade=s)
                self._c(ws, row, 2, t.trans_type,  shade=s)
                self._c(ws, row, 3, p.name,        shade=s)
                self._c(ws, row, 4, p.port_type,   shade=s)
                self._c(ws, row, 5, p.datatype,    shade=s)
                self._c(ws, row, 6,
                        p.expression or "Pass-through",
                        shade=s, wrap=True)
                self._c(ws, row, 7, t.exec_order,  shade=s)
                self._c(ws, row, 8, t.exec_mode,   shade=s)
                row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 4: Lookup Reference ─────────────────────────────────

    def _lookups(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Lookup Name", "Lookup Table", "Lookup Condition",
                "Return Ports", "Cache Type", "Mapping"]
        wids = [28, 30, 45, 25, 14, 28]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)
        row = 2
        for t in mapping.transformations.values():
            if t.trans_type == "Lookup Procedure":
                s = (row % 2 == 0)
                rets = [p.name for p in t.ports
                        if "OUTPUT" in p.port_type]
                self._c(ws, row, 1, t.name, shade=s)
                self._c(ws, row, 2,
                        t.attributes.get("Lookup table name", ""),
                        shade=s)
                self._c(ws, row, 3,
                        t.attributes.get("Lookup condition", ""),
                        shade=s, wrap=True)
                self._c(ws, row, 4, ", ".join(rets), shade=s)
                self._c(ws, row, 5,
                        t.attributes.get(
                            "Lookup cache persistent", "NO"),
                        shade=s)
                self._c(ws, row, 6, mapping.name, shade=s)
                row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 5: SQ SQL ───────────────────────────────────────────

    def _sq_sql(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["SQ Name", "SQL Override", "Source Filter",
                "Sorted Ports", "Mapping"]
        wids = [28, 65, 40, 12, 28]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)
        row = 2
        for t in mapping.transformations.values():
            if t.trans_type == "Source Qualifier":
                s = (row % 2 == 0)
                self._c(ws, row, 1, t.name,  shade=s)
                self._c(ws, row, 2,
                        t.attributes.get("Sql Query", ""),
                        shade=s, wrap=True)
                self._c(ws, row, 3,
                        t.attributes.get("Source Filter", ""),
                        shade=s, wrap=True)
                self._c(ws, row, 4,
                        t.attributes.get(
                            "Number Of Sorted Ports", "0"),
                        shade=s)
                self._c(ws, row, 5, mapping.name, shade=s)
                row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 6: Maplet Detail ────────────────────────────────────

    def _maplets(self, mapping: Mapping,
                 all_maplets: dict, name: str):
        ws   = self._ws(name)
        hdrs = ["Maplet Name", "Port Name", "Port Type",
                "Expression", "Purpose", "Mapping"]
        wids = [28, 22, 14, 55, 30, 28]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)
        row = 2
        for mname in mapping.maplets_used:
            t = all_maplets.get(mname) or \
                mapping.transformations.get(mname)
            if not t:
                continue
            purpose = ""
            if "CRC"  in mname.upper(): purpose = "CRC Checksum for CDC"
            if "HASH" in mname.upper(): purpose = "Hash Key Generation"
            if "SEQ"  in mname.upper(): purpose = "Sequence / Surrogate Key"
            for p in t.ports:
                s = (row % 2 == 0)
                self._c(ws, row, 1, t.name,    shade=s)
                self._c(ws, row, 2, p.name,     shade=s)
                self._c(ws, row, 3, p.port_type,shade=s)
                self._c(ws, row, 4,
                        p.expression or "Pass-through",
                        shade=s, wrap=True)
                self._c(ws, row, 5, purpose,    shade=s)
                self._c(ws, row, 6, mapping.name,shade=s)
                row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 7: Execution Order ──────────────────────────────────

    def _exec_order(self, mapping: Mapping,
                    graph: LineageGraph, name: str):
        ws   = self._ws(name)
        hdrs = ["Exec Order", "Instance Name", "Transform Type",
                "Exec Mode", "In-Degree", "Out-Degree",
                "Pipeline", "Stage", "Notes"]
        wids = [12, 35, 22, 35, 10, 10, 10, 8, 45]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)

        topo = graph.topological_order()
        ts   = mapping.transformations
        row  = 2

        for inst in topo:
            t    = ts.get(inst)
            in_d = graph._instance_graph.in_degree(inst)
            out_d= graph._instance_graph.out_degree(inst)
            s    = (row % 2 == 0)

            note = ""
            if out_d > 1:
                note = f"⚡ SPLIT → {out_d} parallel branches"
            if in_d > 1:
                note = f"🔀 MERGE ← {in_d} branches"
            if t and t.trans_type in ("Aggregator", "Sorter"):
                note = "🛑 BARRIER — all upstream rows must complete first"
            if t and "REPARTITION" in (t.exec_mode or ""):
                note += " | ♻ Data redistributed across partitions"

            mf = None
            if t:
                if "Parallel" in (t.exec_mode or ""):
                    mf = self.GOLD_FILL
                elif "Merge"  in (t.exec_mode or ""):
                    mf = self.GRN_FILL
                elif "Split"  in (t.exec_mode or "") or \
                     "REPARTITION" in (t.exec_mode or ""):
                    mf = self.TEAL_FILL

            # Extract pipeline / stage from exec_order string
            eo = t.exec_order if t else "?"
            pipeline_display = ""
            stage_display    = ""
            if t and t.exec_mode:
                pm = re.search(r'Pipeline (\d+)', t.exec_mode)
                sm = re.search(r'Stage (\d+)',    t.exec_mode)
                if pm: pipeline_display = pm.group(1)
                if sm: stage_display    = sm.group(1)

            self._c(ws, row, 1, eo,   shade=s, bold=True, align="center")
            self._c(ws, row, 2, inst, shade=s)
            self._c(ws, row, 3,
                    t.trans_type if t else "Source/Target",
                    shade=s)
            self._c(ws, row, 4,
                    t.exec_mode if t else "Sequential",
                    shade=s, fill=mf)
            self._c(ws, row, 5, str(in_d),  shade=s, align="center")
            self._c(ws, row, 6, str(out_d), shade=s, align="center")
            self._c(ws, row, 7, pipeline_display, shade=s, align="center")
            self._c(ws, row, 8, stage_display,    shade=s, align="center")
            self._c(ws, row, 9, note, shade=s, wrap=True)
            row += 1
        ws.freeze_panes = "A2"

    # ── Sheet 8: Session Pipeline Detail (NEW) ────────────────────

    def _session_pipeline(self, session: Session, name: str):
        """
        New sheet — shows PIPELINE + STAGE detail direct from the
        SESSION tag, exactly as stored in the XML.
        """
        ws   = self._ws(name)
        hdrs = ["Pipeline", "Stage", "Instance Name",
                "Transform Name", "Transform Type",
                "Is Repartition", "Partition Type",
                "Exec Mode (derived)"]
        wids = [10, 8, 35, 35, 22, 14, 20, 35]
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._h(ws, 1, c, h, w)

        stis = sorted(session.sess_transforms,
                      key=lambda x: (x.pipeline, x.stage))
        pipeline_ids = sorted(set(st.pipeline for st in stis))
        is_parallel  = len(pipeline_ids) > 1

        row = 2
        for sti in stis:
            s = (row % 2 == 0)

            if is_parallel:
                letter   = chr(97 + pipeline_ids.index(sti.pipeline))
                exec_mode = (f"Parallel — Pipeline {sti.pipeline}, "
                             f"Stage {sti.stage}")
                mf = self.GOLD_FILL
            else:
                exec_mode = f"Sequential — Stage {sti.stage}"
                mf = None

            if sti.is_repartition:
                exec_mode += f" [REPARTITION:{sti.partition_type}]"
                mf = self.TEAL_FILL

            self._c(ws, row, 1, str(sti.pipeline), shade=s,
                    align="center", bold=True,
                    fill=self.GOLD_FILL if is_parallel else None)
            self._c(ws, row, 2, str(sti.stage),    shade=s, align="center")
            self._c(ws, row, 3, sti.instance_name, shade=s)
            self._c(ws, row, 4, sti.transformation_name, shade=s)
            self._c(ws, row, 5, sti.transformation_type, shade=s)
            self._c(ws, row, 6, "YES" if sti.is_repartition else "NO",
                    shade=s, align="center",
                    fill=self.TEAL_FILL if sti.is_repartition else None)
            self._c(ws, row, 7, sti.partition_type, shade=s)
            self._c(ws, row, 8, exec_mode,          shade=s, fill=mf)
            row += 1
        ws.freeze_panes = "A2"

    # ── Main ──────────────────────────────────────────────────────

    def generate(self, parser: InformaticaParser,
                 graph_map: dict):
        print(f"\n[STTM] Writing: {self.output_path}")

        self._summary(parser, graph_map)

        for m_name, mapping in parser.mappings.items():
            graph = graph_map[m_name]
            rows  = graph.build_lineage_rows()
            safe  = re.sub(r'[^A-Za-z0-9_]', '_', m_name)[:18]

            self._lineage   (mapping, rows,          f"{safe}_Lineage")
            self._transforms(mapping,                f"{safe}_Transforms")
            self._lookups   (mapping,                f"{safe}_Lookups")
            self._sq_sql    (mapping,                f"{safe}_SQ_SQL")
            self._maplets   (mapping, parser.maplets,f"{safe}_Maplets")
            self._exec_order(mapping, graph,         f"{safe}_ExecOrder")

        # Session pipeline sheets
        for s_name, sess in parser.sessions.items():
            if sess.sess_transforms:
                safe = re.sub(r'[^A-Za-z0-9_]', '_', s_name)[:18]
                self._session_pipeline(sess, f"{safe}_Pipeline")

        self.wb.save(self.output_path)
        print(f"  [OK  ] Saved → {self.output_path}")


# ══════════════════════════════════════════════════════════════════════
# PIPELINE RUNNER
# ══════════════════════════════════════════════════════════════════════

def run_pipeline(xml_path: str,
                 par_files: list,
                 output_path: str) -> None:
    print()
    print("═" * 68)
    print("  LineageIQ  —  Informatica PowerCenter 10.5.8 Lineage Pipeline")
    print("═" * 68)
    print(f"  Input XML : {xml_path}")
    print(f"  Param files: {par_files}")
    print(f"  Output    : {output_path}")
    print("═" * 68)

    # Step 1 ── Resolve parameters
    print("\n[STEP 1] Resolving parameters...")
    resolver = ParamResolver()
    for pf in par_files:
        resolver.load_file(pf)
    resolver.report()

    # Step 2 ── Parse XML
    print("\n[STEP 2] Parsing Informatica XML (PC 10.5.8)...")
    parser = InformaticaParser(resolver)
    parser.parse(xml_path)

    # Step 3 ── Build lineage graphs + annotate execution order
    print("\n[STEP 3] Building lineage graphs & execution order...")
    graph_map: dict = {}

    for m_name, mapping in parser.mappings.items():
        print(f"\n  [GRAPH] {m_name}")
        graph = LineageGraph(mapping)

        # Find matching session to get PIPELINE + STAGE data
        sess = next(
            (s for s in parser.sessions.values()
             if s.mapping_name == m_name),
            None
        )
        if sess and sess.sess_transforms:
            # FIX 3: Use PIPELINE/STAGE for accurate exec order
            graph.annotate_from_session(sess.sess_transforms)
        else:
            # Fallback: topological sort
            print("  [INFO] No session data — using topology fallback")
            graph.annotate_topology()

        graph_map[m_name] = graph

    # Step 4 ── Generate STTM Excel
    print("\n[STEP 4] Generating STTM Excel...")
    gen = STTMGenerator(output_path)
    gen.generate(parser, graph_map)

    print()
    print("═" * 68)
    print(f"  ✅  LineageIQ complete!")
    print(f"  📄  Output file : {Path(output_path).resolve()}")
    print("═" * 68)


# ══════════════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════════════
"""
╔══════════════════════════════════════════════════════════════════════╗
║  INPUT / OUTPUT FILE NAMING GUIDE                                    ║
╠══════════════════════════════════════════════════════════════════════╣
║                                                                      ║
║  INPUT XML (Informatica export from PowerCenter Designer):           ║
║    wf_TCOM_RR.xml                                                    ║
║    wf_RRDW_DLV_CUSTOMER.xml                                          ║
║    m_TSY2_TPR_AM01.xml          ← single mapping export              ║
║                                                                      ║
║  INPUT PAR (Unix parameter file):                                    ║
║    params_prod.par                                                   ║
║    TCOM_EXP_prod.par                                                 ║
║    RRDW_params_uat.par                                               ║
║                                                                      ║
║  OUTPUT XLSX (auto-named if not specified):                          ║
║    LineageIQ_STTM_wf_TCOM_RR.xlsx                                    ║
║    LineageIQ_STTM_m_TSY2_TPR_AM01.xlsx                               ║
║                                                                      ║
║  USAGE EXAMPLES:                                                     ║
║    python lineageiq.py --xml wf_TCOM_RR.xml --par params_prod.par    ║
║    python lineageiq.py --xml wf_TCOM_RR.xml \                        ║
║                        --par params_prod.par \                        ║
║                        --out MySTTM.xlsx                              ║
║    python lineageiq.py   (uses defaults below)                       ║
╚══════════════════════════════════════════════════════════════════════╝
"""

if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="LineageIQ — Informatica Data Lineage Extractor"
    )
    ap.add_argument("--xml", default=None,
                    help="Path to Informatica XML export file")
    ap.add_argument("--par", nargs="*", default=None,
                    help="Path(s) to .par parameter file(s)")
    ap.add_argument("--out", default=None,
                    help="Output Excel file path")
    args = ap.parse_args()

    # ── Default values (edit these for your environment) ──────────
    BASE     = Path(__file__).parent

    xml_path = args.xml or str(BASE / "wf_TCOM_RR.xml")
    par_files= args.par or [str(BASE / "params_prod.par")]

    if args.out:
        out_path = args.out
    else:
        # Auto-name output from input XML stem
        xml_stem = Path(xml_path).stem
        out_path = str(BASE / f"LineageIQ_STTM_{xml_stem}.xlsx")

    run_pipeline(
        xml_path    = xml_path,
        par_files   = par_files,
        output_path = out_path,
    )
