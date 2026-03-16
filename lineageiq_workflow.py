"""
╔══════════════════════════════════════════════════════════════════════╗
║   LineageIQ — WORKFLOW PARSER                                         ║
║   Script : lineageiq_workflow.py                                      ║
║   Purpose: Parse a workflow XML and generate a full STTM Excel        ║
║            covering the ENTIRE workflow — all sessions, worklets,     ║
║            mappings and their execution order                         ║
║                                                                       ║
║   Sheets produced                                                     ║
║     1_WorkflowSummary   — Workflow → Worklets → Sessions → Mappings  ║
║     2_SessionDetail     — All sessions with PIPELINE/STAGE detail     ║
║     3_ExecutionSequence — Complete execution order of the workflow    ║
║     Per Mapping (one set per mapping found):                          ║
║       <M>_MappingParse  — Full 5-section mapping parse               ║
║       <M>_SourceDetail  — Source schema/table/columns/SQ             ║
║       <M>_TargetDetail  — Target schema/table/columns                ║
║       <M>_Transforms    — All transforms in exec order + logic       ║
║       <M>_ColFlowMap    — SQ col → T1 → T2 → … → Target col         ║
║       <M>_Lookups       — Lookup reference                            ║
║       <M>_Maplets       — Maplet details                              ║
║       <M>_ExecOrder     — Pipeline/stage/parallel detection           ║
║                                                                       ║
║   Works WITH or WITHOUT a parameter file                              ║
║                                                                       ║
║   USAGE                                                               ║
║     # With parameter file                                             ║
║     python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║                                   --par params_prod.par               ║
║                                                                       ║
║     # Without parameter file                                          ║
║     python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║                                                                       ║
║     # Custom output                                                   ║
║     python lineageiq_workflow.py --xml wf_TCOM_RR.xml                ║
║                                   --par params_prod.par               ║
║                                   --out MyWorkflow_STTM.xlsx          ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict

import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════

@dataclass
class ColumnDef:
    name: str; datatype: str; precision: str
    scale: str; nullable: str; key_type: str = "NOT A KEY"


@dataclass
class TransformPort:
    name: str; port_type: str; datatype: str
    precision: str; scale: str; expression: str = ""


@dataclass
class Transformation:
    name: str; trans_type: str
    ports:      list = field(default_factory=list)
    attributes: dict = field(default_factory=dict)
    exec_order: str  = ""
    exec_mode:  str  = ""


@dataclass
class Connector:
    from_instance: str; from_field: str
    to_instance:   str; to_field:   str


@dataclass
class Mapping:
    name: str; description: str; folder_name: str = ""
    sources:         dict = field(default_factory=dict)
    targets:         dict = field(default_factory=dict)
    source_schemas:  dict = field(default_factory=dict)
    target_schemas:  dict = field(default_factory=dict)
    transformations: dict = field(default_factory=dict)
    connectors:      list = field(default_factory=list)
    maplets_used:    list = field(default_factory=list)


@dataclass
class SessionTransformInst:
    instance_name:       str
    transformation_name: str
    transformation_type: str
    pipeline:            int
    stage:               int
    is_repartition:      bool
    partition_type:      str


@dataclass
class Session:
    name:            str
    mapping_name:    str
    description:     str  = ""
    src_connection:  str  = ""
    tgt_connection:  str  = ""
    is_reusable:     bool = False
    sess_transforms: list = field(default_factory=list)


@dataclass
class WorkflowTask:
    name:       str
    task_type:  str   # Session / Worklet / Command / Email / Start / End
    task_ref:   str   # referenced task name
    condition:  str = ""


@dataclass
class Workflow:
    name:            str
    description:     str        = ""
    server:          str        = ""
    tasks:           list       = field(default_factory=list)
    execution_order: list       = field(default_factory=list)
    worklets:        dict       = field(default_factory=dict)  # name → [task_names]
    sessions:        list       = field(default_factory=list)  # flat ordered list


# ══════════════════════════════════════════════════════════════════════
# MODULE 1 — PARAMETER RESOLVER
# ══════════════════════════════════════════════════════════════════════

class ParamResolver:
    def __init__(self, par_files: list = None):
        self.params:     dict = {}
        self.unresolved: list = []
        self.has_params: bool = False
        for pf in (par_files or []):
            self._load(pf)

    def _load(self, par_path: str) -> None:
        p = Path(par_path)
        if not p.exists():
            print(f"  [WARN] Par file not found: {par_path} — continuing")
            return
        print(f"  [PAR ] Loading: {par_path}")
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip()
                    if k not in self.params:
                        self.params[k] = v
                        print(f"         {k} = {v}")
        self.has_params = True

    def resolve(self, text: str) -> str:
        if not text or not self.has_params:
            return text

        def sub(m):
            pn = m.group(0)
            if pn in self.params:
                return self.params[pn]
            self.unresolved.append(pn)
            return f"[UNRESOLVED:{pn}]"

        return re.sub(r'\$\$[A-Z0-9_]+', sub, text)

    def report(self) -> None:
        if not self.has_params:
            print("  [INFO] No parameter file — $$PARAMS kept as-is")
            return
        seen   = set()
        unique = [x for x in self.unresolved
                  if not (x in seen or seen.add(x))]
        print(f"  [WARN] Unresolved: {unique}" if unique
              else "  [OK  ] All parameters resolved")


# ══════════════════════════════════════════════════════════════════════
# MODULE 2 — WORKFLOW XML PARSER
# ══════════════════════════════════════════════════════════════════════

class WorkflowParser:
    """
    Parses Informatica PowerCenter workflow XML (PC 10.5.x).
    Extracts: workflows, worklets, sessions, mappings, maplets.
    """

    def __init__(self, resolver: ParamResolver):
        self.resolver         = resolver
        self.workflows: dict  = {}
        self.sessions:  dict  = {}
        self.mappings:  dict  = {}
        self.maplets:   dict  = {}
        self._g_sources: dict = {}
        self._g_targets: dict = {}
        self._src_owner: dict = {}
        self._tgt_owner: dict = {}

    def parse(self, xml_path: str) -> None:
        print(f"\n[PARSE] {xml_path}")
        xp = ET.XMLParser(load_dtd=False, no_network=True,
                           resolve_entities=False, recover=True)
        tree = ET.parse(xml_path, xp)
        root = tree.getroot()
        folder = root.find(".//FOLDER") or root

        self._parse_sources(folder)
        self._parse_targets(folder)
        self._parse_maplets(folder)
        self._parse_mappings(folder)
        self._parse_sessions(folder)
        self._parse_workflows(folder)

        print(f"  [OK  ] workflows={len(self.workflows)}  "
              f"sessions={len(self.sessions)}  "
              f"mappings={len(self.mappings)}  "
              f"maplets={len(self.maplets)}")

    # ── Sources ───────────────────────────────────────────────────

    def _parse_sources(self, f) -> None:
        for src in f.findall(".//SOURCE"):
            n     = src.get("NAME", "")
            owner = self.resolver.resolve(
                src.get("OWNERNAME", src.get("DBDNAME", "")))
            cols  = [ColumnDef(
                sf.get("NAME",""), sf.get("DATATYPE",""),
                sf.get("PRECISION",""), sf.get("SCALE","0"),
                sf.get("NULLABLE",""))
                for sf in src.findall("SOURCEFIELD")]
            self._g_sources[n] = cols
            self._src_owner[n] = owner
            print(f"  [SRC ] {n}  schema={owner}  ({len(cols)} cols)")

    # ── Targets ───────────────────────────────────────────────────

    def _parse_targets(self, f) -> None:
        for tgt in f.findall(".//TARGET"):
            n     = tgt.get("NAME", "")
            owner = self.resolver.resolve(
                tgt.get("OWNERNAME", tgt.get("DBDNAME", "")))
            cols  = [ColumnDef(
                tf.get("NAME",""), tf.get("DATATYPE",""),
                tf.get("PRECISION",""), tf.get("SCALE","0"),
                tf.get("NULLABLE",""), tf.get("KEYTYPE","NOT A KEY"))
                for tf in tgt.findall("TARGETFIELD")]
            self._g_targets[n] = cols
            self._tgt_owner[n] = owner
            print(f"  [TGT ] {n}  schema={owner}  ({len(cols)} cols)")

    # ── Maplets ───────────────────────────────────────────────────

    def _parse_maplets(self, f) -> None:
        for ml in f.findall(".//MAPPLET"):
            n = ml.get("NAME", "")
            t = Transformation(name=n, trans_type="Maplet",
                               attributes={"description":
                                           ml.get("DESCRIPTION","")})
            for inner in ml.findall(".//TRANSFORMATION"):
                for tf in inner.findall("TRANSFORMFIELD"):
                    t.ports.append(TransformPort(
                        tf.get("NAME",""), tf.get("PORTTYPE",""),
                        tf.get("DATATYPE",""), tf.get("PRECISION",""),
                        tf.get("SCALE","0"),
                        self.resolver.resolve(tf.get("EXPRESSION",""))))
            self.maplets[n] = t
            print(f"  [MLET] {n}")

    # ── Mappings ──────────────────────────────────────────────────

    def _parse_mappings(self, f) -> None:
        for mp in f.findall(".//MAPPING"):
            m = Mapping(name=mp.get("NAME","UNKNOWN"),
                        description=mp.get("DESCRIPTION",""))
            for inst in mp.findall("INSTANCE"):
                itype = inst.get("TYPE","")
                iname = inst.get("TRANSFORMATION_NAME",
                                 inst.get("NAME",""))
                if itype == "SOURCE" and iname in self._g_sources:
                    m.sources[iname]        = self._g_sources[iname]
                    m.source_schemas[iname] = self._src_owner.get(iname,"")
                elif itype == "TARGET" and iname in self._g_targets:
                    m.targets[iname]        = self._g_targets[iname]
                    m.target_schemas[iname] = self._tgt_owner.get(iname,"")

            for trans in mp.findall("TRANSFORMATION"):
                t = self._parse_trans(trans)
                m.transformations[t.name] = t
                if t.trans_type == "Maplet":
                    m.maplets_used.append(t.name)

            for inst in mp.findall("INSTANCE"):
                if inst.get("TRANSFORMATION_TYPE") == "Maplet":
                    mname = inst.get("TRANSFORMATION_NAME","")
                    if mname in self.maplets \
                            and mname not in m.transformations:
                        m.transformations[mname] = self.maplets[mname]
                        if mname not in m.maplets_used:
                            m.maplets_used.append(mname)

            for conn in mp.findall("CONNECTOR"):
                m.connectors.append(Connector(
                    conn.get("FROMINSTANCE",""),
                    conn.get("FROMFIELD",""),
                    conn.get("TOINSTANCE",""),
                    conn.get("TOFIELD","")))

            self.mappings[m.name] = m
            print(f"  [MAP ] {m.name}  "
                  f"src={list(m.sources)} "
                  f"tgt={list(m.targets)} "
                  f"trans={len(m.transformations)}")

    def _parse_trans(self, el) -> Transformation:
        t = Transformation(el.get("NAME",""), el.get("TYPE",""))
        for tf in el.findall("TRANSFORMFIELD"):
            t.ports.append(TransformPort(
                tf.get("NAME",""), tf.get("PORTTYPE",""),
                tf.get("DATATYPE",""), tf.get("PRECISION",""),
                tf.get("SCALE","0"),
                self.resolver.resolve(tf.get("EXPRESSION",""))))
        for ta in el.findall("TABLEATTRIBUTE"):
            t.attributes[ta.get("NAME","")] = \
                self.resolver.resolve(ta.get("VALUE",""))
        return t

    # ── Sessions — FIX for PC 10.5.8 ─────────────────────────────

    def _parse_sessions(self, f) -> None:
        for sess in f.findall(".//SESSION"):
            # PC 10.x: MAPPINGNAME is direct attribute
            mapping_name = sess.get("MAPPINGNAME", "")
            # PC 9.x fallback
            if not mapping_name:
                for a in sess.findall("ATTRIBUTE"):
                    if "Mapping name" in a.get("NAME",""):
                        mapping_name = self.resolver.resolve(
                            a.get("VALUE",""))
                        break

            src_conn = tgt_conn = ""
            sti_list = []

            for sti in sess.findall("SESSTRANSFORMATIONINST"):
                ttype = sti.get("TRANSFORMATIONTYPE","")
                sname = sti.get("SINSTANCENAME","")
                pipeline = int(sti.get("PIPELINE","0"))
                stage    = int(sti.get("STAGE","0"))

                for attr in sti.findall("ATTRIBUTE"):
                    aname = attr.get("NAME","")
                    aval  = self.resolver.resolve(attr.get("VALUE",""))
                    if aname == "Table Name Prefix" \
                            and ttype == "Target Definition":
                        tgt_conn = aval
                    if aname == "$Source connection value":
                        src_conn = aval
                    if aname == "$Target connection value" \
                            and not tgt_conn:
                        tgt_conn = aval

                if ttype == "Source Definition" and not src_conn:
                    src_conn = sname

                sti_list.append(SessionTransformInst(
                    sname,
                    sti.get("TRANSFORMATIONNAME", sname),
                    ttype, pipeline, stage,
                    sti.get("ISREPARTITIONPOINT","NO") == "YES",
                    sti.get("PARTITIONTYPE","")))

            s = Session(
                name           = sess.get("NAME",""),
                mapping_name   = mapping_name,
                description    = sess.get("DESCRIPTION",""),
                src_connection = src_conn,
                tgt_connection = tgt_conn,
                is_reusable    = sess.get("REUSABLE","NO") == "YES",
                sess_transforms= sti_list,
            )
            self.sessions[s.name] = s
            print(f"  [SESS] {s.name}  →  {s.mapping_name}")

    # ── Workflows + Worklets ──────────────────────────────────────

    def _parse_workflows(self, f) -> None:
        # Collect worklet definitions
        worklet_defs: dict = {}
        for wkl in f.findall(".//WORKLET"):
            wname = wkl.get("NAME","")
            tasks = [(ti.get("NAME",""), ti.get("TASKTYPE",""),
                      ti.get("TASKNAME",""))
                     for ti in wkl.findall("TASKINSTANCE")]
            worklet_defs[wname] = tasks

        for wf in f.findall(".//WORKFLOW"):
            w = Workflow(
                name        = wf.get("NAME",""),
                description = wf.get("DESCRIPTION",""),
                server      = wf.get("SERVERNAME",""),
            )

            # Collect all task instances
            for ti in wf.findall("TASKINSTANCE"):
                w.tasks.append(WorkflowTask(
                    name      = ti.get("NAME",""),
                    task_type = ti.get("TASKTYPE",""),
                    task_ref  = ti.get("TASKNAME",""),
                ))

            # Flatten session list in execution order
            links: dict = {}
            for lnk in wf.findall("LINK"):
                ft = lnk.get("FROMTASK","")
                tt = lnk.get("TOTASK","")
                cond = lnk.get("CONDITION","")
                links[ft] = (tt, cond)

            # Build ordered task execution sequence
            cur     = "start"
            visited = set()
            while cur in links and cur not in visited:
                visited.add(cur)
                nxt, cond = links[cur]
                if nxt and nxt != "start":
                    w.execution_order.append((nxt, cond))
                cur = nxt

            # Expand worklets to get session list
            for task in w.tasks:
                if task.task_type == "Worklet" \
                        and task.task_ref in worklet_defs:
                    for (tname, ttype, tref) in \
                            worklet_defs[task.task_ref]:
                        if ttype == "Session":
                            w.sessions.append(tname)
                elif task.task_type == "Session":
                    w.sessions.append(task.name)

            w.worklets = {k: [t[0] for t in v]
                          for k, v in worklet_defs.items()}
            self.workflows[w.name] = w
            print(f"  [WF  ] {w.name}  sessions={w.sessions}")


# ══════════════════════════════════════════════════════════════════════
# MODULE 3 — LINEAGE GRAPH (reused from mapping parser)
# ══════════════════════════════════════════════════════════════════════

class LineageGraph:
    def __init__(self, mapping: Mapping):
        self.mapping = mapping
        self.G  = nx.DiGraph()
        self.IG = nx.DiGraph()
        self._build()

    def _build(self) -> None:
        for conn in self.mapping.connectors:
            s = f"{conn.from_instance}.{conn.from_field}"
            t = f"{conn.to_instance}.{conn.to_field}"
            self.G.add_edge(s, t,
                from_instance=conn.from_instance,
                from_field=conn.from_field,
                to_instance=conn.to_instance,
                to_field=conn.to_field)
            if not self.IG.has_edge(conn.from_instance,
                                    conn.to_instance):
                self.IG.add_edge(conn.from_instance,
                                 conn.to_instance)

    def annotate_from_session(self,
                               sti_list: list) -> None:
        """Use PIPELINE/STAGE from session for accurate exec order."""
        ts            = self.mapping.transformations
        pipe_groups   = defaultdict(list)
        for sti in sti_list:
            pipe_groups[sti.pipeline].append(sti)

        pipe_ids    = sorted(pipe_groups.keys())
        is_parallel = len(pipe_ids) > 1

        for pid in pipe_ids:
            stis = sorted(pipe_groups[pid], key=lambda x: x.stage)
            for sti in stis:
                t = ts.get(sti.instance_name)
                if not t:
                    continue
                if is_parallel:
                    letter       = chr(97 + pipe_ids.index(pid))
                    t.exec_order = f"{letter}{sti.stage}"
                    t.exec_mode  = (f"Parallel — "
                                    f"Pipeline {pid}, Stage {sti.stage}")
                else:
                    t.exec_order = str(sti.stage)
                    t.exec_mode  = f"Sequential — Stage {sti.stage}"
                if t.trans_type in ("Aggregator", "Sorter"):
                    t.exec_mode += " [BARRIER]"
                if sti.is_repartition:
                    t.exec_mode += (
                        f" [REPARTITION:{sti.partition_type}]")

    def annotate_topology(self) -> None:
        """Fallback when no session data."""
        try:
            topo = list(nx.topological_sort(self.IG))
        except Exception:
            topo = list(self.IG.nodes)
        ts    = self.mapping.transformations
        split = {n for n in self.IG.nodes
                 if self.IG.out_degree(n) > 1}
        merge = {n for n in self.IG.nodes
                 if self.IG.in_degree(n) > 1}
        level = 0; pl = None; bc: dict = {}; ip: set = set()
        for inst in topo:
            if inst in split:
                level += 1; pl = level + 1; bc[pl] = 0
                _o, _m = str(level), "Sequential — Split Point"
            elif inst in merge:
                ip.discard(inst); pl = None; level += 1
                _o, _m = str(level), "Sequential — Merge Point"
            elif pl is not None:
                ip.add(inst); bc[pl] = bc.get(pl, 0) + 1
                let = chr(96 + bc[pl])
                _o, _m = f"{pl}{let}", f"Parallel Branch {let.upper()}"
            else:
                level += 1; _o, _m = str(level), "Sequential"
            if inst in ts:
                t = ts[inst]; t.exec_order = _o; t.exec_mode = _m
                if t.trans_type in ("Aggregator","Sorter"):
                    t.exec_mode += " [BARRIER]"

    def topo_order(self) -> list:
        try:
            return list(nx.topological_sort(self.IG))
        except Exception:
            return list(self.IG.nodes)

    def trace_back(self, tgt: str, col: str) -> list:
        node = f"{tgt}.{col}"
        if node not in self.G:
            return []
        edges, q, vis = [], [node], set()
        while q:
            n = q.pop()
            if n in vis: continue
            vis.add(n)
            for p in self.G.predecessors(n):
                edges.append(dict(self.G.edges[p, n]))
                q.append(p)
        return edges

    def build_logic(self, t: Transformation, port: str) -> str:
        tt = t.trans_type
        if tt == "Expression":
            for p in t.ports:
                if p.name == port and p.expression:
                    return p.expression
            return "Pass-through"
        elif tt == "Lookup Procedure":
            return (f"LOOKUP: {t.attributes.get('Lookup table name','')}  "
                    f"ON ({t.attributes.get('Lookup condition','')})  "
                    f"RETURN {port}  "
                    f"[Cache={t.attributes.get('Lookup cache persistent','NO')}]")
        elif tt == "Filter":
            return "FILTER: " + t.attributes.get("Filter Condition","")
        elif tt == "Update Strategy":
            return "UPD_STRATEGY: " + \
                   t.attributes.get("Update Strategy Expression","")
        elif tt == "Maplet":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"MAPLET [{t.name}]: {p.expression}"
            return f"MAPLET [{t.name}]: Pass-through"
        elif tt == "Source Qualifier":
            sql = t.attributes.get("Sql Query","")
            flt = t.attributes.get("Source Filter","")
            return (f"SQ_SQL: {sql}" if sql
                    else f"SQ_FILTER: {flt}" if flt
                    else "Pass-through")
        elif tt == "Router":
            return "ROUTER: " + "; ".join(
                f"{k}={v}" for k, v in t.attributes.items()
                if "Group" in k)
        elif tt == "Aggregator":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"AGG: {p.expression}"
            return "Pass-through"
        elif tt == "Joiner":
            return (f"JOINER  condition="
                    f"{t.attributes.get('Join Condition','')}  "
                    f"type={t.attributes.get('Join Type','')}")
        elif tt == "Sequence":
            return (f"SEQUENCE  start="
                    f"{t.attributes.get('Start Value','1')}  "
                    f"incr={t.attributes.get('Increment By','1')}")
        elif tt == "Stored Procedure":
            return "SP: " + t.attributes.get("Stored Procedure Name","")
        return "Pass-through"


# ══════════════════════════════════════════════════════════════════════
# MODULE 4 — EXCEL WRITER (Workflow-level)
# ══════════════════════════════════════════════════════════════════════

class WorkflowExcelWriter:

    C = {
        "navy":"0D1B2A","teal":"065A82","purple":"4A235A",
        "dark":"2C3E50","brown":"7B341E","green_d":"1A6B5A",
        "sub":"1A3A5C","gold":"F4A621","green":"22C55E",
        "cyan":"00B4D8","red":"EF4444","white":"FFFFFF",
        "alt_a":"E8F4FD","alt_b":"E8FDF4","alt_c":"F5EEF8",
        "alt_d":"FEF9E7","alt_e":"FEF5E4","pale":"D6E4F0",
    }
    BORD  = Border(**{s: Side(style="thin", color="CCCCCC")
                      for s in ("left","right","top","bottom")})
    HDR_F = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    BOD_F = Font(name="Calibri", bold=False, color="000000", size=10)
    BLD_F = Font(name="Calibri", bold=True,  color="000000", size=10)

    def __init__(self, output_path: str):
        self.out = output_path
        self.wb  = Workbook()
        self.wb.remove(self.wb.active)

    def _fill(self, c: str) -> PatternFill:
        return PatternFill("solid", fgColor=self.C.get(c, c))

    def _ws(self, t: str):
        ws = self.wb.create_sheet(title=t[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _hc(self, ws, r, c, v, w=None, clr="navy"):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.HDR_F
        cell.fill      = self._fill(clr)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center",
                                   wrap_text=True)
        cell.border    = self.BORD
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[r].height = 30
        return cell

    def _dc(self, ws, r, c, v, shade=False, wrap=False,
            align="left", clr=None, bold=False):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.BLD_F if bold else self.BOD_F
        cell.fill      = (self._fill(clr) if clr
                          else self._fill("alt_a") if shade
                          else self._fill("white"))
        cell.alignment = Alignment(horizontal=align,
                                   vertical="top" if wrap else "center",
                                   wrap_text=wrap)
        cell.border    = self.BORD
        ws.row_dimensions[r].height = 38 if wrap else 22
        return cell

    def _banner(self, ws, row, text, clr="navy", cols=14):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Calibri", bold=True,
                           color="FFFFFF", size=11)
        c.fill      = self._fill(clr)
        c.alignment = Alignment(horizontal="left",
                                vertical="center")
        c.border    = self.BORD
        ws.row_dimensions[row].height = 26
        return row + 1

    def _ch(self, ws, row, hdrs, wids, clr="sub"):
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._hc(ws, row, c, h, w, clr)
        return row + 1

    # ─────────────────────────────────────────────────────────────
    # SHEET 1: Workflow Summary
    # ─────────────────────────────────────────────────────────────

    def sheet_wf_summary(self, parser: WorkflowParser):
        ws   = self._ws("1_WorkflowSummary")
        hdrs = ["Workflow", "Server", "Description",
                "Task Name", "Task Type",
                "Execution Order", "Link Condition",
                "Session Name", "Mapping Name",
                "Source Tables", "Target Tables",
                "Transforms", "Pipelines"]
        wids = [28, 20, 30, 30, 16, 16, 30, 30, 35, 30, 30, 12, 12]
        self._ch(ws, 1, hdrs, wids, "navy")
        row = 2

        for wf_name, wf in parser.workflows.items():
            # Build execution order lookup
            eo_map = {task: (i+1, cond)
                      for i, (task, cond) in
                      enumerate(wf.execution_order)}

            for task in wf.tasks:
                if task.task_type in ("Start", "End", "start"):
                    continue
                sess = parser.sessions.get(task.name)
                m    = parser.mappings.get(
                    sess.mapping_name if sess else "", None)

                eo_num, cond = eo_map.get(task.name, ("?", ""))

                pipe_ids = sorted(set(
                    st.pipeline for st in
                    (sess.sess_transforms if sess else [])))

                shade = (row % 2 == 0)
                vals  = [
                    wf_name, wf.server, wf.description,
                    task.name, task.task_type,
                    str(eo_num), cond,
                    task.name if task.task_type == "Session" else "",
                    sess.mapping_name if sess else "",
                    ", ".join(m.sources.keys()) if m else "",
                    ", ".join(m.targets.keys()) if m else "",
                    str(len(m.transformations)) if m else "",
                    str(pipe_ids) if pipe_ids else "1",
                ]
                for c, v in enumerate(vals, 1):
                    self._dc(ws, row, c, v, shade=shade, wrap=(c==3))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET 2: Session Detail
    # ─────────────────────────────────────────────────────────────

    def sheet_session_detail(self, parser: WorkflowParser):
        ws   = self._ws("2_SessionDetail")
        hdrs = ["Session Name", "Mapping Name", "Description",
                "Src Connection", "Tgt Connection",
                "Pipeline", "Stage",
                "Instance Name", "Transform Type",
                "Is Repartition", "Partition Type",
                "Exec Mode (derived)"]
        wids = [30, 35, 30, 22, 22, 10, 8, 30, 22, 14, 18, 35]
        self._ch(ws, 1, hdrs, wids, "teal")
        row = 2

        for s_name, sess in parser.sessions.items():
            stis = sorted(sess.sess_transforms,
                          key=lambda x: (x.pipeline, x.stage))
            if not stis:
                shade = (row % 2 == 0)
                for c, v in enumerate([
                    s_name, sess.mapping_name, sess.description,
                    sess.src_connection, sess.tgt_connection,
                    "-","-","-","-","-","-","No session transform data"
                ], 1):
                    self._dc(ws, row, c, v, shade=shade)
                row += 1
                continue

            pipe_ids    = sorted(set(st.pipeline for st in stis))
            is_parallel = len(pipe_ids) > 1

            for sti in stis:
                shade = (row % 2 == 0)
                if is_parallel:
                    letter    = chr(97 + pipe_ids.index(sti.pipeline))
                    exec_mode = (f"Parallel — Pipeline {sti.pipeline},"
                                 f" Stage {sti.stage}")
                    pc = "gold"
                else:
                    exec_mode = f"Sequential — Stage {sti.stage}"
                    pc = None
                if sti.is_repartition:
                    exec_mode += f" [REPARTITION:{sti.partition_type}]"
                    pc = "cyan"

                for c, v in enumerate([
                    s_name, sess.mapping_name, sess.description,
                    sess.src_connection, sess.tgt_connection,
                    str(sti.pipeline), str(sti.stage),
                    sti.instance_name, sti.transformation_type,
                    "YES" if sti.is_repartition else "NO",
                    sti.partition_type, exec_mode
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             clr=(pc if c in (6, 12) else
                                  "cyan" if c == 10 and
                                  sti.is_repartition else None))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET 3: Execution Sequence
    # ─────────────────────────────────────────────────────────────

    def sheet_exec_sequence(self, parser: WorkflowParser):
        ws   = self._ws("3_ExecutionSequence")
        hdrs = ["Workflow", "Exec Step", "Task Name",
                "Task Type", "Link Condition",
                "Session", "Mapping",
                "Source Tables", "Target Tables",
                "Pipelines", "Notes"]
        wids = [28, 10, 30, 16, 35, 30, 35, 30, 30, 16, 40]
        self._ch(ws, 1, hdrs, wids, "dark")
        row = 2

        for wf_name, wf in parser.workflows.items():
            for step, (task_name, cond) in \
                    enumerate(wf.execution_order, 1):
                sess = parser.sessions.get(task_name)
                m    = parser.mappings.get(
                    sess.mapping_name if sess else "", None)

                pipe_ids = sorted(set(
                    st.pipeline for st in
                    (sess.sess_transforms if sess else [])))

                note = ""
                if len(pipe_ids) > 1:
                    note = f"⚡ {len(pipe_ids)} parallel pipelines"

                shade = (row % 2 == 0)
                for c, v in enumerate([
                    wf_name, str(step), task_name,
                    "Session" if sess else "Worklet/Other",
                    cond,
                    task_name if sess else "",
                    sess.mapping_name if sess else "",
                    ", ".join(m.sources.keys()) if m else "",
                    ", ".join(m.targets.keys()) if m else "",
                    str(pipe_ids) if pipe_ids else "[1]",
                    note
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             wrap=(c in (5, 11)))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # Per-mapping sheets (reused logic from mapping parser)
    # ─────────────────────────────────────────────────────────────

    def _eo_key(self, t: Transformation):
        eo = t.exec_order or "z99"
        l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
        n  = int(re.sub(r'[^0-9]', '', eo) or "0")
        return (l, n)

    def sheet_mapping_parse(self, mapping: Mapping,
                             graph: LineageGraph,
                             maplets: dict,
                             session: Session,
                             sheet_name: str):
        """5-section mapping parse sheet — same as mapping parser."""
        ws = self._ws(sheet_name)
        row = 1

        # ── A: Source ─────────────────────────────────────────────
        row = self._banner(ws, row,
            f"  A — SOURCE DETAILS     Mapping: {mapping.name}",
            "navy")
        row = self._ch(ws, row,
            ["Source Schema","Source Table","Source Column",
             "Datatype","Precision","Nullable",
             "SQ Name","SQ Column","SQL Override / Filter"],
            [22,25,25,14,10,10,25,25,65], "sub")

        for src_name, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src_name, "")
            sq_name, sq_sql, sq_map = "-", "-", {}
            for conn in mapping.connectors:
                if conn.from_instance == src_name:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_map[conn.from_field] = conn.to_field
                        sql = t.attributes.get("Sql Query","")
                        flt = t.attributes.get("Source Filter","")
                        sq_sql = sql or flt or "Default"
            for i, col in enumerate(cols):
                sq_c = sq_map.get(col.name, "-")
                for c, v in enumerate([
                    schema,src_name,col.name,col.datatype,
                    col.precision,col.nullable,
                    sq_name,sq_c,sq_sql
                ], 1):
                    self._dc(ws,row,c,v,shade=(i%2==0),wrap=(c==9))
                row += 1
        row += 1

        # ── B: Target ─────────────────────────────────────────────
        row = self._banner(ws, row, "  B — TARGET DETAILS", "teal")
        row = self._ch(ws, row,
            ["Target Schema","Target Table","Target Column",
             "Datatype","Precision","Scale","Nullable","Key Type"],
            [22,25,25,14,10,8,10,16], "teal")

        for tgt_name, cols in mapping.targets.items():
            schema = mapping.target_schemas.get(tgt_name, "")
            for i, col in enumerate(cols):
                pk = col.key_type == "PRIMARY KEY"
                for c, v in enumerate([
                    schema,tgt_name,col.name,col.datatype,
                    col.precision,col.scale,col.nullable,col.key_type
                ], 1):
                    self._dc(ws,row,c,v,shade=(i%2==0),
                             clr="gold" if pk else None)
                row += 1
        row += 1

        # ── C: Transformations ────────────────────────────────────
        row = self._banner(ws, row,
            "  C — TRANSFORMATION INVENTORY", "purple")
        row = self._ch(ws, row,
            ["Exec Order","Exec Mode","Transform Name",
             "Transform Type","Port Name","Port Type",
             "Expression / Logic","Attributes"],
            [12,32,28,22,25,16,58,58], "purple")

        for t in sorted(mapping.transformations.values(),
                        key=self._eo_key):
            attr_str = " | ".join(
                f"{k}: {v}" for k, v in t.attributes.items()
                if v and k != "description")
            em_c = ("gold"   if "Parallel"    in (t.exec_mode or "")
                    else "green" if "BARRIER"  in (t.exec_mode or "")
                    else "cyan"  if "REPARTITION" in (t.exec_mode or "")
                    else None)
            ports = t.ports or [TransformPort(
                "-","-","-","-","0")]
            for pi, p in enumerate(ports):
                s = (pi % 2 == 0)
                for c, v in enumerate([
                    t.exec_order if pi==0 else "",
                    t.exec_mode  if pi==0 else "",
                    t.name       if pi==0 else "",
                    t.trans_type if pi==0 else "",
                    p.name, p.port_type,
                    p.expression or "Pass-through",
                    attr_str     if pi==0 else "",
                ], 1):
                    self._dc(ws,row,c,v,shade=s,
                             wrap=(c in (7,8)),
                             clr=em_c if c==2 and pi==0 else None)
                row += 1
        row += 1

        # ── D: Column Flow Map ────────────────────────────────────
        row = self._banner(ws, row,
            "  D — COLUMN FLOW MAP  "
            "[ Source Column → SQ → Transforms → Target Column ]",
            "dark")
        row = self._ch(ws, row,
            ["#","Source Schema","Source Table","Source Column",
             "SQ Name","SQ Column",
             "Transformation Chain\n(Name|Type|Order|Mode)",
             "Logic at Each Step",
             "Target Schema","Target Table",
             "Target Column","Target DT","Key Type","Remarks"],
            [5,18,20,20,20,20,55,62,18,20,20,13,13,14], "dark")

        seq = 1
        for tgt_name, tgt_cols in mapping.targets.items():
            tgt_schema = mapping.target_schemas.get(tgt_name, "")
            for col in tgt_cols:
                edges = graph.trace_back(tgt_name, col.name)
                if not edges:
                    for c, v in enumerate([
                        str(seq),"-","-","-","-","-",
                        "UNCONNECTED","-",
                        tgt_schema,tgt_name,col.name,
                        col.datatype,col.key_type,
                        "No upstream connection"
                    ], 1):
                        self._dc(ws, row, c, v)
                    row += 1; seq += 1
                    continue

                src_table = src_col = src_schema = "-"
                sq_name   = sq_col  = "-"
                chain, logic = [], []

                def inst_ord(inst):
                    t = mapping.transformations.get(inst)
                    if t and t.exec_order:
                        eo = t.exec_order
                        l = ord(eo[0])-96 if eo[0].isalpha() else 0
                        n = int(re.sub(r'[^0-9]','',eo) or "0")
                        return (l,n)
                    return (0,0) if inst in mapping.sources else (99,99)

                seen, ordered = set(), []
                for e in edges:
                    for inst in [e["from_instance"],e["to_instance"]]:
                        if inst not in seen:
                            seen.add(inst); ordered.append(inst)
                ordered.sort(key=inst_ord)

                for inst in ordered:
                    t = mapping.transformations.get(inst)
                    field = "-"
                    for e in edges:
                        if e["from_instance"]==inst:
                            field=e["from_field"]; break
                        if e["to_instance"]==inst:
                            field=e["to_field"]
                    if inst in mapping.sources:
                        src_table=inst; src_col=field
                        src_schema=mapping.source_schemas.get(inst,"")
                    elif t and t.trans_type=="Source Qualifier":
                        sq_name=inst; sq_col=field
                    elif inst not in mapping.targets and t:
                        chain.append(
                            f"{inst}  [{t.trans_type}]  "
                            f"Ord:{t.exec_order}  {t.exec_mode}")
                        lg = graph.build_logic(t, field)
                        if lg and lg != "-":
                            logic.append(f"▶ {inst}: {lg}")

                tc = "\n→ ".join(chain) if chain else "Direct"
                lc = "\n".join(logic)   if logic else "Pass-through"
                shade = (seq % 2 == 0)
                lns   = max(len(chain), len(logic), 1)

                for c, v in enumerate([
                    str(seq),
                    src_schema, src_table, src_col,
                    sq_name, sq_col, tc, lc,
                    tgt_schema, tgt_name,
                    col.name, col.datatype, col.key_type,
                    "PK" if col.key_type=="PRIMARY KEY" else ""
                ], 1):
                    cell = ws.cell(row=row, column=c,
                                   value=str(v) if v else "")
                    cell.font = self.BOD_F
                    cell.fill = (self._fill("alt_d") if shade
                                 else self._fill("white"))
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top",
                        wrap_text=True)
                    cell.border = self.BORD
                ws.row_dimensions[row].height = max(38, lns*18)
                row += 1; seq += 1
        row += 1

        # ── E: Roadmap ────────────────────────────────────────────
        row = self._banner(ws, row,
            "  E — FLOW DIAGRAM ROADMAP", "brown")
        for line in self._roadmap(mapping, graph, session):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=14)
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = Font(name="Consolas", size=9,
                             bold=("STEP" in line),
                             color=("FFFFFF"
                                    if line.startswith("  STEP")
                                    else "1A1A1A"))
            cell.fill = (self._fill("sub")
                         if line.startswith("  STEP")
                         else self._fill("alt_a")
                         if "→" in line
                         else self._fill("white"))
            cell.alignment = Alignment(horizontal="left",
                                        vertical="center")
            cell.border = self.BORD
            ws.row_dimensions[row].height = 15
            row += 1

        ws.freeze_panes = "A4"

    def _roadmap(self, mapping: Mapping,
                  graph: LineageGraph,
                  session: Session) -> list:
        lines = []
        W = 110
        lines.append("═" * W)
        lines.append(
            f"  WORKFLOW FLOW DIAGRAM ROADMAP   Mapping: {mapping.name}")
        lines.append("═" * W); lines.append("")

        lines.append("  STEP 1 — SOURCES")
        lines.append("  " + "─"*60)
        for src, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src, "?")
            sq_name, sq_sql = "-", "-"
            for conn in mapping.connectors:
                if conn.from_instance == src:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_sql  = t.attributes.get(
                            "Sql Query","")[:100]
                        break
            lines.append(
                f"   {schema}.{src}  ({len(cols)} cols)  SQ:{sq_name}")
            if sq_sql:
                lines.append(f"   SQL: {sq_sql}")
        lines.append("")

        lines.append("  STEP 2 — TRANSFORMATION CHAIN")
        lines.append("  " + "─"*60)
        for t in sorted(mapping.transformations.values(),
                        key=self._eo_key):
            split = graph.IG.out_degree(t.name) > 1
            merge = graph.IG.in_degree(t.name) > 1
            flag  = (" ⚡SPLIT" if split else "") + \
                    (" 🔀MERGE" if merge else "") + \
                    (" 🛑BARRIER"
                     if t.trans_type in ("Aggregator","Sorter")
                     else "")
            lines.append(
                f"   [{t.exec_order:>5}]  {t.name:<35}  "
                f"({t.trans_type})  {t.exec_mode}{flag}")
        lines.append("")

        lines.append("  STEP 3 — TARGETS")
        lines.append("  " + "─"*60)
        for tgt, cols in mapping.targets.items():
            schema  = mapping.target_schemas.get(tgt, "?")
            pk_cols = [c.name for c in cols
                       if c.key_type == "PRIMARY KEY"]
            lines.append(
                f"   {schema}.{tgt}  ({len(cols)} cols)"
                f"  PK:{', '.join(pk_cols)}")
        lines.append("")

        lines.append("  ADDITIONAL STEPS — VISUAL DIAGRAM")
        lines.append("  " + "─"*60)
        for s, t, d in [
            ("STEP A","Graph edges from Column Flow Map",
             "Each row = SQ_Col → Transform1 → Transform2 → Target_Col"),
            ("STEP B","Node types & colours",
             "SOURCE=blue  SQ=teal  EXPR=orange  LKP=purple  "
             "FILTER=red  TARGET=green"),
            ("STEP C","Parallel branches → swim-lanes",
             "Group by PIPELINE from Session Detail sheet"),
            ("STEP D","Tools: graphviz / D3.js / draw.io",
             "Export edges as CSV → import to Lucidchart or draw.io"),
            ("STEP E","Cross-mapping TPR→TT→DDM",
             "Match target cols of this mapping to source cols of "
             "next mapping in workflow execution sequence"),
        ]:
            lines.append(f"   {s}  {t}")
            lines.append(f"         • {d}")
        lines.append(""); lines.append("═"*W)
        return lines

    def sheet_lookups(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Lookup Name","Lookup Table","Lookup Condition",
                "Return Cols","Cache","Exec Order","Exec Mode"]
        wids = [28,32,45,28,14,12,32]
        self._ch(ws, 1, hdrs, wids, "teal")
        row = 2
        for t in mapping.transformations.values():
            if t.trans_type == "Lookup Procedure":
                s = (row % 2 == 0)
                rets = [p.name for p in t.ports
                        if "OUTPUT" in p.port_type]
                for c, v in enumerate([
                    t.name,
                    t.attributes.get("Lookup table name",""),
                    t.attributes.get("Lookup condition",""),
                    ", ".join(rets),
                    t.attributes.get("Lookup cache persistent","NO"),
                    t.exec_order, t.exec_mode
                ], 1):
                    self._dc(ws, row, c, v, shade=s, wrap=(c==3))
                row += 1
        ws.freeze_panes = "A2"

    def sheet_maplets(self, mapping: Mapping,
                       maplets: dict, name: str):
        ws   = self._ws(name)
        hdrs = ["Maplet Name","Port Name","Port Type",
                "Datatype","Expression","Purpose"]
        wids = [28,22,14,14,60,30]
        self._ch(ws, 1, hdrs, wids, "purple")
        row = 2
        for mname in mapping.maplets_used:
            t = maplets.get(mname) or \
                mapping.transformations.get(mname)
            if not t: continue
            purpose = ("CRC Checksum" if "CRC"  in mname.upper()
                       else "Hash Key" if "HASH" in mname.upper()
                       else "Sequence" if "SEQ"  in mname.upper()
                       else "")
            for p in t.ports:
                s = (row % 2 == 0)
                for c, v in enumerate([
                    t.name, p.name, p.port_type, p.datatype,
                    p.expression or "Pass-through", purpose
                ], 1):
                    self._dc(ws, row, c, v, shade=s, wrap=(c==5))
                row += 1
        ws.freeze_panes = "A2"

    def sheet_exec_order(self, mapping: Mapping,
                          graph: LineageGraph, name: str):
        ws   = self._ws(name)
        hdrs = ["Exec Order","Instance","Transform Type",
                "Exec Mode","In-Deg","Out-Deg","Notes"]
        wids = [12,35,22,38,10,10,50]
        self._ch(ws, 1, hdrs, wids, "navy")
        row  = 2
        for inst in graph.topo_order():
            t    = mapping.transformations.get(inst)
            in_d = graph.IG.in_degree(inst)
            out_d= graph.IG.out_degree(inst)
            s    = (row % 2 == 0)
            note = ""
            if out_d > 1: note = f"⚡ SPLIT → {out_d} branches"
            if in_d  > 1: note = f"🔀 MERGE ← {in_d} branches"
            if t and t.trans_type in ("Aggregator","Sorter"):
                note += " 🛑 BARRIER"
            em_c = ("gold"  if t and "Parallel" in (t.exec_mode or "")
                    else "green" if t and "Merge"  in (t.exec_mode or "")
                    else "cyan"  if t and "Split"  in (t.exec_mode or "")
                    else None)
            for c, v in enumerate([
                t.exec_order if t else "?",
                inst,
                t.trans_type if t else "Source/Target",
                t.exec_mode  if t else "Sequential",
                str(in_d), str(out_d), note
            ], 1):
                self._dc(ws, row, c, v, shade=s,
                         align="center" if c in (1,5,6) else "left",
                         clr=em_c if c==4 else None, wrap=(c==7))
            row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # GENERATE — all sheets
    # ─────────────────────────────────────────────────────────────

    def generate(self, parser: WorkflowParser) -> None:
        print(f"\n[WRITE] {self.out}")

        # Workflow-level sheets
        self.sheet_wf_summary(parser)
        self.sheet_session_detail(parser)
        self.sheet_exec_sequence(parser)

        # Per-mapping sheets
        for m_name, mapping in parser.mappings.items():
            # Match session
            sess = next(
                (s for s in parser.sessions.values()
                 if s.mapping_name == m_name), None)

            # Build graph
            graph = LineageGraph(mapping)
            if sess and sess.sess_transforms:
                graph.annotate_from_session(sess.sess_transforms)
            else:
                graph.annotate_topology()

            safe = re.sub(r'[^A-Za-z0-9_]', '_', m_name)[:14]

            self.sheet_mapping_parse(
                mapping, graph, parser.maplets, sess,
                f"{safe}_MappingParse")
            self.sheet_lookups (mapping,                f"{safe}_Lookups")
            self.sheet_maplets (mapping, parser.maplets,f"{safe}_Maplets")
            self.sheet_exec_order(mapping, graph,       f"{safe}_ExecOrder")

            print(f"  [OK  ] Mapping sheets: {m_name}")

        self.wb.save(self.out)
        print(f"\n  ✅  Saved → {Path(self.out).resolve()}")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="LineageIQ Workflow Parser — "
                    "parses an Informatica workflow XML end-to-end",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
--------
  # With parameter file
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml --par params_prod.par

  # Without parameter file  ($$PARAMS kept as-is in output)
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml

  # Multiple par files  (PROD takes priority)
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml
                                --par params_prod.par params_uat.par

  # Custom output name
  python lineageiq_workflow.py --xml wf_TCOM_RR.xml --par params_prod.par
                                --out WF_TCOM_RR_STTM.xlsx
        """
    )
    ap.add_argument("--xml", required=True,
                    help="Path to Informatica workflow XML file")
    ap.add_argument("--par", nargs="*", default=None,
                    help="Path(s) to .par file(s) — optional")
    ap.add_argument("--out", default=None,
                    help="Output Excel path — auto-named if not given")
    args = ap.parse_args()

    xml_path  = args.xml
    par_files = args.par or []
    xml_stem  = Path(xml_path).stem
    out_path  = args.out or f"LineageIQ_Workflow_{xml_stem}.xlsx"

    print()
    print("═" * 65)
    print("  LineageIQ — WORKFLOW PARSER")
    print("═" * 65)
    print(f"  XML  : {xml_path}")
    print(f"  PAR  : {par_files or 'None ($$PARAMS kept as-is)'}")
    print(f"  OUT  : {out_path}")
    print("═" * 65)

    resolver = ParamResolver(par_files)
    resolver.report()

    parser = WorkflowParser(resolver)
    parser.parse(xml_path)

    if not parser.workflows and not parser.mappings:
        print("\n  [ERROR] No workflow or mapping tags found in XML.")
        return

    writer = WorkflowExcelWriter(out_path)
    writer.generate(parser)

    print()
    print("═" * 65)
    print(f"  ✅  Done!  Output → {Path(out_path).resolve()}")
    print("═" * 65)


if __name__ == "__main__":
    main()
