"""
╔══════════════════════════════════════════════════════════════════════╗
║  LineageIQ — MAPPING PARSER  v4                                      ║
║  Script : lineageiq_mapping.py                                       ║
║                                                                      ║
║  Informatica Maplet Architecture (correctly handled):                ║
║  ┌─ FOLDER ──────────────────────────────────────────────────────┐  ║
║  │  <MAPPLET NAME="ep_CRC_GEN">          ← global definition     │  ║
║  │    <TRANSFORMATION TYPE="Expression"> ← internal transforms   │  ║
║  │      <TRANSFORMFIELD .../>            ← ports + expressions   │  ║
║  │  <MAPPING NAME="m_TPR_CUST_BAL">                              │  ║
║  │    <TRANSFORMATION TYPE="Maplet"      ← inline maplet (rare)  │  ║
║  │      NAME="ep_CRC_GEN" REUSABLE="NO"> ← maplet defined inside │  ║
║  │    <INSTANCE TYPE="TRANSFORMATION"    ← maplet used as step   │  ║
║  │      TRANSFORMATION_TYPE="Maplet"     ← via INSTANCE ref      │  ║
║  │      TRANSFORMATION_NAME="ep_CRC_GEN">                        │  ║
║  │    <CONNECTOR FROMINSTANCE="EXP_DERIVE" FROMFIELD="CUST_ID"   │  ║
║  │               TOINSTANCE="ep_CRC_GEN"  TOFIELD="IN_CUST_ID"/> │  ║
║  │    <CONNECTOR FROMINSTANCE="ep_CRC_GEN" FROMFIELD="OUT_CRC"   │  ║
║  │               TOINSTANCE="FIL_ACTIVE"  TOFIELD="CRC_VAL"/>    │  ║
║                                                                      ║
║  Key maplet behaviours captured:                                     ║
║    1. Global MAPPLET tag  → parsed into maplet registry              ║
║    2. Inline TRANSFORMATION TYPE="Maplet" inside MAPPING             ║
║       → parsed directly, merged into mapping transformations         ║
║    3. INSTANCE TRANSFORMATION_TYPE="Maplet"                          ║
║       → resolved from registry, injected into mapping                ║
║    4. CONNECTOR tags referencing maplet instance name                ║
║       → lineage graph traces through the maplet's ports              ║
║    5. Maplet internal expressions                                     ║
║       → captured verbatim on each OUTPUT port                        ║
║    6. Nested maplets (maplet inside maplet)                           ║
║       → recursive resolution with cycle detection                    ║
║                                                                      ║
║  Works WITH or WITHOUT a parameter file                              ║
║                                                                      ║
║  USAGE                                                               ║
║    python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║    python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║                                 --par params_prod.par                ║
║    python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml             ║
║                                 --par params_prod.par                ║
║                                 --out MySTTM.xlsx                    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import re
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict
from typing import Optional

import lxml.etree as ET
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════

@dataclass
class ColumnDef:
    name:      str
    datatype:  str
    precision: str
    scale:     str
    nullable:  str
    key_type:  str = "NOT A KEY"


@dataclass
class TransformPort:
    name:       str
    port_type:  str    # INPUT / OUTPUT / INPUT/OUTPUT / VARIABLE
    datatype:   str
    precision:  str
    scale:      str
    expression: str = ""


@dataclass
class MapletPort:
    """
    Represents a port declared on a MAPPLET boundary
    (MAPPLETINPUT / MAPPLETOUTPUT).
    These are the external-facing ports that CONNECTORs attach to.
    """
    name:      str
    direction: str   # INPUT / OUTPUT


@dataclass
class Transformation:
    name:       str
    trans_type: str
    ports:      list = field(default_factory=list)   # [TransformPort]
    attributes: dict = field(default_factory=dict)
    exec_order: str  = ""
    exec_mode:  str  = ""
    # Maplet-specific
    is_maplet:          bool = False
    maplet_input_ports: list = field(default_factory=list)   # [MapletPort]
    maplet_output_ports:list = field(default_factory=list)
    maplet_internals:   list = field(default_factory=list)   # [Transformation] internal transforms


@dataclass
class Connector:
    from_instance: str
    from_field:    str
    to_instance:   str
    to_field:      str


@dataclass
class Mapping:
    name:            str
    description:     str
    folder_name:     str = ""
    sources:         dict = field(default_factory=dict)
    targets:         dict = field(default_factory=dict)
    source_schemas:  dict = field(default_factory=dict)
    target_schemas:  dict = field(default_factory=dict)
    transformations: dict = field(default_factory=dict)
    connectors:      list = field(default_factory=list)
    # Separate maplet tracking
    maplet_instances: list = field(default_factory=list)  # names of maplets used


# ══════════════════════════════════════════════════════════════════════
# MODULE 1 — PARAMETER RESOLVER
# ══════════════════════════════════════════════════════════════════════

class ParamResolver:
    """
    Reads .par files and substitutes $$PARAM references.
    If no par file is provided $$PARAMS are kept as-is in output.
    """

    def __init__(self, par_files: list = None):
        self.params:     dict = {}
        self.unresolved: list = []
        self.has_params: bool = False
        for pf in (par_files or []):
            self._load(pf)

    def _load(self, par_path: str) -> None:
        p = Path(par_path)
        if not p.exists():
            print(f"  [WARN] Par file not found: {par_path}")
            return
        print(f"  [PAR ] Loading: {par_path}")
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip()
                    if k not in self.params:
                        self.params[k] = v
                        print(f"         {k} = {v}")
        self.has_params = True

    def resolve(self, text: str) -> str:
        if not text or not self.has_params:
            return text

        def sub(m):
            pn = m.group(0)
            if pn in self.params:
                return self.params[pn]
            self.unresolved.append(pn)
            return f"[UNRESOLVED:{pn}]"

        return re.sub(r'\$\$[A-Z0-9_]+', sub, text)

    def report(self) -> None:
        if not self.has_params:
            print("  [INFO] No parameter file — $$PARAMS kept as-is")
            return
        seen   = set()
        unique = [x for x in self.unresolved
                  if not (x in seen or seen.add(x))]
        print(f"  [WARN] Unresolved params: {unique}" if unique
              else "  [OK  ] All parameters resolved")


# ══════════════════════════════════════════════════════════════════════
# MODULE 2 — MAPPING XML PARSER  (full maplet support)
# ══════════════════════════════════════════════════════════════════════

class MappingParser:
    """
    Parses Informatica PowerCenter mapping XML.

    Maplet resolution strategy
    --------------------------
    Pass 1:  Parse all global MAPPLET tags at FOLDER level.
             Each MAPPLET contains one or more TRANSFORMATION tags
             (its internal logic) plus optional MAPPLETINPUT /
             MAPPLETOUTPUT boundary port declarations.

    Pass 2:  Parse each MAPPING:
             a. Scan TRANSFORMATION tags inside MAPPING.
                If TYPE="Maplet" → this is an inline maplet definition.
                  - Parse its ports and add to mapping.transformations.
                  - Also register in the global maplet registry so
                    other mappings can reference it.
             b. Scan INSTANCE tags.
                If TRANSFORMATION_TYPE="Maplet" → look up the maplet
                  by TRANSFORMATION_NAME in the registry.
                  - Inject the resolved Transformation object into
                    mapping.transformations under the INSTANCE NAME
                    (which may differ from TRANSFORMATION_NAME for
                    renamed instances).
                  - Record in mapping.maplet_instances.

    Pass 3:  Parse CONNECTOR tags.
             CONNECTORs reference instances by NAME (not by
             TRANSFORMATION_NAME). The lineage graph is built on
             instance names so traces through maplets work correctly.

    Nested maplets
    --------------
    If a MAPPLET's internal TRANSFORMATION references another MAPPLET
    (TYPE="Maplet"), resolve_maplet() is called recursively.
    A visited-set prevents infinite loops.
    """

    def __init__(self, resolver: ParamResolver):
        self.resolver          = resolver
        self.mappings:  dict   = {}    # name → Mapping
        self.maplets:   dict   = {}    # name → Transformation (global registry)
        self._g_sources: dict  = {}
        self._g_targets: dict  = {}
        self._src_owner: dict  = {}
        self._tgt_owner: dict  = {}

    # ── Entry point ───────────────────────────────────────────────

    def parse(self, xml_path: str) -> None:
        print(f"\n[PARSE] {xml_path}")
        xp = ET.XMLParser(load_dtd=False, no_network=True,
                           resolve_entities=False, recover=True)
        tree  = ET.parse(xml_path, xp)
        root  = tree.getroot()
        folder = (root.find(".//FOLDER")
                  if root.find(".//FOLDER") is not None
                  else root)

        # Pass 1 — global definitions
        self._parse_sources(folder)
        self._parse_targets(folder)
        self._parse_global_maplets(folder)

        # Pass 2 — mappings (includes inline maplet detection)
        self._parse_mappings(folder)

        # Report
        m_with_maplets = sum(
            1 for m in self.mappings.values()
            if m.maplet_instances)
        print(f"  [OK  ] {len(self.mappings)} mapping(s), "
              f"{len(self.maplets)} maplet(s) in registry, "
              f"{m_with_maplets} mapping(s) use maplet(s)")

    # ── Sources ───────────────────────────────────────────────────

    def _parse_sources(self, parent) -> None:
        for src in parent.findall(".//SOURCE"):
            n     = src.get("NAME", "")
            owner = self.resolver.resolve(
                src.get("OWNERNAME", src.get("DBDNAME", "")))
            cols  = [
                ColumnDef(sf.get("NAME",""), sf.get("DATATYPE",""),
                          sf.get("PRECISION",""), sf.get("SCALE","0"),
                          sf.get("NULLABLE",""))
                for sf in src.findall("SOURCEFIELD")
            ]
            self._g_sources[n] = cols
            self._src_owner[n] = owner
            print(f"  [SRC ] {n}  schema={owner}  ({len(cols)} cols)")

    # ── Targets ───────────────────────────────────────────────────

    def _parse_targets(self, parent) -> None:
        for tgt in parent.findall(".//TARGET"):
            n     = tgt.get("NAME", "")
            owner = self.resolver.resolve(
                tgt.get("OWNERNAME", tgt.get("DBDNAME", "")))
            cols  = [
                ColumnDef(tf.get("NAME",""), tf.get("DATATYPE",""),
                          tf.get("PRECISION",""), tf.get("SCALE","0"),
                          tf.get("NULLABLE",""), tf.get("KEYTYPE","NOT A KEY"))
                for tf in tgt.findall("TARGETFIELD")
            ]
            self._g_targets[n] = cols
            self._tgt_owner[n] = owner
            print(f"  [TGT ] {n}  schema={owner}  ({len(cols)} cols)")

    # ── Global MAPPLET definitions ────────────────────────────────

    def _parse_global_maplets(self, parent) -> None:
        """
        Parse all <MAPPLET> tags at FOLDER level.
        Each MAPPLET:
          - Has 1..N <TRANSFORMATION> children (its internal logic)
          - May have <MAPPLETINPUT> / <MAPPLETOUTPUT> boundary ports
          - Its external ports are the union of all INPUT ports of its
            internal transformations (for inputs) and all OUTPUT ports
            (for outputs) — or explicitly declared by MAPPLETINPUT/OUTPUT
        """
        for ml in parent.findall("MAPPLET"):
            t = self._build_maplet_from_element(ml)
            self.maplets[t.name] = t
            print(f"  [MLET] {t.name}  "
                  f"({len(t.maplet_input_ports)} in-ports, "
                  f"{len(t.maplet_output_ports)} out-ports, "
                  f"{len(t.maplet_internals)} internal transforms)")

    def _build_maplet_from_element(self, ml_el,
                                    visited: set = None) -> Transformation:
        """
        Build a Transformation object representing a maplet.
        visited: set of maplet names currently being resolved (cycle guard).
        """
        if visited is None:
            visited = set()

        name = ml_el.get("NAME", "")
        t = Transformation(
            name       = name,
            trans_type = "Maplet",
            is_maplet  = True,
            attributes = {"description": ml_el.get("DESCRIPTION", "")}
        )

        # ── Collect explicit MAPPLETINPUT / MAPPLETOUTPUT ports ───
        for mi in ml_el.findall("MAPPLETINPUT"):
            for mp in mi.findall("MPORTFIELD"):
                t.maplet_input_ports.append(
                    MapletPort(mp.get("NAME",""), "INPUT"))

        for mo in ml_el.findall("MAPPLETOUTPUT"):
            for mp in mo.findall("MPORTFIELD"):
                t.maplet_output_ports.append(
                    MapletPort(mp.get("NAME",""), "OUTPUT"))

        # ── Parse internal TRANSFORMATION tags ────────────────────
        for inner in ml_el.findall("TRANSFORMATION"):
            inner_type = inner.get("TYPE","")

            # Nested maplet reference inside maplet
            if inner_type == "Maplet":
                inner_name = inner.get("NAME","")
                if inner_name in visited:
                    print(f"  [WARN] Cycle detected: maplet {name} "
                          f"→ {inner_name} — skipping")
                    continue
                # Resolve from registry if already parsed
                if inner_name in self.maplets:
                    resolved = self.maplets[inner_name]
                else:
                    # Try to find the MAPPLET element and parse it
                    resolved = self._parse_inline_transform(inner)
                t.maplet_internals.append(resolved)
            else:
                inner_t = self._parse_inline_transform(inner)
                t.maplet_internals.append(inner_t)

        # ── Derive external ports if not explicitly declared ───────
        # Collect all ports from internal transforms
        all_internal_ports = []
        for it in t.maplet_internals:
            all_internal_ports.extend(it.ports)

        # If no explicit MAPPLETINPUT/OUTPUT, derive from internal ports
        if not t.maplet_input_ports:
            for p in all_internal_ports:
                if "INPUT" in p.port_type:
                    t.maplet_input_ports.append(
                        MapletPort(p.name, "INPUT"))

        if not t.maplet_output_ports:
            for p in all_internal_ports:
                if "OUTPUT" in p.port_type:
                    t.maplet_output_ports.append(
                        MapletPort(p.name, "OUTPUT"))

        # ── Build combined ports list (external-facing) ────────────
        # These are the ports that CONNECTORs attach to from outside
        seen_ports = set()
        for mp in t.maplet_input_ports:
            if mp.name not in seen_ports:
                seen_ports.add(mp.name)
                # Find matching internal port for datatype
                matching = next(
                    (p for it in t.maplet_internals
                     for p in it.ports
                     if p.name == mp.name), None)
                t.ports.append(TransformPort(
                    name       = mp.name,
                    port_type  = "INPUT",
                    datatype   = matching.datatype if matching else "string",
                    precision  = matching.precision if matching else "",
                    scale      = matching.scale if matching else "0",
                    expression = ""
                ))

        for mp in t.maplet_output_ports:
            if mp.name not in seen_ports:
                seen_ports.add(mp.name)
                matching = next(
                    (p for it in t.maplet_internals
                     for p in it.ports
                     if p.name == mp.name), None)
                expr = matching.expression if matching else ""
                t.ports.append(TransformPort(
                    name       = mp.name,
                    port_type  = "OUTPUT",
                    datatype   = matching.datatype if matching else "string",
                    precision  = matching.precision if matching else "",
                    scale      = matching.scale if matching else "0",
                    expression = expr
                ))

        return t

    def _parse_inline_transform(self, el) -> Transformation:
        """Parse a <TRANSFORMATION> element into a Transformation object."""
        t = Transformation(
            name       = el.get("NAME",""),
            trans_type = el.get("TYPE",""),
        )
        for tf in el.findall("TRANSFORMFIELD"):
            t.ports.append(TransformPort(
                name       = tf.get("NAME",""),
                port_type  = tf.get("PORTTYPE",""),
                datatype   = tf.get("DATATYPE",""),
                precision  = tf.get("PRECISION",""),
                scale      = tf.get("SCALE","0"),
                expression = self.resolver.resolve(
                    tf.get("EXPRESSION","")),
            ))
        for ta in el.findall("TABLEATTRIBUTE"):
            k = ta.get("NAME","")
            v = self.resolver.resolve(ta.get("VALUE",""))
            t.attributes[k] = v
        return t

    # ── Mappings (with inline maplet detection) ───────────────────

    def _parse_mappings(self, parent) -> None:
        for mp in parent.findall(".//MAPPING"):
            m = Mapping(
                name        = mp.get("NAME","UNKNOWN"),
                description = mp.get("DESCRIPTION",""),
            )

            # Source / target instances
            for inst in mp.findall("INSTANCE"):
                itype = inst.get("TYPE","")
                iname = inst.get("TRANSFORMATION_NAME",
                                 inst.get("NAME",""))
                if itype == "SOURCE" and iname in self._g_sources:
                    m.sources[iname]        = self._g_sources[iname]
                    m.source_schemas[iname] = self._src_owner.get(iname,"")
                elif itype == "TARGET" and iname in self._g_targets:
                    m.targets[iname]        = self._g_targets[iname]
                    m.target_schemas[iname] = self._tgt_owner.get(iname,"")

            # ── TRANSFORMATION tags inside MAPPING ────────────────
            for trans in mp.findall("TRANSFORMATION"):
                trans_type = trans.get("TYPE","")
                trans_name = trans.get("NAME","")

                if trans_type == "Maplet":
                    # Inline maplet definition inside mapping
                    # Build it using the maplet builder
                    inline_ml = self._build_maplet_from_inline_trans(
                        trans)
                    m.transformations[trans_name] = inline_ml
                    # Register globally so INSTANCE refs can find it
                    if trans_name not in self.maplets:
                        self.maplets[trans_name] = inline_ml
                    if trans_name not in m.maplet_instances:
                        m.maplet_instances.append(trans_name)
                    print(f"  [MLET] Inline in {m.name}: {trans_name}")
                else:
                    t = self._parse_inline_transform(trans)
                    m.transformations[trans_name] = t

            # ── INSTANCE tags — resolve maplet references ──────────
            for inst in mp.findall("INSTANCE"):
                itype     = inst.get("TRANSFORMATION_TYPE","")
                inst_name = inst.get("NAME","")
                ref_name  = inst.get("TRANSFORMATION_NAME", inst_name)

                if itype == "Maplet":
                    # Resolve from global registry
                    resolved = self.maplets.get(ref_name)
                    if resolved is None:
                        print(f"  [WARN] Maplet '{ref_name}' referenced "
                              f"in {m.name} but not found in registry")
                        continue

                    # Inject under the INSTANCE NAME (may differ)
                    # Create a copy so rename doesn't pollute registry
                    injected = Transformation(
                        name              = inst_name,
                        trans_type        = "Maplet",
                        is_maplet         = True,
                        ports             = resolved.ports,
                        attributes        = {
                            **resolved.attributes,
                            "maplet_definition": ref_name,
                        },
                        maplet_input_ports = resolved.maplet_input_ports,
                        maplet_output_ports= resolved.maplet_output_ports,
                        maplet_internals   = resolved.maplet_internals,
                    )
                    if inst_name not in m.transformations:
                        m.transformations[inst_name] = injected

                    if inst_name not in m.maplet_instances:
                        m.maplet_instances.append(inst_name)

            # ── CONNECTOR tags ─────────────────────────────────────
            for conn in mp.findall("CONNECTOR"):
                m.connectors.append(Connector(
                    from_instance = conn.get("FROMINSTANCE",""),
                    from_field    = conn.get("FROMFIELD",""),
                    to_instance   = conn.get("TOINSTANCE",""),
                    to_field      = conn.get("TOFIELD",""),
                ))

            self.mappings[m.name] = m
            print(f"  [MAP ] {m.name}: "
                  f"src={list(m.sources)}  "
                  f"tgt={list(m.targets)}  "
                  f"transforms={len(m.transformations)}  "
                  f"connectors={len(m.connectors)}  "
                  f"maplets={m.maplet_instances}")

    def _build_maplet_from_inline_trans(self, el) -> Transformation:
        """
        Handle: <TRANSFORMATION TYPE="Maplet" NAME="ep_CRC_GEN"
                                 REUSABLE="NO">
          <TRANSFORMFIELD .../>  ← ports are directly on the element
        </TRANSFORMATION>
        This is an inline definition where ports are TRANSFORMFIELD
        children (not wrapped in an inner TRANSFORMATION).
        """
        t = Transformation(
            name       = el.get("NAME",""),
            trans_type = "Maplet",
            is_maplet  = True,
            attributes = {"description": el.get("DESCRIPTION","")}
        )
        # Ports are TRANSFORMFIELD children of this element
        for tf in el.findall("TRANSFORMFIELD"):
            port_type  = tf.get("PORTTYPE","")
            expression = self.resolver.resolve(tf.get("EXPRESSION",""))
            port = TransformPort(
                name       = tf.get("NAME",""),
                port_type  = port_type,
                datatype   = tf.get("DATATYPE",""),
                precision  = tf.get("PRECISION",""),
                scale      = tf.get("SCALE","0"),
                expression = expression,
            )
            t.ports.append(port)
            # Classify as input/output port
            if "INPUT" in port_type:
                t.maplet_input_ports.append(
                    MapletPort(port.name, "INPUT"))
            if "OUTPUT" in port_type:
                t.maplet_output_ports.append(
                    MapletPort(port.name, "OUTPUT"))
                # Create an internal "pseudo-transform" so internals are visible
                internal = Transformation(
                    name       = f"{t.name}_EXPR",
                    trans_type = "Expression",
                    ports      = [port],
                )
                if not t.maplet_internals:
                    t.maplet_internals.append(internal)

        # TABLEATTRIBUTE
        for ta in el.findall("TABLEATTRIBUTE"):
            k = ta.get("NAME","")
            v = self.resolver.resolve(ta.get("VALUE",""))
            t.attributes[k] = v

        return t


# ══════════════════════════════════════════════════════════════════════
# MODULE 3 — LINEAGE GRAPH
# ══════════════════════════════════════════════════════════════════════

class LineageGraph:
    """
    Builds a directed graph from CONNECTOR tags.

    Maplet tracing:
    ---------------
    CONNECTORs connect TO/FROM the maplet's INSTANCE NAME using the
    maplet's external port names (TRANSFORMFIELD NAME or MPORTFIELD
    NAME).  The graph traces through the maplet as a single node —
    i.e.  EXP_DERIVE.BAL_USD_AMT → ep_CRC_GEN.IN_BAL_AMT
          ep_CRC_GEN.OUT_CRC_VAL → FIL_ACTIVE.CRC_VAL

    When building logic for a maplet port, we look up the internal
    expression from maplet_internals to show the actual formula.
    """

    def __init__(self, mapping: Mapping):
        self.mapping = mapping
        self.G  = nx.DiGraph()   # port-level  node = "INSTANCE.FIELD"
        self.IG = nx.DiGraph()   # instance-level
        self._build()

    def _build(self) -> None:
        for conn in self.mapping.connectors:
            src = f"{conn.from_instance}.{conn.from_field}"
            tgt = f"{conn.to_instance}.{conn.to_field}"
            self.G.add_edge(src, tgt,
                from_instance = conn.from_instance,
                from_field    = conn.from_field,
                to_instance   = conn.to_instance,
                to_field      = conn.to_field)
            if not self.IG.has_edge(conn.from_instance,
                                    conn.to_instance):
                self.IG.add_edge(conn.from_instance,
                                 conn.to_instance)

    def annotate_exec_order(self) -> None:
        """Topological sort — detects parallel branches and merges."""
        try:
            topo = list(nx.topological_sort(self.IG))
        except nx.NetworkXUnfeasible:
            topo = list(self.IG.nodes)

        ts    = self.mapping.transformations
        split = {n for n in self.IG.nodes
                 if self.IG.out_degree(n) > 1}
        merge = {n for n in self.IG.nodes
                 if self.IG.in_degree(n) > 1}

        level = 0
        pl    = None
        bc:   dict = {}
        ip:   set  = set()

        for inst in topo:
            if inst in split:
                level += 1; pl = level + 1; bc[pl] = 0
                _o, _m = str(level), "Sequential — Split Point"
            elif inst in merge:
                ip.discard(inst); pl = None; level += 1
                _o, _m = str(level), "Sequential — Merge Point"
            elif pl is not None:
                ip.add(inst)
                bc[pl] = bc.get(pl, 0) + 1
                let = chr(96 + bc[pl])
                _o  = f"{pl}{let}"
                _m  = f"Parallel Branch {let.upper()}"
            else:
                level += 1
                _o, _m = str(level), "Sequential"

            if inst in ts:
                t = ts[inst]
                t.exec_order = _o
                t.exec_mode  = _m
                if t.trans_type in ("Aggregator", "Sorter"):
                    t.exec_mode += " [BARRIER]"
                if t.is_maplet:
                    t.exec_mode += " [MAPLET]"

    def topo_order(self) -> list:
        try:
            return list(nx.topological_sort(self.IG))
        except Exception:
            return list(self.IG.nodes)

    def trace_back(self, tgt_inst: str, tgt_col: str) -> list:
        """BFS backward from target port — returns edge dicts."""
        node = f"{tgt_inst}.{tgt_col}"
        if node not in self.G:
            return []
        edges, q, vis = [], [node], set()
        while q:
            n = q.pop()
            if n in vis:
                continue
            vis.add(n)
            for pred in self.G.predecessors(n):
                edges.append(dict(self.G.edges[pred, n]))
                q.append(pred)
        return edges

    def build_logic(self, t: Transformation, port: str) -> str:
        """
        Build a human-readable logic string for a port.
        For maplets: look into maplet_internals for the expression.
        """
        tt = t.trans_type

        # ── Maplet: look inside internal transforms ───────────────
        if tt == "Maplet" or t.is_maplet:
            # Search internal transforms for the port
            for internal in t.maplet_internals:
                for p in internal.ports:
                    if p.name == port and p.expression:
                        defn = t.attributes.get(
                            "maplet_definition", t.name)
                        return (f"MAPLET [{defn}] → "
                                f"{internal.trans_type} "
                                f"{internal.name}: {p.expression}")
            # Fallback: check ports directly on the maplet object
            for p in t.ports:
                if p.name == port and p.expression:
                    defn = t.attributes.get(
                        "maplet_definition", t.name)
                    return f"MAPLET [{defn}]: {p.expression}"
            defn = t.attributes.get("maplet_definition", t.name)
            return f"MAPLET [{defn}]: Pass-through"

        # ── Standard transformations ──────────────────────────────
        if tt == "Expression":
            for p in t.ports:
                if p.name == port and p.expression:
                    return p.expression
            return "Pass-through"
        elif tt == "Lookup Procedure":
            table = t.attributes.get("Lookup table name","")
            cond  = t.attributes.get("Lookup condition","")
            cache = t.attributes.get("Lookup cache persistent","NO")
            return (f"LOOKUP: {table}  ON ({cond})  "
                    f"RETURN {port}  [Cache={cache}]")
        elif tt == "Filter":
            return "FILTER: " + t.attributes.get(
                "Filter Condition","")
        elif tt == "Update Strategy":
            return "UPD_STRATEGY: " + t.attributes.get(
                "Update Strategy Expression","")
        elif tt == "Source Qualifier":
            sql = t.attributes.get("Sql Query","")
            flt = t.attributes.get("Source Filter","")
            return (f"SQ_SQL: {sql}" if sql
                    else f"SQ_FILTER: {flt}" if flt
                    else "Pass-through")
        elif tt == "Router":
            return "ROUTER: " + "; ".join(
                f"{k}={v}" for k, v in t.attributes.items()
                if "Group" in k)
        elif tt == "Aggregator":
            for p in t.ports:
                if p.name == port and p.expression:
                    return f"AGG: {p.expression}"
            return "Pass-through"
        elif tt == "Joiner":
            return (f"JOINER condition="
                    f"{t.attributes.get('Join Condition','')} "
                    f"type={t.attributes.get('Join Type','')}")
        elif tt == "Sequence":
            return (f"SEQUENCE start="
                    f"{t.attributes.get('Start Value','1')} "
                    f"incr={t.attributes.get('Increment By','1')}")
        elif tt == "Normalizer":
            return "NORMALIZER"
        elif tt == "Stored Procedure":
            return "SP: " + t.attributes.get(
                "Stored Procedure Name","")
        return "Pass-through"


# ══════════════════════════════════════════════════════════════════════
# MODULE 4 — EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════

class MappingExcelWriter:

    C = {
        "navy":    "0D1B2A",
        "teal":    "065A82",
        "purple":  "4A235A",
        "dark":    "2C3E50",
        "brown":   "7B341E",
        "maplet":  "7B2D8B",   # distinct maplet colour
        "sub":     "1A3A5C",
        "gold":    "F4A621",
        "green":   "22C55E",
        "cyan":    "00B4D8",
        "coral":   "E8593C",
        "white":   "FFFFFF",
        "alt_a":   "E8F4FD",
        "alt_b":   "E8FDF4",
        "alt_c":   "F5EEF8",
        "alt_d":   "FEF9E7",
        "alt_maplet": "F9EAF9",  # light purple for maplet rows
    }
    BORD  = Border(**{s: Side(style="thin", color="CCCCCC")
                      for s in ("left","right","top","bottom")})
    HDR_F = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
    BOD_F = Font(name="Calibri", bold=False, color="000000", size=10)
    BLD_F = Font(name="Calibri", bold=True,  color="000000", size=10)

    def __init__(self, output_path: str):
        self.out = output_path
        self.wb  = Workbook()
        self.wb.remove(self.wb.active)

    def _fill(self, color_key: str) -> PatternFill:
        hex_c = self.C.get(color_key, color_key)
        hex_c = hex_c.lstrip("#")
        if len(hex_c) not in (6, 8):
            hex_c = "FFFFFF"
        return PatternFill("solid", fgColor=hex_c)

    def _ws(self, title: str):
        ws = self.wb.create_sheet(title=title[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def _hc(self, ws, r, c, v, w=None, clr="navy"):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.HDR_F
        cell.fill      = self._fill(clr)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center",
                                   wrap_text=True)
        cell.border    = self.BORD
        if w:
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[r].height = 30
        return cell

    def _dc(self, ws, r, c, v, shade=False, wrap=False,
            align="left", clr=None, bold=False):
        cell = ws.cell(row=r, column=c,
                       value=str(v) if v is not None else "")
        cell.font      = self.BLD_F if bold else self.BOD_F
        cell.fill      = (self._fill(clr) if clr
                          else self._fill("alt_a") if shade
                          else self._fill("white"))
        cell.alignment = Alignment(horizontal=align,
                                   vertical="top" if wrap else "center",
                                   wrap_text=wrap)
        cell.border    = self.BORD
        ws.row_dimensions[r].height = 38 if wrap else 22
        return cell

    def _banner(self, ws, row, text, clr="navy", cols=14):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Calibri", bold=True,
                           color="FFFFFF", size=11)
        c.fill      = self._fill(clr)
        c.alignment = Alignment(horizontal="left",
                                vertical="center")
        c.border    = self.BORD
        ws.row_dimensions[row].height = 26
        return row + 1

    def _ch(self, ws, row, hdrs, wids, clr="sub"):
        for c, (h, w) in enumerate(zip(hdrs, wids), 1):
            self._hc(ws, row, c, h, w, clr)
        return row + 1

    def _eo_key(self, t: Transformation):
        eo = t.exec_order or "z99"
        l  = ord(eo[0]) - 96 if eo[0].isalpha() else 0
        n  = int(re.sub(r'[^0-9]', '', eo) or "0")
        return (l, n)

    # ── Maplet colour helper ──────────────────────────────────────

    def _maplet_fill(self, shade: bool) -> str:
        return "alt_maplet" if shade else "F9F0FA"

    # ─────────────────────────────────────────────────────────────
    # SHEET A: Full Mapping Parse (5 sections)
    # ─────────────────────────────────────────────────────────────

    def sheet_mapping_parse(self, mapping: Mapping,
                             graph: LineageGraph,
                             sheet_name: str):
        ws = self._ws(sheet_name)
        row = 1

        # ── SECTION A: SOURCE ─────────────────────────────────────
        row = self._banner(ws, row,
            f"  A — SOURCE DETAILS     Mapping: {mapping.name}",
            "navy")
        row = self._ch(ws, row,
            ["Source Schema","Source Table","Source Column",
             "Datatype","Precision","Nullable",
             "SQ Name","SQ Column","SQL Override / Filter"],
            [22,25,25,14,10,10,25,25,65], "sub")

        for src_name, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src_name,"")
            sq_name, sq_sql, sq_map = "-","-",{}
            for conn in mapping.connectors:
                if conn.from_instance == src_name:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_map[conn.from_field] = conn.to_field
                        sql = t.attributes.get("Sql Query","")
                        flt = t.attributes.get("Source Filter","")
                        sq_sql = sql or flt or "Default"
            for i, col in enumerate(cols):
                sq_c = sq_map.get(col.name,"-")
                for c, v in enumerate([
                    schema, src_name, col.name, col.datatype,
                    col.precision, col.nullable,
                    sq_name, sq_c, sq_sql
                ], 1):
                    self._dc(ws, row, c, v, shade=(i%2==0),
                             wrap=(c==9))
                row += 1
        row += 1

        # ── SECTION B: TARGET ─────────────────────────────────────
        row = self._banner(ws, row, "  B — TARGET DETAILS", "teal")
        row = self._ch(ws, row,
            ["Target Schema","Target Table","Target Column",
             "Datatype","Precision","Scale","Nullable","Key Type"],
            [22,25,25,14,10,8,10,16], "teal")

        for tgt_name, cols in mapping.targets.items():
            schema = mapping.target_schemas.get(tgt_name,"")
            for i, col in enumerate(cols):
                pk = col.key_type == "PRIMARY KEY"
                for c, v in enumerate([
                    schema, tgt_name, col.name, col.datatype,
                    col.precision, col.scale, col.nullable, col.key_type
                ], 1):
                    self._dc(ws, row, c, v, shade=(i%2==0),
                             clr="gold" if pk else None)
                row += 1
        row += 1

        # ── SECTION C: TRANSFORMATION INVENTORY ──────────────────
        row = self._banner(ws, row,
            "  C — TRANSFORMATION INVENTORY  "
            "(with full Maplet expansion)", "purple")
        row = self._ch(ws, row,
            ["Exec Order","Exec Mode","Transform Name","Transform Type",
             "Is Maplet","Maplet Definition","Port Name","Port Type",
             "Datatype","Expression / Logic","Attributes"],
            [12,32,25,20,9,22,25,14,12,55,55], "purple")

        for t in sorted(mapping.transformations.values(),
                        key=self._eo_key):
            is_maplet = t.is_maplet or t.trans_type == "Maplet"
            attr_str  = " | ".join(
                f"{k}: {v}" for k, v in t.attributes.items()
                if v and k not in ("description","maplet_definition"))
            maplet_def = (t.attributes.get("maplet_definition","")
                          if is_maplet else "")

            em_c = ("gold"        if "Parallel" in (t.exec_mode or "")
                    else "green"  if "BARRIER"  in (t.exec_mode or "")
                    else "cyan"   if "Merge"    in (t.exec_mode or "")
                    else "coral"  if is_maplet
                    else None)

            # Show external ports + internal expressions
            display_ports = list(t.ports)
            # If maplet and has internals, also show internal expressions
            if is_maplet and t.maplet_internals:
                for internal in t.maplet_internals:
                    for p in internal.ports:
                        if p.expression and not any(
                                dp.name == p.name
                                for dp in display_ports):
                            display_ports.append(p)

            if not display_ports:
                display_ports = [TransformPort(
                    "-","-","-","-","0")]

            for pi, p in enumerate(display_ports):
                shade = (pi % 2 == 0)
                mfill = (self._maplet_fill(shade)
                         if is_maplet else None)
                for c, v in enumerate([
                    t.exec_order  if pi==0 else "",
                    t.exec_mode   if pi==0 else "",
                    t.name        if pi==0 else "",
                    t.trans_type  if pi==0 else "",
                    "YES"         if (pi==0 and is_maplet) else
                    ("" if pi==0 else ""),
                    maplet_def    if pi==0 else "",
                    p.name, p.port_type, p.datatype,
                    p.expression or "Pass-through",
                    attr_str      if pi==0 else "",
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             wrap=(c in (10,11)),
                             clr=(em_c if c==2 and pi==0 else
                                  mfill if is_maplet else None))
                row += 1
        row += 1

        # ── SECTION D: COLUMN FLOW MAP ────────────────────────────
        row = self._banner(ws, row,
            "  D — COLUMN FLOW MAP  "
            "[ Source → SQ → Transforms (incl. Maplets) → Target ]",
            "dark")
        row = self._ch(ws, row,
            ["#","Source Schema","Source Table","Source Column",
             "SQ Name","SQ Column",
             "Transformation Chain\n(Name|Type|Order|Maplet?)",
             "Logic at Each Step\n(expression/condition/maplet formula)",
             "Target Schema","Target Table",
             "Target Column","Target DT","Key Type","Remarks"],
            [5,18,20,20,20,20,55,65,18,20,20,13,13,14], "dark")

        seq = 1
        for tgt_name, tgt_cols in mapping.targets.items():
            tgt_schema = mapping.target_schemas.get(tgt_name,"")
            for col in tgt_cols:
                edges = graph.trace_back(tgt_name, col.name)
                if not edges:
                    for c, v in enumerate([
                        str(seq),"-","-","-","-","-",
                        "UNCONNECTED","-",
                        tgt_schema, tgt_name, col.name,
                        col.datatype, col.key_type,
                        "No upstream connection"
                    ], 1):
                        self._dc(ws, row, c, v)
                    ws.row_dimensions[row].height = 22
                    row += 1; seq += 1
                    continue

                src_table = src_col = src_schema = "-"
                sq_name   = sq_col  = "-"
                chain, logic = [], []
                has_maplet   = False

                def inst_ord(inst):
                    t = mapping.transformations.get(inst)
                    if t and t.exec_order:
                        eo = t.exec_order
                        l  = ord(eo[0])-96 if eo[0].isalpha() else 0
                        n  = int(re.sub(r'[^0-9]','',eo) or "0")
                        return (l, n)
                    return (0,0) if inst in mapping.sources else (99,99)

                seen, ordered = set(), []
                for e in edges:
                    for inst in [e["from_instance"],
                                 e["to_instance"]]:
                        if inst not in seen:
                            seen.add(inst); ordered.append(inst)
                ordered.sort(key=inst_ord)

                for inst in ordered:
                    t = mapping.transformations.get(inst)
                    field = "-"
                    for e in edges:
                        if e["from_instance"] == inst:
                            field = e["from_field"]; break
                        if e["to_instance"] == inst:
                            field = e["to_field"]

                    if inst in mapping.sources:
                        src_table  = inst; src_col = field
                        src_schema = mapping.source_schemas.get(
                            inst,"")
                    elif t and t.trans_type == "Source Qualifier":
                        sq_name = inst; sq_col = field
                    elif inst not in mapping.targets and t:
                        is_ml = t.is_maplet or t.trans_type=="Maplet"
                        if is_ml:
                            has_maplet = True
                        tag = " [MAPLET]" if is_ml else ""
                        chain.append(
                            f"{inst}  [{t.trans_type}]{tag}  "
                            f"Ord:{t.exec_order}  {t.exec_mode}")
                        lg = graph.build_logic(t, field)
                        if lg and lg != "-":
                            logic.append(f"▶ {inst}: {lg}")

                tc  = "\n→ ".join(chain) if chain else "Direct"
                lc  = "\n".join(logic)   if logic else "Pass-through"
                shade = (seq % 2 == 0)
                lns   = max(len(chain), len(logic), 1)

                # Highlight rows that pass through a maplet
                row_clr = "alt_maplet" if (has_maplet and shade) else \
                          ("F9F0FA" if has_maplet else None)

                for c, v in enumerate([
                    str(seq),
                    src_schema, src_table, src_col,
                    sq_name, sq_col, tc, lc,
                    tgt_schema, tgt_name,
                    col.name, col.datatype, col.key_type,
                    "PK" if col.key_type=="PRIMARY KEY" else
                    ("via MAPLET" if has_maplet else "")
                ], 1):
                    cell = ws.cell(row=row, column=c,
                                   value=str(v) if v else "")
                    cell.font = self.BOD_F
                    cell.fill = (self._fill(row_clr)
                                 if row_clr and c == 7
                                 else self._fill("alt_d") if shade
                                 else self._fill("white"))
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top", wrap_text=True)
                    cell.border = self.BORD
                ws.row_dimensions[row].height = max(38, lns*18)
                row += 1; seq += 1
        row += 1

        # ── SECTION E: FLOW DIAGRAM ROADMAP ──────────────────────
        row = self._banner(ws, row,
            "  E — FLOW DIAGRAM ROADMAP  "
            "(incl. maplet swim-lane guidance)",
            "brown")
        for line in self._roadmap(mapping, graph):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=14)
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = Font(name="Consolas", size=9,
                             bold=("STEP" in line or
                                   "MAPLET" in line),
                             color=("FFFFFF"
                                    if line.startswith("  STEP")
                                    else "4A235A"
                                    if "MAPLET" in line
                                    else "1A1A1A"))
            cell.fill = (self._fill("sub")
                         if line.startswith("  STEP")
                         else self._fill("alt_maplet")
                         if "MAPLET" in line
                         else self._fill("alt_a")
                         if "→" in line
                         else self._fill("white"))
            cell.alignment = Alignment(horizontal="left",
                                        vertical="center")
            cell.border = self.BORD
            ws.row_dimensions[row].height = 15
            row += 1

        ws.freeze_panes = "A4"

    def _roadmap(self, mapping: Mapping,
                  graph: LineageGraph) -> list:
        lines = []; W = 110
        lines.append("═" * W)
        lines.append(
            f"  FLOW DIAGRAM ROADMAP   Mapping: {mapping.name}")
        lines.append("═" * W); lines.append("")

        # Sources
        lines.append("  STEP 1 — SOURCES")
        lines.append("  " + "─"*60)
        for src, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src,"?")
            sq_name, sq_sql = "-", "-"
            for conn in mapping.connectors:
                if conn.from_instance == src:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_sql  = t.attributes.get("Sql Query","")[:90]
                        break
            lines.append(
                f"   {schema}.{src}  ({len(cols)} cols)  SQ:{sq_name}")
            if sq_sql:
                lines.append(f"   SQL: {sq_sql}")
        lines.append("")

        # Transformation chain with maplet expansion
        lines.append("  STEP 2 — TRANSFORMATION CHAIN  (with Maplet expansion)")
        lines.append("  " + "─"*60)
        for t in sorted(mapping.transformations.values(),
                        key=self._eo_key):
            is_ml  = t.is_maplet or t.trans_type == "Maplet"
            split_ = graph.IG.out_degree(t.name) > 1
            merge_ = graph.IG.in_degree(t.name) > 1
            flag   = (" ⚡SPLIT" if split_ else "") + \
                     (" 🔀MERGE" if merge_ else "") + \
                     (" 🛑BARRIER"
                      if t.trans_type in ("Aggregator","Sorter")
                      else "")
            maplet_tag = "  ◈ MAPLET" if is_ml else ""
            lines.append(
                f"   [{t.exec_order:>5}]  {t.name:<35}  "
                f"({t.trans_type}){maplet_tag}  "
                f"{t.exec_mode}{flag}")

            if is_ml:
                defn = t.attributes.get("maplet_definition", t.name)
                lines.append(
                    f"             ◈ Maplet definition: {defn}")
                # Show input and output ports
                in_ports  = [p.name for p in t.ports
                             if "INPUT" in p.port_type]
                out_ports = [p.name for p in t.ports
                             if "OUTPUT" in p.port_type]
                lines.append(
                    f"             ◈ Input  ports : "
                    f"{', '.join(in_ports) or 'none'}")
                lines.append(
                    f"             ◈ Output ports : "
                    f"{', '.join(out_ports) or 'none'}")
                # Show internal expressions
                for internal in t.maplet_internals:
                    for p in internal.ports:
                        if p.expression:
                            lines.append(
                                f"             ◈ INTERNAL "
                                f"{internal.trans_type} "
                                f"{internal.name}:"
                                f"  {p.name} = {p.expression[:70]}")
            elif t.trans_type == "Expression":
                for p in t.ports:
                    if p.expression and "OUTPUT" in p.port_type:
                        lines.append(
                            f"             ▶ {p.name} = "
                            f"{p.expression[:80]}")
            elif t.trans_type == "Lookup Procedure":
                lines.append(
                    f"             ▶ LOOKUP: "
                    f"{t.attributes.get('Lookup table name','')}  "
                    f"ON {t.attributes.get('Lookup condition','')}")
            elif t.trans_type == "Filter":
                lines.append(
                    f"             ▶ FILTER: "
                    f"{t.attributes.get('Filter Condition','')}")
        lines.append("")

        # Targets
        lines.append("  STEP 3 — TARGETS")
        lines.append("  " + "─"*60)
        for tgt, cols in mapping.targets.items():
            schema  = mapping.target_schemas.get(tgt,"?")
            pk_cols = [c.name for c in cols
                       if c.key_type == "PRIMARY KEY"]
            lines.append(
                f"   {schema}.{tgt}  ({len(cols)} cols)"
                f"  PK:{', '.join(pk_cols)}")
        lines.append("")

        # Additional steps
        lines.append("  ADDITIONAL STEPS — VISUAL FLOW DIAGRAM")
        lines.append("  " + "─"*60)
        add = [
            ("STEP A","Graph edges from Section D",
             "Each row = SQ_Col → Transform1 → [MAPLET] → Transform2 → Target"),
            ("STEP B","Node colours",
             "SOURCE=blue  SQ=teal  EXPR=orange  LKP=purple  "
             "MAPLET=coral  FILTER=red  TARGET=green"),
            ("STEP C","Maplet swim-lane",
             "Draw each maplet as a sub-diagram box. Show its input ports "
             "on left, output ports on right, internal transforms inside."),
            ("STEP D","Parallel branches",
             "Group by PIPELINE from session sheet → swim-lanes"),
            ("STEP E","Tools",
             "Python: graphviz / networkx  |  Web: D3.js  |  "
             "Manual: draw.io (import from Section D as CSV)"),
        ]
        for sid, title, detail in add:
            lines.append(f"   {sid}  {title}")
            lines.append(f"         • {detail}")
        lines.append(""); lines.append("═"*W)
        return lines

    # ─────────────────────────────────────────────────────────────
    # SHEET B: Source Detail
    # ─────────────────────────────────────────────────────────────

    def sheet_source_detail(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Source Schema","Source Table","Source Column",
                "Datatype","Precision","Scale","Nullable",
                "SQ Name","SQ Column","SQL Override / Filter"]
        wids = [22,25,25,14,10,8,10,25,25,70]
        self._ch(ws, 1, hdrs, wids, "navy")
        row = 2
        for src_name, cols in mapping.sources.items():
            schema = mapping.source_schemas.get(src_name,"")
            sq_name, sq_sql, sq_map = "-","-",{}
            for conn in mapping.connectors:
                if conn.from_instance == src_name:
                    t = mapping.transformations.get(conn.to_instance)
                    if t and t.trans_type == "Source Qualifier":
                        sq_name = t.name
                        sq_map[conn.from_field] = conn.to_field
                        sql = t.attributes.get("Sql Query","")
                        flt = t.attributes.get("Source Filter","")
                        sq_sql = sql or flt or "Default"
            for i, col in enumerate(cols):
                sq_c = sq_map.get(col.name,"-")
                for c, v in enumerate([
                    schema, src_name, col.name, col.datatype,
                    col.precision, col.scale, col.nullable,
                    sq_name, sq_c, sq_sql
                ], 1):
                    self._dc(ws, row, c, v, shade=(i%2==0),
                             wrap=(c==10))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET C: Target Detail
    # ─────────────────────────────────────────────────────────────

    def sheet_target_detail(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Target Schema","Target Table","Target Column",
                "Datatype","Precision","Scale","Nullable","Key Type"]
        wids = [22,25,25,14,10,8,10,16]
        self._ch(ws, 1, hdrs, wids, "teal")
        row = 2
        for tgt_name, cols in mapping.targets.items():
            schema = mapping.target_schemas.get(tgt_name,"")
            for i, col in enumerate(cols):
                pk = col.key_type == "PRIMARY KEY"
                for c, v in enumerate([
                    schema, tgt_name, col.name, col.datatype,
                    col.precision, col.scale, col.nullable, col.key_type
                ], 1):
                    self._dc(ws, row, c, v, shade=(i%2==0),
                             clr="gold" if pk else None)
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET D: Transformation Detail (with maplet expansion)
    # ─────────────────────────────────────────────────────────────

    def sheet_transforms(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Exec Order","Exec Mode","Transform Name",
                "Transform Type","Is Maplet","Maplet Definition",
                "Port Name","Port Type","Datatype",
                "Expression / Logic","Attributes"]
        wids = [12,32,25,20,9,22,25,14,12,58,55]
        self._ch(ws, 1, hdrs, wids, "purple")
        row = 2

        for t in sorted(mapping.transformations.values(),
                        key=self._eo_key):
            is_ml    = t.is_maplet or t.trans_type == "Maplet"
            attr_str = " | ".join(
                f"{k}: {v}" for k, v in t.attributes.items()
                if v and k not in ("description","maplet_definition"))
            maplet_def = (t.attributes.get("maplet_definition","")
                          if is_ml else "")
            em_c = ("gold"   if "Parallel" in (t.exec_mode or "")
                    else "green" if "BARRIER" in (t.exec_mode or "")
                    else "cyan"  if "Merge"   in (t.exec_mode or "")
                    else "coral" if is_ml else None)

            # Show all ports including internal maplet expressions
            display_ports = list(t.ports)
            if is_ml:
                for internal in t.maplet_internals:
                    for p in internal.ports:
                        if p.expression and not any(
                                dp.name == p.name
                                for dp in display_ports):
                            display_ports.append(p)

            if not display_ports:
                display_ports = [TransformPort("-","-","-","-","0")]

            for pi, p in enumerate(display_ports):
                shade = (pi % 2 == 0)
                for c, v in enumerate([
                    t.exec_order  if pi==0 else "",
                    t.exec_mode   if pi==0 else "",
                    t.name        if pi==0 else "",
                    t.trans_type  if pi==0 else "",
                    "YES"         if pi==0 and is_ml else
                    ("" if pi==0 else ""),
                    maplet_def    if pi==0 else "",
                    p.name, p.port_type, p.datatype,
                    p.expression or "Pass-through",
                    attr_str      if pi==0 else "",
                ], 1):
                    self._dc(ws, row, c, v, shade=shade,
                             wrap=(c in (10,11)),
                             clr=(em_c if c==2 and pi==0 else
                                  self._maplet_fill(shade)
                                  if is_ml else None))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET E: Column Flow Map (dedicated)
    # ─────────────────────────────────────────────────────────────

    def sheet_column_flow(self, mapping: Mapping,
                           graph: LineageGraph, name: str):
        ws   = self._ws(name)
        hdrs = ["#","Source Schema","Source Table","Source Column",
                "SQ Name","SQ Column",
                "Transformation Chain\n(Name|Type|Order|Maplet?)",
                "Logic at Each Step\n(expr/condition/maplet formula)",
                "Target Schema","Target Table",
                "Target Column","Target DT","Key Type","Remarks"]
        wids = [5,18,22,22,22,22,55,65,18,22,22,14,13,16]
        self._ch(ws, 1, hdrs, wids, "dark")
        row = 2; seq = 1

        for tgt_name, tgt_cols in mapping.targets.items():
            tgt_schema = mapping.target_schemas.get(tgt_name,"")
            for col in tgt_cols:
                edges = graph.trace_back(tgt_name, col.name)
                if not edges:
                    for c, v in enumerate([
                        str(seq),"-","-","-","-","-",
                        "UNCONNECTED","-",
                        tgt_schema, tgt_name, col.name,
                        col.datatype, col.key_type,
                        "No upstream connection"
                    ], 1):
                        self._dc(ws, row, c, v)
                    row += 1; seq += 1
                    continue

                src_table = src_col = src_schema = "-"
                sq_name   = sq_col  = "-"
                chain, logic = [], []
                has_maplet   = False

                def inst_ord(inst):
                    t = mapping.transformations.get(inst)
                    if t and t.exec_order:
                        eo = t.exec_order
                        l  = ord(eo[0])-96 if eo[0].isalpha() else 0
                        n  = int(re.sub(r'[^0-9]','',eo) or "0")
                        return (l,n)
                    return (0,0) if inst in mapping.sources else (99,99)

                seen, ordered = set(), []
                for e in edges:
                    for inst in [e["from_instance"],e["to_instance"]]:
                        if inst not in seen:
                            seen.add(inst); ordered.append(inst)
                ordered.sort(key=inst_ord)

                for inst in ordered:
                    t = mapping.transformations.get(inst)
                    field = "-"
                    for e in edges:
                        if e["from_instance"] == inst:
                            field = e["from_field"]; break
                        if e["to_instance"] == inst:
                            field = e["to_field"]

                    if inst in mapping.sources:
                        src_table  = inst; src_col = field
                        src_schema = mapping.source_schemas.get(inst,"")
                    elif t and t.trans_type == "Source Qualifier":
                        sq_name = inst; sq_col = field
                    elif inst not in mapping.targets and t:
                        is_ml = t.is_maplet or t.trans_type=="Maplet"
                        if is_ml:
                            has_maplet = True
                        tag = "  [MAPLET]" if is_ml else ""
                        chain.append(
                            f"{inst}  [{t.trans_type}]{tag}  "
                            f"Ord:{t.exec_order}  {t.exec_mode}")
                        lg = graph.build_logic(t, field)
                        if lg and lg != "-":
                            logic.append(f"▶ {inst}: {lg}")

                tc  = "\n→ ".join(chain) if chain else "Direct"
                lc  = "\n".join(logic)   if logic else "Pass-through"
                shade = (seq % 2 == 0)
                lns   = max(len(chain), len(logic), 1)

                row_clr = (self._maplet_fill(shade)
                           if has_maplet else None)

                for c, v in enumerate([
                    str(seq),
                    src_schema, src_table, src_col,
                    sq_name, sq_col, tc, lc,
                    tgt_schema, tgt_name,
                    col.name, col.datatype, col.key_type,
                    "PK" if col.key_type=="PRIMARY KEY" else
                    ("via MAPLET" if has_maplet else "")
                ], 1):
                    cell = ws.cell(row=row, column=c,
                                   value=str(v) if v else "")
                    cell.font = self.BOD_F
                    cell.fill = (self._fill(row_clr)
                                 if row_clr
                                 else self._fill("alt_d") if shade
                                 else self._fill("white"))
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top",
                        wrap_text=True)
                    cell.border = self.BORD
                ws.row_dimensions[row].height = max(38, lns*18)
                row += 1; seq += 1
        ws.freeze_panes = "B2"

    # ─────────────────────────────────────────────────────────────
    # SHEET F: Lookups
    # ─────────────────────────────────────────────────────────────

    def sheet_lookups(self, mapping: Mapping, name: str):
        ws   = self._ws(name)
        hdrs = ["Lookup Name","Lookup Table","Lookup Condition",
                "Return Cols","Cache","Exec Order","Exec Mode"]
        wids = [28,32,45,28,14,12,32]
        self._ch(ws, 1, hdrs, wids, "teal")
        row = 2
        for t in mapping.transformations.values():
            if t.trans_type == "Lookup Procedure":
                s    = (row % 2 == 0)
                rets = [p.name for p in t.ports
                        if "OUTPUT" in p.port_type]
                for c, v in enumerate([
                    t.name,
                    t.attributes.get("Lookup table name",""),
                    t.attributes.get("Lookup condition",""),
                    ", ".join(rets),
                    t.attributes.get("Lookup cache persistent","NO"),
                    t.exec_order, t.exec_mode
                ], 1):
                    self._dc(ws, row, c, v, shade=s, wrap=(c==3))
                row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET G: Maplet Detail (expanded — the new key sheet)
    # ─────────────────────────────────────────────────────────────

    def sheet_maplet_detail(self, mapping: Mapping,
                             global_maplets: dict, name: str):
        """
        Full maplet breakdown:
        - External input ports (what CONNECTORs feed in)
        - Internal transformation(s) with expressions
        - External output ports (what CONNECTORs take out)
        - Upstream connectors (what feeds INTO the maplet)
        - Downstream connectors (what the maplet feeds INTO)
        """
        ws = self._ws(name)
        row = 1

        if not mapping.maplet_instances:
            row = self._banner(ws, row,
                "No maplets used in this mapping", "teal")
            ws.freeze_panes = "A2"
            return

        for mname in mapping.maplet_instances:
            t = mapping.transformations.get(mname)
            if not t:
                continue

            defn_name = t.attributes.get("maplet_definition", mname)

            # ── Maplet header banner ──────────────────────────────
            row = self._banner(ws, row,
                f"  MAPLET: {mname}  "
                f"(definition: {defn_name})  "
                f"[{t.exec_order}]  {t.exec_mode}",
                "maplet")

            # ── A: External Input Ports ───────────────────────────
            row = self._banner(ws, row,
                "    Input Ports  (fed by CONNECTORs from upstream)",
                "purple")
            row = self._ch(ws, row,
                ["Port Name","Datatype","Upstream Instance",
                 "Upstream Field","CONNECTOR #"],
                [25,14,30,25,12], "sub")

            # Find all connectors feeding into this maplet instance
            conn_idx = 1
            for conn in mapping.connectors:
                if conn.to_instance == mname:
                    shade = (conn_idx % 2 == 0)
                    for c, v in enumerate([
                        conn.to_field, "-",
                        conn.from_instance, conn.from_field,
                        str(conn_idx)
                    ], 1):
                        # Fill datatype from port definition
                        if c == 2:
                            matching_port = next(
                                (p for p in t.ports
                                 if p.name == conn.to_field), None)
                            v = (matching_port.datatype
                                 if matching_port else "-")
                        self._dc(ws, row, c, v, shade=shade)
                    row += 1; conn_idx += 1
            row += 1

            # ── B: Internal Transformations ───────────────────────
            row = self._banner(ws, row,
                "    Internal Transformations  (logic inside the maplet)",
                "purple")
            row = self._ch(ws, row,
                ["Internal Transform","Type","Port Name",
                 "Port Type","Datatype","Expression (verbatim)"],
                [28,20,25,14,14,65], "sub")

            if t.maplet_internals:
                for internal in t.maplet_internals:
                    ports = internal.ports or [
                        TransformPort("-","-","-","-","0")]
                    for pi, p in enumerate(ports):
                        shade = (pi % 2 == 0)
                        for c, v in enumerate([
                            internal.name if pi==0 else "",
                            internal.trans_type if pi==0 else "",
                            p.name, p.port_type, p.datatype,
                            p.expression or "Pass-through"
                        ], 1):
                            self._dc(ws, row, c, v, shade=shade,
                                     wrap=(c==6))
                        row += 1
            else:
                # Inline maplet — show ports directly
                for pi, p in enumerate(t.ports):
                    shade = (pi % 2 == 0)
                    for c, v in enumerate([
                        mname if pi==0 else "",
                        "Maplet (inline)" if pi==0 else "",
                        p.name, p.port_type, p.datatype,
                        p.expression or "Pass-through"
                    ], 1):
                        self._dc(ws, row, c, v, shade=shade,
                                 wrap=(c==6))
                    row += 1
            row += 1

            # ── C: External Output Ports ──────────────────────────
            row = self._banner(ws, row,
                "    Output Ports  (fed to CONNECTORs downstream)",
                "purple")
            row = self._ch(ws, row,
                ["Port Name","Datatype","Expression (derived)",
                 "Downstream Instance","Downstream Field","CONNECTOR #"],
                [25,14,50,30,25,12], "sub")

            conn_idx = 1
            for conn in mapping.connectors:
                if conn.from_instance == mname:
                    shade = (conn_idx % 2 == 0)
                    # Find expression from internal transforms
                    expr = "-"
                    for internal in t.maplet_internals:
                        for p in internal.ports:
                            if (p.name == conn.from_field
                                    and p.expression):
                                expr = p.expression
                                break
                    if expr == "-":
                        for p in t.ports:
                            if (p.name == conn.from_field
                                    and p.expression):
                                expr = p.expression
                                break
                    matching_port = next(
                        (p for p in t.ports
                         if p.name == conn.from_field), None)
                    dt = (matching_port.datatype
                          if matching_port else "-")
                    for c, v in enumerate([
                        conn.from_field, dt, expr,
                        conn.to_instance, conn.to_field,
                        str(conn_idx)
                    ], 1):
                        self._dc(ws, row, c, v, shade=shade,
                                 wrap=(c==3))
                    row += 1; conn_idx += 1
            row += 2   # spacer between maplets

        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # SHEET H: Execution Order
    # ─────────────────────────────────────────────────────────────

    def sheet_exec_order(self, mapping: Mapping,
                          graph: LineageGraph, name: str):
        ws   = self._ws(name)
        hdrs = ["Exec Order","Instance Name","Transform Type",
                "Is Maplet","Exec Mode","In-Deg","Out-Deg","Notes"]
        wids = [12,35,22,9,38,10,10,50]
        self._ch(ws, 1, hdrs, wids, "navy")
        row = 2
        for inst in graph.topo_order():
            t    = mapping.transformations.get(inst)
            in_d = graph.IG.in_degree(inst)
            out_d= graph.IG.out_degree(inst)
            s    = (row % 2 == 0)
            is_ml = t and (t.is_maplet or t.trans_type=="Maplet")
            note = ""
            if out_d > 1: note = f"⚡ SPLIT → {out_d} branches"
            if in_d  > 1: note = f"🔀 MERGE ← {in_d} branches"
            if t and t.trans_type in ("Aggregator","Sorter"):
                note += " 🛑 BARRIER"
            if is_ml:
                defn = t.attributes.get("maplet_definition", inst)
                note += f" ◈ MAPLET def:{defn}"
            em_c = ("gold"   if t and "Parallel" in (t.exec_mode or "")
                    else "green" if t and "Merge"  in (t.exec_mode or "")
                    else "cyan"  if t and "Split"  in (t.exec_mode or "")
                    else "coral" if is_ml else None)
            for c, v in enumerate([
                t.exec_order if t else "?",
                inst,
                t.trans_type if t else "Source/Target",
                "YES" if is_ml else "NO",
                t.exec_mode  if t else "Sequential",
                str(in_d), str(out_d), note
            ], 1):
                self._dc(ws, row, c, v, shade=s,
                         align="center" if c in (1,4,6,7) else "left",
                         clr=em_c if c==5 else None, wrap=(c==8))
            row += 1
        ws.freeze_panes = "A2"

    # ─────────────────────────────────────────────────────────────
    # GENERATE
    # ─────────────────────────────────────────────────────────────

    def generate(self, parser: MappingParser) -> None:
        print(f"\n[WRITE] {self.out}")
        for m_name, mapping in parser.mappings.items():
            graph = LineageGraph(mapping)
            graph.annotate_exec_order()
            safe  = re.sub(r'[^A-Za-z0-9_]', '_', m_name)[:16]

            self.sheet_mapping_parse(mapping, graph,
                                      f"{safe}_A_MappingParse")
            self.sheet_source_detail (mapping,
                                      f"{safe}_B_Sources")
            self.sheet_target_detail (mapping,
                                      f"{safe}_C_Targets")
            self.sheet_transforms    (mapping,
                                      f"{safe}_D_Transforms")
            self.sheet_column_flow   (mapping, graph,
                                      f"{safe}_E_ColFlow")
            self.sheet_lookups       (mapping,
                                      f"{safe}_F_Lookups")
            self.sheet_maplet_detail (mapping, parser.maplets,
                                      f"{safe}_G_MapletDetail")
            self.sheet_exec_order    (mapping, graph,
                                      f"{safe}_H_ExecOrder")

            print(f"  [OK  ] {m_name}  "
                  f"(maplets: {mapping.maplet_instances})")

        self.wb.save(self.out)
        print(f"\n  ✅  Saved → {Path(self.out).resolve()}")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="LineageIQ Mapping Parser v4 — "
                    "full maplet architecture support",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples
--------
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml --par params_prod.par
  python lineageiq_mapping.py --xml m_TPR_CUST_BAL.xml --par params_prod.par --out MySTTM.xlsx
        """
    )
    ap.add_argument("--xml", required=True)
    ap.add_argument("--par", nargs="*", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    xml_path  = args.xml
    par_files = args.par or []
    out_path  = args.out or \
        f"LineageIQ_Mapping_{Path(xml_path).stem}.xlsx"

    print()
    print("═" * 65)
    print("  LineageIQ — MAPPING PARSER  v4  (full maplet support)")
    print("═" * 65)
    print(f"  XML : {xml_path}")
    print(f"  PAR : {par_files or 'None ($$PARAMS kept as-is)'}")
    print(f"  OUT : {out_path}")
    print("═" * 65)

    resolver = ParamResolver(par_files)
    resolver.report()

    parser = MappingParser(resolver)
    parser.parse(xml_path)

    if not parser.mappings:
        print("\n  [ERROR] No <MAPPING> tags found in XML.")
        return

    writer = MappingExcelWriter(out_path)
    writer.generate(parser)

    print()
    print("═" * 65)
    print(f"  ✅  Done!  → {Path(out_path).resolve()}")
    print("═" * 65)


if __name__ == "__main__":
    main()
